"""Paste IAMActionHunter privilege-escalation CSV exports verbatim into new
sheets of the Cloud VA/VAPT Excel tracker.

IAMActionHunter (https://github.com/RhinoSecurityLabs/IAMActionHunter) produces a
CSV when run with its privilege-escalation config. The consultant uploads that CSV
in the Cloud VA importer; unlike the Steampipe/Prowler scanner CSVs, it is NOT
parsed into findings — its rows are pasted as-is into their own worksheet so the
consultant can reference / triage the raw privilege-escalation paths inside the
same tracker workbook.
"""
from __future__ import annotations

import csv
import io
from pathlib import Path

# Characters Excel forbids in a sheet title, plus the 31-char length cap.
_BAD_SHEET_CHARS = set('[]:*?/\\')
_MAX_SHEET = 31


def _safe_sheet_name(base: str, used: set[str]) -> str:
    """Return an Excel-legal, unique (case-insensitive) sheet name."""
    name = "".join("_" if c in _BAD_SHEET_CHARS else c for c in (base or "Sheet")).strip()
    name = (name or "Sheet")[:_MAX_SHEET]
    lower_used = {u.lower() for u in used}
    cand = name
    i = 2
    while cand.lower() in lower_used:
        suffix = f" ({i})"
        cand = name[: _MAX_SHEET - len(suffix)] + suffix
        i += 1
    used.add(cand)
    return cand


def append_iam_csv_sheets(xlsx_path: Path, csv_entries: list[dict]) -> int:
    """Open the tracker at ``xlsx_path`` and append one new sheet per IAM CSV,
    writing each CSV's rows verbatim. ``csv_entries`` is a list of
    ``{"filename": str, "path": str}`` dicts. Returns the number of sheets added.

    Best-effort: a missing/unreadable CSV is skipped; the workbook is saved only
    if at least one sheet was added.
    """
    import openpyxl

    xlsx_path = Path(xlsx_path)
    if not csv_entries:
        return 0

    wb = openpyxl.load_workbook(str(xlsx_path))
    used = set(wb.sheetnames)
    added = 0

    for entry in csv_entries:
        raw_path = str((entry or {}).get("path") or "")
        if not raw_path:
            continue
        p = Path(raw_path)
        if not p.exists():
            continue
        try:
            text = p.read_bytes().decode("utf-8-sig", errors="replace")
        except Exception:
            continue

        stem = Path((entry or {}).get("filename") or p.name).stem
        base = f"IAM {stem}" if stem else "IAMActionHunter"
        title = _safe_sheet_name(base, used)
        ws = wb.create_sheet(title=title)

        # Sniff the delimiter (IAMActionHunter emits standard commas, but be
        # tolerant of tab/semicolon exports) and write every cell verbatim.
        sample = text[:4096]
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
        except Exception:
            dialect = csv.excel
        reader = csv.reader(io.StringIO(text), dialect)
        for r, row in enumerate(reader, start=1):
            for c, val in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=val)
        added += 1

    if added:
        wb.save(str(xlsx_path))
    return added
