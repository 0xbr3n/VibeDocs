"""Project CRUD."""
import re
import shutil
import tempfile
from datetime import datetime
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified
from ..database import get_db
from ..models import Project, User
from ..schemas import ProjectCreate, ProjectOut
from ..auth import get_current_user
from ..services import scope_import as _scope_import
from ..services.upload_utils import stream_save as _stream_save
import logging as _log_p

_logger = _log_p.getLogger(__name__)

router = APIRouter(prefix="/api/projects", tags=["projects"])

@router.get("", response_model=list[ProjectOut])
def list_projects(db: Session = Depends(get_db), user: User = Depends(get_current_user)):
    # IDOR fix: only return projects the caller has a relationship to.
    # Previously did db.query(Project).all() then called user_can_see_project()
    # per row — that was O(n) queries (N+1). Now we push the SQL-expressible
    # conditions into a single query with subqueries; the JSON-based
    # assigned_user_ids check requires one extra fetch at most.
    from sqlalchemy import or_
    from ..models import Report, ReportAccess, Role

    if user.role in (Role.admin, Role.senior):
        return db.query(Project).order_by(Project.created_at.desc()).all()

    # Subqueries for owned-report and grant-based project IDs
    owned_pids = db.query(Report.project_id).filter(
        Report.created_by_id == user.id
    ).subquery()
    granted_pids = (
        db.query(Report.project_id)
          .join(ReportAccess, ReportAccess.report_id == Report.id)
          .filter(ReportAccess.user_id == user.id)
          .subquery()
    )

    rows = (
        db.query(Project)
          .filter(
              or_(
                  Project.lead_id == user.id,
                  Project.id.in_(owned_pids),
                  Project.id.in_(granted_pids),
              )
          )
          .order_by(Project.created_at.desc())
          .all()
    )

    # JSON-based assignment (project.details["assigned_user_ids"]) can't be
    # expressed portably in SQL. One extra query for the remaining projects.
    seen_ids = {p.id for p in rows}
    remaining = (
        db.query(Project).filter(Project.id.notin_(seen_ids)).all()
        if seen_ids else db.query(Project).all()
    )
    json_rows = [
        p for p in remaining
        if user.id in ((p.details or {}).get("assigned_user_ids") or [])
    ]

    combined = rows + json_rows
    combined.sort(key=lambda p: p.created_at or "", reverse=True)
    return combined

@router.post("", response_model=ProjectOut)
def create_project(
    payload: ProjectCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    p = Project(**payload.model_dump(), lead_id=user.id)
    db.add(p); db.commit(); db.refresh(p)
    return p

@router.get("/{project_id}", response_model=ProjectOut)
def get_project(project_id: int, db: Session = Depends(get_db),
                user: User = Depends(get_current_user)):
    from .permissions import require_project_visibility
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")
    require_project_visibility(db, user, p)
    return p

@router.put("/{project_id}", response_model=ProjectOut)
def update_project(
    project_id: int,
    payload: ProjectCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    from ..models import Role
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")
    # IDOR fix: only the project lead or admin/senior may mutate project
    # metadata. Previously any authenticated user could PUT a project.
    if user.role not in (Role.admin, Role.senior) and p.lead_id != user.id:
        raise HTTPException(403, "Only the project lead, senior, or admin can edit a project")
    for k, v in payload.model_dump().items():
        setattr(p, k, v)
    db.commit(); db.refresh(p)
    return p


@router.post("/{project_id}/close")
def close_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Mark a project as Completed.

    Authorisation:
      * **Project owner** (the user whose id is on `Project.lead_id`)
        is the canonical signer-off. Admin retains an override for
        operational cleanup — but senior + everyone else no longer
        has the close power, even on projects they were granted view
        access to. Earlier any senior could close a project they
        weren't leading, which kept slipping past internal review.
      * **Gate:** every report under the project must be in a final
        review state — ``approved`` or ``published`` — before close
        is allowed. A project with a draft / in-review / rejected
        report on any of its reports is refused with 409 + a list of
        the blocking reports.

    Idempotent: closing an already-completed project is a no-op.
    Use POST /reopen to undo.
    """
    from ..models import Role, AuditLog, ReportReviewStatus
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")

    # Authorisation — owner OR admin only.
    if current_user.role != Role.admin and p.lead_id != current_user.id:
        raise HTTPException(
            403,
            "Only the project owner (lead) or an admin can close a project.",
        )

    if p.status == "completed":
        return {"ok": True, "status": p.status, "noop": True}

    # Gate — every report on the project must have its CURRENT version
    # in approved/published. We check the latest version per report
    # (the consultant ships off the highest version they generated).
    FINAL_STATES = {
        ReportReviewStatus.approved.value,
        ReportReviewStatus.published.value,
    }
    not_final: list[dict] = []
    for r in p.reports:
        cv = r.current_version_obj if hasattr(r, "current_version_obj") else None
        # Fall back to highest-version row if the convenience property
        # isn't loaded — same logic as the version-list endpoints.
        if cv is None and getattr(r, "versions", None):
            cv = max(r.versions, key=lambda x: x.version)
        if cv is None:
            not_final.append({"report_id": r.id, "report_name": r.name,
                              "version": None, "review_status": "missing"})
            continue
        if (cv.review_status or "draft") not in FINAL_STATES:
            not_final.append({
                "report_id":     r.id,
                "report_name":   r.name,
                "version":       cv.version,
                "review_status": cv.review_status or "draft",
            })
    if not_final:
        raise HTTPException(409, detail={
            "error":   "reports_not_final",
            "message": (f"{len(not_final)} report(s) aren't in a final "
                        "review state yet. Approve or publish the latest "
                        "version of each before closing the project."),
            "blocking_reports": not_final,
        })

    prev = p.status
    p.status = "completed"
    db.add(AuditLog(actor_id=current_user.id, action="project.close",
                    object_type="project", object_id=project_id,
                    detail={"previous_status": prev, "project_name": p.name,
                            "reports_at_close": len(p.reports)}))
    db.commit()
    return {"ok": True, "status": p.status}


@router.post("/{project_id}/reopen")
def reopen_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Flip a completed project back to active. Mirror of /close — same
    authorisation, idempotent on an already-active project."""
    from ..models import Role, AuditLog
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")
    # Mirror the close-side authorisation: owner OR admin only. The
    # "senior can close someone else's project" backdoor that used
    # to live here is gone — reopen authority needs to match close
    # authority or it'd be possible to bypass the all-reports-final
    # gate by closing-then-reopening with a senior account.
    if current_user.role != Role.admin and p.lead_id != current_user.id:
        raise HTTPException(403, "Only the project owner (lead) or an admin can reopen a project.")
    if p.status == "active":
        return {"ok": True, "status": p.status, "noop": True}
    prev = p.status
    p.status = "active"
    db.add(AuditLog(actor_id=current_user.id, action="project.reopen",
                    object_type="project", object_id=project_id,
                    detail={"previous_status": prev, "project_name": p.name}))
    db.commit()
    return {"ok": True, "status": p.status}


@router.delete("/{project_id}")
def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a project and every report under it. Only the project lead,
    a senior, or an admin can perform this. Notifies every user who had
    access to the project — by email when SMTP is configured, otherwise
    via a persisted .eml fallback so dev installs still see the message.

    Cascades follow SQLAlchemy's relationship config — Reports / Versions
    / Findings under this project disappear with it.
    """
    from ..models import AuditLog, Report, ReportAccess, ReportTemplate
    from ..services.email_send import send_mail

    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")

    # Authorisation: lead OR admin OR senior. Consultants who only have a
    # report grant inside the project cannot blow the whole project away.
    from ..models import Role
    if (current_user.role not in (Role.admin, Role.senior)
            and p.lead_id != current_user.id):
        raise HTTPException(403, "Only the project lead, senior, or admin can delete a project")

    # Build the recipient list BEFORE deletion: project lead + every user
    # holding any ReportAccess on a report inside this project + every
    # report owner. De-dupe by user id.
    recipients: dict[int, User] = {}
    if p.lead_id:
        lead = db.get(User, p.lead_id)
        if lead: recipients[lead.id] = lead
    reports_in_project = db.query(Report).filter(Report.project_id == project_id).all()
    for r in reports_in_project:
        if r.created_by_id:
            owner = db.get(User, r.created_by_id)
            if owner: recipients[owner.id] = owner
        for g in (r.access_grants or []):
            u = db.get(User, g.user_id)
            if u: recipients[u.id] = u
    # The deleter shouldn't email themselves
    recipients.pop(current_user.id, None)

    project_name = p.name
    client_name = p.client_name

    db.add(AuditLog(actor_id=current_user.id, action="project.delete",
                    object_type="project", object_id=project_id,
                    detail={"name": project_name,
                            "client": client_name,
                            "reports_deleted": len(reports_in_project),
                            "notified_user_ids": list(recipients.keys())}))
    # Manually delete reports first so the cascade clears its children
    # (versions / findings / access grants / etc) in the right order. Some
    # earlier installs may have relationships without explicit cascade.
    for r in reports_in_project:
        db.delete(r)
    db.delete(p)
    db.commit()

    # Send the notifications (best-effort, never raises)
    from ..services import email_templates as _email_tmpls
    for u in recipients.values():
        if not u.email:
            continue
        try:
            subject, body_text, body_html = _email_tmpls.render_template(
                db, "project_deleted",
                {"user": u, "actor_username": current_user.username,
                 "project_name": project_name, "client_name": client_name},
            )
            send_mail(u.email, subject=subject,
                      body_text=body_text, body_html=body_html)
        except Exception:
            pass

    return {
        "ok": True,
        "project_id": project_id,
        "reports_deleted": len(reports_in_project),
        "notified": len(recipients),
    }


# ============================================================
# Scope auto-population from an uploaded Nessus CSV / Excel tracker
# ============================================================
# Two endpoints to keep the workflow flexible:
#   POST /{pid}/scope/parse  -> dry-run: returns the extracted targets but
#                               doesn't touch the project. Frontend shows
#                               the preview so the consultant can prune.
#   POST /{pid}/scope/apply  -> takes a JSON list of confirmed targets and
#                               writes them onto project.scope_targets.
# This split is deliberate — every consultant tracker contains noise that
# we don't want to commit blindly.


@router.post("/{project_id}/scope/parse")
def parse_scope_file(
    project_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Dry-run parse of a Nessus CSV or Excel tracker. Returns the targets
    we'd add — caller previews + confirms via /scope/apply."""
    from .permissions import require_project_visibility
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")
    require_project_visibility(db, user, p)
    suffix = Path(file.filename or "").suffix.lower()
    if suffix not in (".csv", ".xlsx", ".xls", ".xlsm"):
        raise HTTPException(400, f"Unsupported file type: {suffix or '(none)'}. "
                                  "Use a Nessus CSV or an Excel tracker.")

    # Persist to a temp file so the parser can re-open it (openpyxl needs a path).
    import os as _os
    tmp_fd, tmp_path_str = tempfile.mkstemp(suffix=suffix)
    _os.close(tmp_fd)
    tmp_path = Path(tmp_path_str)
    try:
        _stream_save(file.file, tmp_path, max_bytes=10 * 1024 * 1024)
        try:
            result = _scope_import.parse_any(tmp_path)
        except HTTPException:
            raise
        except Exception as e:
            _logger.warning("Scope file parse failed for project %s: %s", project_id, e)
            raise HTTPException(400, "Could not parse the uploaded scope file. Please check the format and try again.")
    finally:
        try: tmp_path.unlink(missing_ok=True)
        except Exception: pass

    # Pre-merge with the project's existing targets so the consultant can see
    # what's new vs already present.
    existing = list(p.scope_targets or [])
    new_targets = [t for t in result["targets"] if t not in existing]
    return {
        "project_id": project_id,
        "source": result["source"],
        "host_count": result["host_count"],
        "warnings": result["warnings"],
        "existing_targets": existing,
        "new_targets": new_targets,
        "all_targets": existing + new_targets,
    }


@router.post("/{project_id}/scope/import-from-scan")
async def import_scope_from_scan(
    project_id: int,
    files: list[UploadFile] = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Pull every host out of one or more Nessus CSV / Excel exports
    and return the deduplicated list as candidate scope targets.

    Designed for the Infra VA / Infra VAPT / OT VAPT workflow where
    the consultant re-runs the scan a few times during an engagement
    and wants the project scope to reflect "every IP this scan
    actually saw" without hand-pasting. Handles BOTH the standard
    Nessus column layout (one IP per row in the Host column) and
    the pipeline's grouped output (multiple comma-joined IPs in a
    single cell, produced when `group_ips_in_by_category=True`).

    This is a PREVIEW endpoint — it doesn't mutate the project. The
    UI gets back a sorted unique IP list and renders it into the
    scope textarea; the consultant then clicks Save (which calls
    `/scope/apply` in `replace` or `merge` mode) to persist.

    Permissions follow the same rules as `apply_scope` — only
    admin / senior / project-lead can mutate scope, but anyone with
    view access can PREVIEW. We don't gate the preview here so
    consultants can use it to spot drift even on projects they
    don't lead.
    """
    from ..models import Role  # local — keeps the top-of-file import block clean
    from .permissions import require_project_visibility

    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")
    require_project_visibility(db, user, p)

    if not files:
        raise HTTPException(400, "Upload at least one CSV or Excel file.")

    accepted_csv  = {".csv"}
    accepted_xlsx = {".xlsx", ".xls"}
    rejected: list[str] = []
    accepted: list[tuple[str, bytes, str]] = []
    for f in files:
        name = (f.filename or "").strip()
        if not name:
            continue
        ext = ("." + name.rsplit(".", 1)[-1].lower()) if "." in name else ""
        if ext in accepted_csv:
            accepted.append((name, await f.read(), "csv"))
        elif ext in accepted_xlsx:
            accepted.append((name, await f.read(), "xlsx"))
        else:
            rejected.append(name)
    if not accepted:
        raise HTTPException(
            400,
            "No valid files in the upload. Accepts .csv / .xlsx / .xls. "
            "Rejected: " + (", ".join(rejected) or "(none)"),
        )

    # Parse + extract — keep the per-file summary so the UI can show
    # "12 IPs from internal.csv + 5 IPs from external.csv = 17 unique".
    per_file: list[dict] = []
    all_ips: set[str] = set()
    for name, data, kind in accepted:
        try:
            ips = _extract_hosts_from_scan(data, kind)
        except Exception as e:                              # pragma: no cover
            ips = []
            per_file.append({"filename": name, "ips": 0,
                              "error": f"{type(e).__name__}: {e}"})
            continue
        all_ips.update(ips)
        per_file.append({"filename": name, "ips": len(ips)})

    sorted_ips = sorted(all_ips, key=_scope_sort_key)
    return {
        "ok": True,
        "project_id": project_id,
        "unique_ips": sorted_ips,
        "count": len(sorted_ips),
        "per_file": per_file,
        "rejected_files": rejected,
        # `existing` is what the project currently holds — handy for
        # the UI to compute the diff before showing the consultant
        # what would change on `replace` vs `merge`.
        "existing": list(p.scope_targets or []),
    }


def _scope_sort_key(s: str):
    """Sort key that orders dotted-IPv4 numerically, falling back to
    a plain alphabetical sort for hostnames / URLs / IPv6.
    """
    parts = s.split(".")
    if len(parts) == 4 and all(p.isdigit() for p in parts):
        try:
            return (0,) + tuple(int(p) for p in parts)
        except ValueError:
            pass
    return (1, s.lower())


def _extract_hosts_from_scan(data: bytes, kind: str) -> list[str]:
    """Pull every IP / hostname out of the Host column of a Nessus
    CSV or Excel export.

    Returns a deduplicated list (order preserved by first
    appearance). Empty input or no Host column returns an empty
    list — the caller treats that as "this file contributed
    nothing" rather than an error.

    Handles BOTH layouts:
      * one host per row (canonical Nessus CSV)
      * multiple hosts per cell, comma-joined (what the pipeline
        emits when `group_ips_in_by_category=True`)
    Also splits on `;` and whitespace as a defensive measure in
    case other tools join with different separators.
    """
    if kind == "csv":
        return _extract_hosts_from_csv_bytes(data)
    if kind == "xlsx":
        return _extract_hosts_from_xlsx_bytes(data)
    return []


_HOST_HEADER_ALIASES = (
    "host", "hosts", "ip", "ips", "ip address", "ip addresses",
    "asset", "affected host", "target", "address",
)


def _is_host_header(cell: str) -> bool:
    s = (cell or "").strip().lower()
    return s in _HOST_HEADER_ALIASES


# Tokens that look like IPv4 dotted-quads or hostnames. Used to
# filter out empty / "N/A" / junk values after splitting a cell.
_TOKEN_RE = re.compile(
    r"[A-Za-z0-9][A-Za-z0-9\.\-:/]{0,254}[A-Za-z0-9]|[A-Za-z0-9]"
)


def _split_host_cell(cell: str) -> list[str]:
    """Split a Host cell into individual tokens. Handles
    comma-joined IPs (`10.0.0.1, 10.0.0.2`), space-separated
    (`10.0.0.1 10.0.0.2`), and semicolon-joined. Trims surrounding
    quotes / whitespace.
    """
    if not cell:
        return []
    text = str(cell).strip().strip('"').strip("'")
    if not text:
        return []
    # Replace common separators with commas, then split.
    for sep in (";", "\n", "\r", "\t"):
        text = text.replace(sep, ",")
    out = []
    for tok in text.split(","):
        tok = tok.strip().strip('"').strip("'")
        if not tok:
            continue
        if tok.lower() in ("n/a", "none", "null", "-"):
            continue
        # Defensive: pull any inline token via the regex so a stray
        # parenthesis / port suffix doesn't sneak in.
        m = _TOKEN_RE.search(tok)
        if m:
            out.append(m.group(0))
        else:
            out.append(tok)
    return out


def _extract_hosts_from_csv_bytes(data: bytes) -> list[str]:
    import csv as _csv
    from io import StringIO
    # Nessus CSVs are utf-8 with a BOM. Strip it.
    try:
        text = data.decode("utf-8-sig", errors="replace")
    except Exception:
        text = data.decode("latin-1", errors="replace")
    reader = _csv.reader(StringIO(text))
    rows = list(reader)
    if not rows:
        return []
    # Header is the first row that has ≥ 2 non-empty cells (defends
    # against leading comment lines some scanners prepend).
    header_idx = None
    for i, row in enumerate(rows):
        if sum(1 for c in row if (c or "").strip()) >= 2:
            header_idx = i
            break
    if header_idx is None:
        return []
    header = rows[header_idx]
    host_cols = [i for i, h in enumerate(header) if _is_host_header(h)]
    if not host_cols:
        return []
    out: list[str] = []
    seen: set[str] = set()
    for row in rows[header_idx + 1:]:
        for col in host_cols:
            if col >= len(row):
                continue
            for tok in _split_host_cell(row[col]):
                if tok not in seen:
                    seen.add(tok)
                    out.append(tok)
    return out


def _extract_hosts_from_xlsx_bytes(data: bytes) -> list[str]:
    from io import BytesIO
    try:
        from openpyxl import load_workbook
    except ImportError:                                     # pragma: no cover
        return []
    try:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
    except Exception:
        return []
    out: list[str] = []
    seen: set[str] = set()
    # Look at EVERY sheet — Nessus exports usually have one sheet,
    # but the pipeline output may have a few. Per-sheet host
    # extraction is independent so multiple sheets just add to the
    # union.
    for ws in wb.worksheets:
        if ws.max_row is None or ws.max_row < 2:
            continue
        header_row = None
        header_idx = 0
        for i, row in enumerate(
                ws.iter_rows(min_row=1, max_row=5, values_only=True), start=1):
            if not row:
                continue
            if sum(1 for c in row if c and str(c).strip()) >= 2:
                header_row = row
                header_idx = i
                break
        if not header_row:
            continue
        host_cols = [
            j for j, h in enumerate(header_row) if _is_host_header(h or "")
        ]
        if not host_cols:
            continue
        for row in ws.iter_rows(min_row=header_idx + 1, values_only=True):
            if not row:
                continue
            for col in host_cols:
                if col >= len(row):
                    continue
                cell = row[col]
                if cell is None:
                    continue
                for tok in _split_host_cell(str(cell)):
                    if tok not in seen:
                        seen.add(tok)
                        out.append(tok)
    wb.close()
    return out


@router.post("/{project_id}/scope/apply")
def apply_scope(
    project_id: int,
    targets: list[str] = Body(..., embed=True),
    mode: str = Body("merge", embed=True),  # "merge" | "replace"
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Persist the user-confirmed scope. `mode=merge` (default) appends new
    targets to the existing list; `mode=replace` overwrites the scope
    outright. Either way the order is preserved and duplicates dropped."""
    from ..models import Role
    p = db.get(Project, project_id)
    if not p:
        raise HTTPException(404, "Project not found")
    # Mutating scope is a senior/lead/admin action — random consultants
    # shouldn't be rewriting the in-scope targets list on someone else's
    # engagement.
    if user.role not in (Role.admin, Role.senior) and p.lead_id != user.id:
        raise HTTPException(403, "Only the project lead, senior, or admin can change scope")
    if mode not in ("merge", "replace"):
        raise HTTPException(400, "mode must be 'merge' or 'replace'")
    base = list(p.scope_targets or []) if mode == "merge" else []
    seen = set(base)
    out = list(base)
    for t in targets or []:
        t = (t or "").strip()
        if t and t not in seen:
            seen.add(t); out.append(t)
    p.scope_targets = out
    flag_modified(p, "scope_targets")
    db.commit()
    return {"ok": True, "scope_targets": out, "count": len(out)}


# ============================================
# TEAM ASSIGNMENT ENDPOINTS
# ============================================

@router.post("/{project_id}/assign")
def assign_user_to_project(
    project_id: int,
    user_id: int = Body(..., embed=True),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Assign a user to a project.

    Membership is stored as a list of user ids on `project.details["assigned_user_ids"]`.
    Previously this code path tried to use a `ProjectMember` table that
    doesn't exist in the schema, fell into an `ImportError` branch, and
    then set a `project.assigned_users` attribute that isn't a real DB
    column — so every assignment silently disappeared on the next page
    load. Persisting inside the existing JSON column avoids a migration
    and gets picked up by `user_can_see_project()` for the IDOR gate.
    """
    from ..models import Role, AuditLog
    from ..services.email_send import send_mail
    from ..services import email_templates as _email_tmpls

    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    if (current_user.role not in (Role.admin, Role.senior)
            and project.lead_id != current_user.id):
        raise HTTPException(403, "Only the project lead, senior, or admin can assign team members")

    user = db.get(User, user_id)
    if not user:
        raise HTTPException(404, "User not found")
    if not user.is_active:
        raise HTTPException(400, "Cannot assign a disabled user")
    if user.id == project.lead_id:
        raise HTTPException(400, "User is already the project lead")

    details = dict(project.details or {})
    assigned = list(details.get("assigned_user_ids") or [])
    if user.id in assigned:
        raise HTTPException(400, "User already assigned to this project")
    assigned.append(user.id)
    details["assigned_user_ids"] = assigned
    project.details = details
    flag_modified(project, "details")

    # AuditLog row is the source of truth for the in-app notification
    # bell — see /api/notifications below. We record the actor (assigner)
    # so the user can see who added them.
    db.add(AuditLog(
        actor_id=current_user.id, action="project.member.assigned",
        object_type="project", object_id=project.id,
        detail={"assigned_user_id": user.id, "project_name": project.name,
                "client_name": project.client_name},
    ))
    db.commit()

    # Best-effort email notification. Goes through `notify_user` so the
    # recipient's per-user opt-out is honoured. Never raises — the
    # assignment commit already happened above.
    from ..services.notifier import notify_user
    from ..services.url_helpers import absolute_url
    notify_user(
        db, user, "project_assigned",
        {"user": user, "actor_username": current_user.username,
         "project_name": project.name, "client_name": project.client_name,
         "project_url": absolute_url(f"/projects/{project.id}")},
        actor_user_id=current_user.id,
    )

    return {"success": True, "message": f"User {user.username} assigned to project"}


@router.get("/{project_id}/team")
def get_project_team(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Get all users assigned to a project."""
    from .permissions import require_project_visibility
    # Check project exists
    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    require_project_visibility(db, current_user, project)

    # Read from the persisted JSON store. The previous `ProjectMember` /
    # `project.assigned_users` mix never worked — neither the table nor
    # the attribute existed, so every team listing came back empty.
    assigned_ids = (project.details or {}).get("assigned_user_ids") or []
    if not assigned_ids:
        members = []
    else:
        members = db.query(User).filter(User.id.in_(assigned_ids)).all()

    return [
        {
            "id": m.id,
            "username": m.username,
            "email": m.email,
            "role": m.role.value
        }
        for m in members
    ]


@router.delete("/{project_id}/team/{user_id}")
def remove_user_from_project(
    project_id: int,
    user_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Remove a user from a project."""
    from ..models import Role
    project = db.get(Project, project_id)
    if not project:
        raise HTTPException(404, "Project not found")
    if (current_user.role not in (Role.admin, Role.senior)
            and project.lead_id != current_user.id):
        raise HTTPException(403, "Only the project lead, senior, or admin can remove team members")

    from ..models import AuditLog
    details = dict(project.details or {})
    assigned = list(details.get("assigned_user_ids") or [])
    if user_id not in assigned:
        raise HTTPException(404, "User not assigned to this project")
    assigned.remove(user_id)
    details["assigned_user_ids"] = assigned
    project.details = details
    flag_modified(project, "details")
    db.add(AuditLog(
        actor_id=current_user.id, action="project.member.removed",
        object_type="project", object_id=project.id,
        detail={"removed_user_id": user_id, "project_name": project.name},
    ))
    db.commit()

    return {"success": True, "message": "User removed from project"}