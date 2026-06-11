"""
Two-way Risk Register sync between an Excel tracker and the VAPT Reporter
report findings.

Import (XLSX -> findings)
-------------------------
  parse_risk_register(path)
    Opens the workbook, finds a sheet named (case-insensitively) "Risk
    Register", locates the header row, fuzzy-matches each known column to a
    finding field, then yields one dict per data row. Skips rows whose every
    cell is blank.

Export (findings -> XLSX)
-------------------------
  write_risk_register(findings, *, template_path=None, output_path)
    If `template_path` points at the consultant's master tracker template,
    we load a *copy* of it via openpyxl, clear the data rows in the Risk
    Register sheet (preserving the header + all other sheets), then write
    one row per finding using the same column mapping the parser learned.
    Without a template, we synthesize a fresh single-sheet workbook with
    sane defaults.

Field mapping is fuzzy: we lower-case + strip both the spreadsheet header
and the candidate aliases below, then compare. This survives slightly
different consultant templates (e.g. "Vulnerability Title" vs "Finding
Title" vs just "Title").

The risk register sample shipped with this repo lives in OWASP-2025
nomenclature; aliases below cover that and the common VibeDocs tracker
columns.
"""
from __future__ import annotations
import html as _html_module
import io
import logging
import ipaddress
import re
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable, Optional

log = logging.getLogger(__name__)


# ── HTML-to-plain-text helper ─────────────────────────────────────────────────

def _strip_html_to_text(value: str) -> str:
    """Convert a Quill-editor HTML string to plain text suitable for an Excel cell.

    Steps:
    1. Remove base64 data-URL <img> tags entirely — they contain kilobytes of raw
       bytes that would otherwise spill into the Excel cell as unreadable text.
    2. Replace block-level tags (<p>, <br>, <li>, headings) with newlines so
       paragraph structure survives as multi-line cell text.
    3. Strip all remaining tags.
    4. Decode HTML entities (& amp; → &, &lt; → <, etc.).
    5. Collapse excessive blank lines.
    """
    if not isinstance(value, str) or "<" not in value:
        return value
    # 1. Drop base64-embedded images — these are data URLs from Quill paste
    text = re.sub(
        r'<img\b[^>]*\bsrc=["\']data:[^"\']*["\'][^>]*/?>',
        '',
        value,
        flags=re.IGNORECASE | re.DOTALL,
    )
    # 2. Block elements → newlines
    text = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    text = re.sub(r'</(p|div|li|h[1-6]|blockquote|pre)>', '\n', text, flags=re.IGNORECASE)
    # 3. Strip remaining tags
    text = re.sub(r'<[^>]+>', '', text)
    # 4. Decode HTML entities
    text = _html_module.unescape(text)
    # 5. Collapse runs of blank lines to at most one blank line
    text = re.sub(r'\n{3,}', '\n\n', text)
    return text.strip()


def _extract_base64_images(html_value: str) -> list[bytes]:
    """Return a list of decoded image byte-blobs from base64 data-URL <img> tags.

    Used to embed pasted screenshots from Quill into the appropriate Excel column
    when the finding's text fields (description, impact, etc.) contain pasted images.
    Returns an empty list when there are no embedded images or on any decode error.
    """
    import base64 as _b64
    blobs: list[bytes] = []
    for m in re.finditer(
        r'<img\b[^>]*\bsrc=["\']data:image/[^;]+;base64,([A-Za-z0-9+/=\s]+)["\'][^>]*/?>',
        html_value or "",
        flags=re.IGNORECASE | re.DOTALL,
    ):
        try:
            blobs.append(_b64.b64decode(re.sub(r'\s+', '', m.group(1))))
        except Exception:
            pass
    return blobs


# Risk Register sheet name candidates (case-insensitive contains).
SHEET_NAME_CANDIDATES = [
    "risk register",
    "findings",
    "risks",
    "vulnerabilities",
    "issues",
]


# Header text -> canonical finding-dict key. Aliases are matched in a
# priority chain:  exact -> word-boundary -> substring.  This avoids the
# previous bug where "Implications" (a common name for the impact column
# in consultant trackers) was matching "title" via accidental substring
# overlap.  The aliases are listed in order of preference: the first hit
# wins, so put the most specific phrasing first.
COLUMN_ALIASES: dict[str, list[str]] = {
    # S/N is a metadata column, not a finding field. It anchors row parsing
    # (start at S/N == 1, stop at first blank S/N below the data block) but
    # never lands in the finding dict.
    "_sn":              ["s/n", "s.n.", "sn", "s no", "serial", "serial no",
                          "serial number"],
    "index":            ["#", "no.", "no", "id", "finding id", "finding #",
                          "finding no", "ref", "ref.", "ref no", "reference id"],
    "title":            ["issue title", "finding title", "vulnerability title",
                          "vuln title", "title", "finding", "issue", "issue name",
                          "vulnerability", "name"],
    # Severity = the text rating (Critical/High/Medium/Low/Info). The VibeDocs
    # trackers ship two related columns next to each other:
    #   * "CVSS Risk Rating" / "Overall Risk Rating" — text severity
    #   * "CVSS Score"                                — numeric score
    # The most-specific aliases come first so a header like "CVSS Risk
    # Rating" gets mapped to severity, not to cvss_score via substring.
    "severity":         ["cvss risk rating", "overall risk rating",
                          "risk rating", "severity",
                          "risk level", "risk", "rating", "criticality"],
    "cvss_score":       ["cvss v4 score", "cvss v3 score", "cvss base score",
                          "cvss score", "cvss", "base score", "score"],
    "cvss_vector":      ["cvss 4.0 vector", "cvss v4.0 vector",
                          "cvss 4 vector", "cvss4 vector",
                          "cvss vector", "cvss string",
                          "vector string", "vector"],
    "description":      ["issue description", "vulnerability description",
                          "observations", "observation", "description",
                          "details"],
    "impact":           ["business impact", "potential impact", "implications",
                          "implication", "impact"],
    "remediation":      ["recommendations", "recommendation",
                          "remediation", "mitigation", "suggested fix", "fix"],
    "references":       ["external references", "references", "reference",
                          "links", "cve"],
    # Affected Port(s) / Protocol — primarily used by the infra (NVA/VAPT)
    # trackers and the per-group sheets so each host's port + protocol land
    # in their own columns.
    "port":             ["affected port(s)", "affected port", "affected ports",
                          "port(s)", "port", "service port"],
    "protocol":         ["protocol", "proto", "transport"],
    # Benchmark / Benchmark Clauses — Cloud VAPT tracker columns (CIS refs).
    "benchmark":         ["benchmark"],
    "benchmark_clauses": ["benchmark clauses", "benchmark clause"],
    # Order matters: catch the longer, more-specific VibeDocs column
    # names first ("Affected File : Affected Line", "Affected Module / URL"
    # etc.) so we don't accidentally map the generic "Affected" header
    # before a more specific one wins.
    "affected_asset":   ["affected file : affected line",
                          "affected resource(s)/instance(s)",
                          "affected resources/instances",
                          "affected resource(s)",
                          "affected resources",
                          "affected resource",
                          "affected module / url",
                          "affected module/url",
                          "affected module",
                          "affected endpoint(s)",
                          "affected endpoint",
                          "affected url", "affected host",
                          "affected asset", "affected",
                          "asset", "target", "host", "url", "endpoint", "ip"],
    "poc_steps":        ["steps to replicate", "steps to reproduce",
                          "reproduction steps", "poc",
                          "proof of concept", "steps", "evidence",
                          "reproduction"],
    "status":           ["finding status", "status", "state"],
    # VibeDocs trackers use "Post Review Observations" for the retest
    # notes column. The generic "retest …" aliases stay as a fallback
    # for non-VibeDocs trackers.
    "retest_notes":     ["post review observations", "post review observation",
                          "retest notes", "retest result", "retest",
                          "follow-up", "follow up"],
    # NOTE: "client owner" is intentionally NOT in this list — it's
    # the *name* of the engagement POC on the client side (one value
    # per report, repeated on every row of the tracker), not a free-
    # text statement. It has its own canonical key below so we can
    # auto-populate it from project.client_poc / report.details
    # without polluting the per-finding statement field.
    "client_statement": ["client statement", "client response", "client comments",
                          "management comments",
                          "customer response", "customer statement",
                          "client"],
    # Per-row Client Owner (POC name). The importer reads it onto
    # report.details.client_owner; the exporter writes it back into
    # every row's Client Owner column so the consultant doesn't have
    # to type it N times.
    "client_owner":     ["client owner", "client poc", "client contact",
                          "owner", "asset owner"],
    "owasp_category":   ["owasp top 10", "owasp 2025", "owasp 2021",
                          "owasp category", "owasp"],
    "cwe":              ["cwe id", "cwe-id", "cwe"],
    # Date Raised — the date the finding was first reported to the
    # client. VibeDocs trackers expose this as a single column that's
    # the same for every row (= the report-creation date). The export
    # writes `report.details.report_date` into every row's cell; the
    # importer picks it back up and merges it onto `report.details`.
    "date_raised":      ["date raised", "date found", "date identified",
                          "date reported", "raised on", "date of finding",
                          "report date", "date"],
    # Tester column captured per-row. We aggregate distinct values across
    # all rows into report.details.tester_names on import; on export the
    # same column is filled from that list (or per-row added_by user).
    # NOTE: "dt tester2" intentionally NOT listed here — it maps to the
    # separate `retest_tester_name` key below so initial and retest
    # testers round-trip to distinct columns.
    "tester_name":      ["dt tester", "dt tester 1", "tester",
                          "tested by", "consultant"],
    # Retest tester — the "DT Tester2" column on VibeDocs trackers.
    # Populated on export for retest / report-update versions only.
    "retest_tester_name": ["dt tester2", "dt tester 2", "dt tester 2nd",
                            "retest tester", "retest consultant", "retester"],
    # System / application name — the per-row "System" column on the
    # VibeDocs tracker. On export, filled from report.details.application_name
    # so every row carries the application being tested.
    "system":           ["system", "system name", "system / application",
                          "application name", "application", "asset name",
                          "system under test"],
    # Area of Review — e.g. "Web PT", "API VAPT". Derived from the
    # report template code and written to every row on export.
    "review_type":      ["area of review", "review", "review type",
                          "type of review", "engagement type",
                          "area of test", "assessment type"],
    # Date Follow-Up — the retest / follow-up date column. Populated on
    # export for retest and report-update versions from report_date.
    "date_follow_up":   ["date follow-up", "date follow up",
                          "follow-up date", "follow up date",
                          "retest date", "date retest",
                          "date of follow-up"],
    # Screenshot anchor columns. These aren't finding text fields —
    # `_screenshot` and `_retest_screenshot` are sentinels (underscore
    # prefix matches the `_sn` convention) used to know which column
    # an embedded image is anchored to. The image extraction pass
    # (extract_risk_register_images) returns row+col coordinates;
    # the tracker importer matches col against these indices so the
    # screenshot lands on the right finding field (screenshots vs
    # retest_evidence).
    "_screenshot":         ["screenshot", "screenshots", "evidence",
                             "evidence screenshot"],
    "_retest_screenshot":  ["post review screenshot",
                             "post-review screenshot",
                             "retest screenshot", "retest evidence",
                             "follow-up screenshot"],
}


STATUS_NORMALISE = {
    "open": "Open",
    "new": "Open",
    "closed": "Closed",
    "resolved": "Closed",
    "fixed": "Closed",
    "risk accepted": "Risk Accepted",
    "accepted": "Risk Accepted",
    "false positive": "False Positive",
    "fp": "False Positive",
    "n/a": "N/A",
    "na": "N/A",
    "not applicable": "N/A",
    "in remediation": "In Remediation",
    "in progress": "In Remediation",
    "remediation": "In Remediation",
}


SEVERITY_NORMALISE = {
    "critical": "Critical",
    "crit": "Critical",
    "high": "High",
    "h": "High",
    "medium": "Medium",
    "med": "Medium",
    "m": "Medium",
    "moderate": "Medium",
    "low": "Low",
    "l": "Low",
    "informational": "Informational",
    "info": "Informational",
    "information": "Informational",
    "i": "Informational",
    "none": "Informational",
}


# ============================================================
# Import
# ============================================================

def parse_risk_register(path: Path, *,
                          override_mapping: Optional[dict[str, int]] = None
                          ) -> dict[str, Any]:
    """Read findings out of the tracker. Returns:
        {
          "sheet": <chosen sheet name>,
          "header_row": <1-based row index of detected header>,
          "headers": [<header text per column>],
          "column_map": {<finding_field>: <column_index_0based>},
          "rows": [<finding dict>, ...],
        }

    Anchoring + row range:
      * If the workbook has an "S/N" column anywhere on the header row,
        we treat it as the row driver: parsing starts at the first row
        where S/N is a non-blank value of `1` (or just non-blank if a
        literal 1 isn't found) and stops at the first row below the data
        block where S/N is empty.
      * Without an S/N column we fall back to "first row with a title,
        stop at the next blank-row".

    Column matching:
      * If `override_mapping` is supplied (from the preview walkthrough),
        it wins outright — the user is the final arbiter.
      * Otherwise the alias chain (exact -> word-boundary -> substring) is
        applied, with earlier aliases winning to defeat accidental matches.
    """
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        ws_name = _pick_sheet(wb.sheetnames)
        if not ws_name:
            raise ValueError(
                "Could not find a Risk Register sheet in this workbook. "
                f"Expected one of: {SHEET_NAME_CANDIDATES}. "
                f"Found sheets: {wb.sheetnames}"
            )
        ws = wb[ws_name]
        header_row, headers = _find_header_row(ws)
        if not headers:
            raise ValueError(
                f"Could not locate a header row in sheet '{ws_name}'. "
                "Make sure your tracker has columns like Title / Severity / "
                "Description in the first 20 rows."
            )

        col_map = dict(override_mapping or _map_columns(headers))

        # Identify the S/N column (if any) — used as the row anchor below.
        # _map_columns places "_sn" into col_map; pop it so it doesn't leak
        # into the finding dicts.
        sn_col = col_map.pop("_sn", None)
        # Same for screenshot anchor columns. They're keyed with a
        # leading underscore so they get popped here and never end up
        # in the finding fields. Surfaced separately on the return
        # dict so the import path can route embedded images to the
        # right finding attribute (screenshots vs retest_evidence).
        screenshot_col        = col_map.pop("_screenshot", None)
        retest_screenshot_col = col_map.pop("_retest_screenshot", None)

        # Locate where the data block starts and where it ends, anchored to
        # S/N if available. We tolerate the first numeric S/N value being
        # something other than literally 1 (some templates start with "1.0"
        # or " 1 "), but require it to be non-blank.
        data_start = header_row + 1
        data_rows: list[tuple[int, tuple]] = []   # (sheet_row, row_tuple)
        for row_idx, row in enumerate(
                ws.iter_rows(min_row=data_start, values_only=True),
                start=data_start):
            sn_val = row[sn_col] if (sn_col is not None and sn_col < len(row)) else None
            row_blank = all(c is None or str(c).strip() == "" for c in row)
            if sn_col is not None:
                # Anchored to S/N
                if sn_val is None or (isinstance(sn_val, str) and not sn_val.strip()):
                    # Allow leading blank rows above the first 1; bail only
                    # AFTER we've started collecting.
                    if data_rows:
                        break
                    continue
                data_rows.append((row_idx, row))
            else:
                # Heuristic: blank-row terminator
                if row_blank:
                    if data_rows:
                        break
                    continue
                data_rows.append((row_idx, row))

        out: list[dict[str, Any]] = []
        for row_idx, row in data_rows:
            d: dict[str, Any] = {}
            for field, col_idx in col_map.items():
                if col_idx is None or col_idx >= len(row):
                    continue
                cell = row[col_idx]
                val = cell.value if hasattr(cell, "value") else cell
                if isinstance(val, str): val = val.strip()
                if val == "" or val is None: continue
                d[field] = val
            if not d.get("title"):
                # Without a title we can't surface this row as a finding.
                continue
            _normalize_finding(d)
            d["_sheet_row"] = row_idx
            out.append(d)

        return {
            "sheet": ws_name,
            "header_row": header_row,
            "headers": headers,
            "column_map": col_map,
            "sn_col": sn_col,
            "screenshot_col":        screenshot_col,
            "retest_screenshot_col": retest_screenshot_col,
            "rows": out,
        }
    finally:
        wb.close()


def _pick_sheet(names: list[str]) -> Optional[str]:
    lower = {n.lower(): n for n in names}
    for cand in SHEET_NAME_CANDIDATES:
        for low, orig in lower.items():
            if cand in low:
                return orig
    # Fall back to the first sheet only if it looks like findings (header row
    # with title-ish columns). Otherwise give up.
    return None


def _find_header_row(ws) -> tuple[int, list[str]]:
    """Score each of the first 20 rows by how many cells look like
    column-header text from our alias list, and return the highest-
    scoring row. Tie-break by row-with-the-most-non-empty-cells.

    Why scoring instead of the previous "first row with >=2 matches":
    VibeDocs trackers use ROW 1 for merged GROUP labels — "Risk",
    "Findings", "Ownership", "Follow-up" — and ROW 2 for the actual
    column names (S/N, System, CVSS Risk Rating, Issue Title, …).
    With the earlier first-match-wins logic, the four group labels
    accidentally substring-match severity / title / retest aliases and
    row 1 would win, locking us into a useless 4-column mapping. The
    score below makes row 2 dominate because it has 17+ matching
    headers vs row 1's 4 partial substring hits.
    """
    flat_aliases = {a.lower() for aliases in COLUMN_ALIASES.values() for a in aliases}
    best_row = 0
    best_score = (0, 0)   # (alias_matches, nonblank_cells)
    best_headers: list[str] = []
    for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=20, values_only=True), start=1):
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        nonblank = [c for c in cells if c]
        if len(nonblank) < 3:
            continue
        matches = sum(
            1 for c in nonblank
            if any(a == c or (len(a) > 3 and a in c) or
                    (len(c) > 3 and c in a)
                    for a in flat_aliases)
        )
        if matches < 2:
            continue
        score = (matches, len(nonblank))
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_headers = [str(c).strip() if c is not None else "" for c in row]
    return best_row, best_headers


def _normalize_header(h: str) -> str:
    return re.sub(r"\s+", " ", (h or "").strip().lower())


def _map_columns(headers: list[str]) -> dict[str, int]:
    """Return {finding_field: column_index_0_based}.

    Match priority:
      1. **Exact**         normalized header == normalized alias.
      2. **Word-boundary** alias appears as a whole word inside the header.
      3. **Substring**     alias appears anywhere (last-resort fallback).

    The most-specific alias for a field is tried first (the list is in
    priority order). A column is claimed by at most one field; subsequent
    field iterations skip already-claimed columns. This avoids the
    previous bug where "Implications" matched both `impact` and `title`
    via substring greediness.
    """
    norm_headers = [_normalize_header(h) for h in headers]
    claimed: set[int] = set()
    out: dict[str, int] = {}

    def try_pass(mode: str) -> None:
        for field, aliases in COLUMN_ALIASES.items():
            if field in out:
                continue
            for alias in aliases:
                a = alias.lower()
                hit = None
                for idx, h in enumerate(norm_headers):
                    if not h or idx in claimed:
                        continue
                    if mode == "exact" and h == a:
                        hit = idx; break
                    if mode == "word":
                        # Word-boundary using regex; escape the alias.
                        if re.search(r"(?:^|\W)" + re.escape(a) + r"(?:$|\W)", h):
                            hit = idx; break
                    if mode == "substr" and (a in h or h in a):
                        hit = idx; break
                if hit is not None:
                    out[field] = hit
                    claimed.add(hit)
                    break

    try_pass("exact")
    try_pass("word")
    try_pass("substr")
    return out


def _normalize_finding(d: dict[str, Any]) -> None:
    """Normalise types in place: severity / status to enum spelling,
    CVSS to float, tester name trimmed to a clean string."""
    # Severity
    sev = d.get("severity")
    if isinstance(sev, str):
        key = sev.strip().lower()
        d["severity"] = SEVERITY_NORMALISE.get(key, sev.strip().title())
    # Status — VibeDocs trackers use mixed-case + slightly varied
    # wording ("In Progress" vs "In Remediation"); normalise so the
    # FindingStatus enum accepts whatever lands in d["status"].
    st = d.get("status")
    if isinstance(st, str):
        key = st.strip().lower()
        d["status"] = STATUS_NORMALISE.get(key, st.strip().title())
    # CVSS
    score = d.get("cvss_score")
    if score is not None:
        try:
            d["cvss_score"] = float(score)
        except (ValueError, TypeError):
            d.pop("cvss_score", None)
    # Tester name: collapse whitespace, drop empties so the aggregator
    # downstream doesn't pick up phantom values.
    tn = d.get("tester_name")
    if isinstance(tn, str):
        tn = re.sub(r"\s+", " ", tn).strip()
        if tn:
            d["tester_name"] = tn
        else:
            d.pop("tester_name", None)
    # Coerce numeric-cell strings: openpyxl sometimes returns
    # description / impact text as float (when a sole numeric like "7"
    # is in the cell). Keep the value but stringify so downstream code
    # doesn't trip on type checks.
    for k in ("description", "impact", "remediation", "references",
              "poc_steps", "retest_notes", "client_statement"):
        v = d.get(k)
        if isinstance(v, (int, float)):
            d[k] = str(v)


# ============================================================
# Embedded-image extraction
# ============================================================
#
# Consultants often paste screenshots directly into the Risk Register
# spreadsheet's `Screenshot` or `Post Review Screenshot` columns. The
# regular parser uses `read_only=True` (faster, lower memory) which
# DROPS embedded images. This separate pass loads the workbook in
# full mode and pulls images by their cell anchor so we can attach
# them to the right ReportFinding.
#
# Notes:
#   * Excel anchors are 0-indexed at the XML layer; openpyxl exposes
#     `_from.row` / `_from.col`. We surface them 1-indexed to match
#     the rest of this module's row / column numbering.
#   * Some images use a `OneCellAnchor` (single top-left cell);
#     others use `TwoCellAnchor` (top-left + bottom-right). For
#     mapping to a finding, only the top-left matters — that's the
#     row the user dropped the screenshot onto.
#   * Image bytes live in the worksheet's `Image` object's `._data`
#     callable. We resolve to bytes here so the caller can write to
#     disk without holding the workbook open.

def extract_risk_register_images(path: Path) -> list[dict]:
    """Return every image embedded in the Risk Register sheet,
    keyed by its top-left anchor cell.

    Each returned dict carries::
        {
          "row":  <1-based sheet row>,
          "col":  <1-based sheet column>,
          "bytes": <raw image bytes>,
          "ext":  <file extension, lowercase, leading dot> e.g. ".png",
        }

    Empty list when the sheet has no images, when the workbook can't
    be opened in full mode, or when the Risk Register sheet itself is
    missing — never raises.
    """
    from openpyxl import load_workbook
    try:
        wb = load_workbook(str(path), data_only=True, read_only=False)
    except Exception as e:                                # pragma: no cover
        log.warning("Could not open workbook for image extraction: %s", e)
        return []
    try:
        ws_name = _pick_sheet(wb.sheetnames)
        if not ws_name:
            return []
        ws = wb[ws_name]
        images = list(getattr(ws, "_images", None) or [])
        if not images:
            return []
        out: list[dict] = []
        for img in images:
            anchor = getattr(img, "anchor", None)
            if not anchor:
                continue
            frm = getattr(anchor, "_from", None) or getattr(anchor, "from", None)
            if frm is None:
                continue
            row_0 = getattr(frm, "row", None)
            col_0 = getattr(frm, "col", None)
            if row_0 is None or col_0 is None:
                continue
            blob = _image_bytes(img)
            if not blob:
                continue
            ext = _image_ext_from_bytes(blob)
            out.append({
                "row":  int(row_0) + 1,
                "col":  int(col_0) + 1,
                "bytes": blob,
                "ext":  ext,
            })
        return out
    finally:
        try: wb.close()
        except Exception: pass


def _image_bytes(img) -> Optional[bytes]:
    """Resolve an openpyxl Image object's raw bytes. openpyxl exposes
    images through a callable `_data()` on its `Image` wrapper; older
    code paths also expose `.ref` (BytesIO) directly. Try both."""
    # Newer openpyxl: image.ref is a BytesIO-like with .getvalue().
    ref = getattr(img, "ref", None)
    if ref is not None:
        try:
            if hasattr(ref, "getvalue"):
                v = ref.getvalue()
                if isinstance(v, (bytes, bytearray)):
                    return bytes(v)
        except Exception:                                 # pragma: no cover
            pass
    # Older fallback: image._data is the raw blob loader callable.
    data_attr = getattr(img, "_data", None)
    if data_attr is not None:
        try:
            v = data_attr() if callable(data_attr) else data_attr
            if isinstance(v, (bytes, bytearray)):
                return bytes(v)
        except Exception:                                 # pragma: no cover
            pass
    return None


_MAGIC_BYTES = (
    (b"\x89PNG\r\n\x1a\n", ".png"),
    (b"\xff\xd8\xff",       ".jpg"),
    (b"GIF87a",             ".gif"),
    (b"GIF89a",             ".gif"),
    (b"BM",                 ".bmp"),
    (b"RIFF",               ".webp"),   # WebP starts with RIFF + WEBP
)


def _image_ext_from_bytes(blob: bytes) -> str:
    """Sniff a sensible extension from the image magic bytes. Falls
    back to `.png` so the screenshot still saves cleanly."""
    for magic, ext in _MAGIC_BYTES:
        if blob.startswith(magic):
            # WebP starts with RIFF but isn't always image/webp — verify.
            if ext == ".webp" and b"WEBP" not in blob[:16]:
                continue
            return ext
    return ".png"


# ============================================================
# Export
# ============================================================

# VibeDocs tracker convention for the severity column. Both the
# displayed text AND the cell fill colour are set per-row so the
# delivered XLSX matches the printed Word-report colour scheme.
#
# Display text quirks:
#   * "Informational" is rendered as "INFO" in the spreadsheet — the
#     full word doesn't fit cleanly in the narrow severity column on
#     the trackers and the team's house style abbreviates it.
#
# Shared Alignment object — center-horizontal for all written data cells.
# Defined at module level so every write helper references one object.
def _make_center_align():
    try:
        from openpyxl.styles import Alignment
        return Alignment(horizontal="center", vertical="center", wrap_text=True)
    except Exception:
        return None

_CENTER_ALIGN = _make_center_align()

# Cell fills (ARGB without the alpha byte — openpyxl prefixes "00"):
SEVERITY_DISPLAY = {
    "Critical":      "Critical",
    "High":          "High",
    "Medium":        "Medium",
    "Low":           "Low",
    # House style abbreviates "Informational" to "Info" in the tracker.
    "Informational": "Info",
    "Info":          "Info",
}
# IMPORTANT: this palette is for the EXCEL Risk-Register tracker only.
# The Word-document renderer uses a different palette
# (`SEVERITY_CELL_PALETTE` in services/docx_generator.py) per the
# team's house style. The two are intentionally decoupled — Excel
# trackers stay on the original red/amber/green so importers + the
# VibeDocs master tracker template keep matching shades, while the
# Word deliverable can iterate on its visual style independently.
SEVERITY_FILL_HEX = {
    "Critical":      "C00000",  # dark red
    "High":          "FF0000",  # red
    "Medium":        "FFC000",  # amber
    "Low":           "92D050",  # green
    "Informational": "00B0F0",  # light blue
    "Info":          "00B0F0",
}
# White text contrasts well on Critical / High; Medium / Low / Info
# fills are bright enough that black text reads more cleanly.
# (`#00B0F0` light-blue against white text was hard to read on
# screen + on printouts — black text matches the team's sample.)
SEVERITY_FONT_HEX = {
    "Critical":      "FFFFFF",
    "High":          "FFFFFF",
    "Medium":        "000000",
    "Low":           "000000",
    "Informational": "000000",
    "Info":          "000000",
}


def _normalize_severity_label(raw: Any) -> Optional[str]:
    """Return the canonical Severity-enum spelling for the input value
    so the colour / display lookup tables hit. Returns None when the
    input is None / empty / unrecognised."""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    key = s.lower()
    aliases = {
        "critical": "Critical", "crit": "Critical",
        "high": "High", "h": "High",
        "medium": "Medium", "med": "Medium", "moderate": "Medium",
        "low": "Low", "l": "Low",
        "info": "Informational", "informational": "Informational",
        "information": "Informational", "none": "Informational",
    }
    return aliases.get(key, s.title())


STATUS_FILL_HEX = {
    "open":          "FFC7CE",  # light red
    "closed":        "C6EFCE",  # light green
    "resolved":      "C6EFCE",
    "risk accepted": "FFEB9C",  # light yellow
    "in progress":   "DDEBF7",  # light blue
}
STATUS_FONT_HEX = {
    "open":          "9C0006",  # dark red
    "closed":        "276221",  # dark green
    "resolved":      "276221",
    "risk accepted": "9C6500",  # dark amber
    "in progress":   "2E75B6",  # dark blue
}


def _apply_status_styling(ws, col_idx: int, row: int, raw_value: Any) -> None:
    """Write the Status value and paint the cell with a traffic-light colour.
    Applied consistently to every row so the output looks the same regardless
    of whether the row was within the template's original data block or was
    appended beyond it.
    """
    if raw_value is None:
        return
    text = str(raw_value).strip()
    if not text:
        return
    cell = ws.cell(row=row, column=col_idx)
    cell.value = text
    key = text.lower()
    fill_hex = STATUS_FILL_HEX.get(key)
    font_hex = STATUS_FONT_HEX.get(key)
    from openpyxl.styles import PatternFill, Font
    from copy import copy as _copy_style
    if fill_hex:
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex,
                                fill_type="solid")
        try:
            base = _copy_style(cell.font)
            base.color = font_hex
            cell.font = base
        except Exception:
            cell.font = Font(color=font_hex)
    else:
        # Unmapped status (e.g. "NA" for informational findings) — CLEAR any
        # red "Bad" example fill/font the template shipped with so the cell
        # renders as a normal colourless cell, not a red one.
        cell.fill = PatternFill(fill_type=None)
        try:
            base = _copy_style(cell.font)
            base.color = "FF000000"   # default black (ARGB)
            cell.font = base
        except Exception:
            cell.font = Font(color="FF000000")
    cell.alignment = _CENTER_ALIGN


def _apply_severity_styling(ws, col_idx: int, row: int, raw_value: Any) -> None:
    """Write the severity text + paint the cell using the team palette.
    Safe to call with raw_value=None — leaves the cell untouched in
    that case so we don't blow away the template's example styling.
    """
    canonical = _normalize_severity_label(raw_value)
    if canonical is None:
        return
    cell = ws.cell(row=row, column=col_idx)
    cell.value = SEVERITY_DISPLAY.get(canonical, canonical)
    fill_hex = SEVERITY_FILL_HEX.get(canonical)
    font_hex = SEVERITY_FONT_HEX.get(canonical)
    if fill_hex:
        # Local import keeps the module light when no export is happening.
        from openpyxl.styles import PatternFill, Font
        from copy import copy as _copy_style
        cell.fill = PatternFill(start_color=fill_hex, end_color=fill_hex,
                                 fill_type="solid")
        # Preserve existing font attributes (bold, name, size, alignment)
        # while just overriding the colour. Cells in the VibeDocs template
        # ship with a specific font; we want to keep that and only change
        # the colour for contrast.
        try:
            base = _copy_style(cell.font)
            base.color = font_hex
            cell.font = base
        except Exception:                                 # pragma: no cover
            cell.font = Font(color=font_hex)
    cell.alignment = _CENTER_ALIGN


# Standard column layout for the synthesised tracker if no template exists.
DEFAULT_EXPORT_COLUMNS = [
    ("#",              "index"),
    ("Date Raised",    "date_raised"),
    ("Title",          "title"),
    ("Severity",       "severity"),
    ("CVSS Score",     "cvss_score"),
    ("CVSS Vector",    "cvss_vector"),
    ("Status",         "status"),
    ("Affected Asset", "affected_asset"),
    ("Description",    "description"),
    ("Impact",         "impact"),
    ("Steps to Reproduce", "poc_steps"),
    ("Remediation",    "remediation"),
    ("References",     "references"),
    ("CWE ID",         "cwe"),
    # VibeDocs master tracker template uses "OWASP Top 10" as the
    # column header; the importer also accepts "OWASP Category" /
    # "OWASP 2021" / "OWASP 2025" via the alias map above, so existing
    # trackers keep importing cleanly.
    ("OWASP Top 10",   "owasp_category"),
    ("Retest Notes",   "retest_notes"),
    ("Client Statement", "client_statement"),
]


# ── Info-sheet population ──────────────────────────────────────────────────

def _looks_like_ip_or_cidr(s: str) -> bool:
    """Return True if s is a valid IPv4/IPv6 address or CIDR notation."""
    try:
        ipaddress.ip_network(s, strict=False)
        return True
    except ValueError:
        return False


def _classify_scope_targets(targets: list) -> tuple[list[str], list[str]]:
    """Split a project's scope_targets into (urls, ips).

    Only strings that parse as a valid IPv4/IPv6 address or CIDR range go into
    ``ips``; everything else (http/https URLs, hostnames with or without dots,
    bare hostname tokens like "azure-fullstack-prod") goes into ``urls``.
    """
    urls: list[str] = []
    ips: list[str] = []
    for t in (targets or []):
        s = str(t).strip()
        if not s:
            continue
        if _looks_like_ip_or_cidr(s):
            ips.append(s)
        else:
            urls.append(s)
    return urls, ips


def _info_write(ws, row: int, col: int, value) -> None:
    """Write `value` to (row, col) only when the cell is currently blank."""
    if value is None or value == "":
        return
    cell = ws.cell(row=row, column=col)
    if cell.value in (None, ""):
        cell.value = _no_formula(value)


def _info_force_write(ws, row: int, col: int, value) -> None:
    """Write `value` to (row, col), overwriting any existing content.

    Formula-injection safe: values that begin with a formula-trigger char
    (`=`, `+`, `-`, `@`, …) — most commonly a phone number like
    "+65 1234 5678" — are stored verbatim with the cell's quote-prefix
    flag set, so Excel renders them as plain text WITHOUT showing a leading
    apostrophe and never evaluates them. Falls back to the literal-apostrophe
    guard if the openpyxl build doesn't expose quotePrefix.
    """
    if value is None or value == "":
        return
    cell = ws.cell(row=row, column=col)
    if isinstance(value, str) and value and value[0] in _FORMULA_TRIGGERS:
        try:
            cell.value = value
            cell.quotePrefix = True
        except Exception:                                  # pragma: no cover
            cell.value = "'" + value
    else:
        cell.value = value


def populate_info_sheet(wb, info_context: dict) -> None:
    """Auto-populate the Info sheet of *wb* with engagement details.

    The function locates the first sheet whose name contains "info"
    (case-insensitive), then scans rows 1-100 looking for known label
    keywords in column B. When a match is found it writes the
    corresponding context value into column C (simple fields) or the
    correct sub-columns for tabular sections.

    Tabular sections handled:
        Client PIC         → col C=Name, D=Email
        DT Tester (1)      → col C=Name, D=Email, E=Contact No
        DT Tester 2        → col C=Name, D=Email, E=Contact No (retest only)
        Login Credentials  → col C=Role, D=Username, E=Password

    Only blank cells are overwritten.

    ``info_context`` keys (all optional, missing/empty → skipped):
        system_name         – project name (System Name / ID)
        scope_description   – project scope description
        urls                – list[str] of URL scope targets
        ips                 – list[str] of IP scope targets
        testing_start       – str, formatted start date
        testing_end         – str, formatted end date
        date_of_retest      – str, formatted retest date
        client_pic_name     – Client PIC contact name
        client_pic_email    – Client PIC contact email
        tester_name         – DT Tester name(s)
        tester_email        – DT Tester email address
        retest_tester_name  – DT Tester 2 name (retest only)
        retest_tester_email – DT Tester 2 email (retest only)
        source_ip           – Source IP address of tester machine
        login_credentials   – list of {role, username, password} dicts
    """
    info_ws = next(
        (wb[s] for s in wb.sheetnames if "info" in s.lower()),
        None,
    )
    if info_ws is None:
        return

    urls: list[str] = list(info_context.get("urls") or [])
    ips: list[str] = list(info_context.get("ips") or [])
    # AWS / cloud account IDs (Cloud VAPT). On a cloud tracker the "Target IP
    # Address" row is where the scanned account IDs belong, so fold them into
    # the IP target list when no real IPs were supplied.
    aws_accounts: list[str] = [str(a).strip() for a in
                               (info_context.get("aws_account_ids") or []) if str(a).strip()]
    if aws_accounts and not ips:
        ips = aws_accounts
    url_str = "\n".join(urls)
    ip_str  = "\n".join(ips)
    aws_str = "\n".join(aws_accounts)

    system_name          = info_context.get("system_name") or ""
    scope_desc           = info_context.get("scope_description") or ""
    testing_start        = info_context.get("testing_start") or ""
    testing_end          = info_context.get("testing_end") or ""
    date_of_retest       = info_context.get("date_of_retest") or ""
    client_pic_name      = info_context.get("client_pic_name") or ""
    client_pic_email     = info_context.get("client_pic_email") or ""
    tester_name          = info_context.get("tester_name") or ""
    tester_email         = info_context.get("tester_email") or ""
    tester_phone         = info_context.get("tester_phone") or ""
    retest_tester_name   = info_context.get("retest_tester_name") or ""
    retest_tester_email  = info_context.get("retest_tester_email") or ""
    retest_tester_phone  = info_context.get("retest_tester_phone") or ""
    source_ip            = info_context.get("source_ip") or ""
    login_credentials    = list(info_context.get("login_credentials") or [])

    # State flags for tabular sections.
    # `_cred_idx` tracks which credential row to write next (-1 = not in section).
    next_is_client_pic    = False
    next_is_dt_tester     = False
    next_is_dt_tester_2   = False
    cred_idx              = -1   # index into login_credentials list

    for row_idx in range(1, 101):
        label_cell = info_ws.cell(row=row_idx, column=2)   # col B
        label_val  = label_cell.value

        # --- Tabular data rows: col B holds a row-number integer ---
        if isinstance(label_val, (int, float)) and not isinstance(label_val, bool):
            if next_is_client_pic:
                _info_force_write(info_ws, row_idx, 3, client_pic_name)
                _info_force_write(info_ws, row_idx, 4, client_pic_email)
                next_is_client_pic = False
            elif next_is_dt_tester:
                _info_force_write(info_ws, row_idx, 3, tester_name)
                _info_force_write(info_ws, row_idx, 4, tester_email)
                _info_force_write(info_ws, row_idx, 5, tester_phone)   # Contact No
                next_is_dt_tester = False
            elif next_is_dt_tester_2:
                _info_force_write(info_ws, row_idx, 3, retest_tester_name)
                _info_force_write(info_ws, row_idx, 4, retest_tester_email)
                _info_force_write(info_ws, row_idx, 5, retest_tester_phone)  # Contact No
                next_is_dt_tester_2 = False
            elif cred_idx >= 0:
                # Write the next login credential row (cols C=Role, D=Username, E=Password)
                if cred_idx < len(login_credentials):
                    cred = login_credentials[cred_idx]
                    _info_force_write(info_ws, row_idx, 3, cred.get("role") or "")
                    _info_force_write(info_ws, row_idx, 4, cred.get("username") or "")
                    _info_force_write(info_ws, row_idx, 5, cred.get("password") or "")
                    cred_idx += 1
                # Keep cred_idx active — subsequent integer rows get next credentials
            continue

        # --- Empty label (blank col B) ---
        label_lc = str(label_val or "").strip().lower()
        if not label_lc:
            # A pending Client PIC / DT Tester section whose data row carries
            # a BLANK col B (some templates number these rows, e.g. "1"; the
            # Web/2025 tracker leaves col B empty). Flush the write here on the
            # first blank row after the header, then clear the flag — otherwise
            # the tester name/email/phone never get written at all.
            if next_is_client_pic:
                _info_force_write(info_ws, row_idx, 3, client_pic_name)
                _info_force_write(info_ws, row_idx, 4, client_pic_email)
                next_is_client_pic = False
                continue
            if next_is_dt_tester:
                _info_force_write(info_ws, row_idx, 3, tester_name)
                _info_force_write(info_ws, row_idx, 4, tester_email)
                _info_force_write(info_ws, row_idx, 5, tester_phone)
                next_is_dt_tester = False
                continue
            if next_is_dt_tester_2:
                _info_force_write(info_ws, row_idx, 3, retest_tester_name)
                _info_force_write(info_ws, row_idx, 4, retest_tester_email)
                _info_force_write(info_ws, row_idx, 5, retest_tester_phone)
                next_is_dt_tester_2 = False
                continue
            # Blank col B can mean a merged-cell continuation row inside the
            # Login Credentials section — keep writing credentials rather than
            # resetting the counter. Only reset cred_idx when a new named
            # label appears (handled below after this block).
            if cred_idx >= 0 and cred_idx < len(login_credentials):
                cred = login_credentials[cred_idx]
                _info_force_write(info_ws, row_idx, 3, cred.get("role") or "")
                _info_force_write(info_ws, row_idx, 4, cred.get("username") or "")
                _info_force_write(info_ws, row_idx, 5, cred.get("password") or "")
                cred_idx += 1
            continue

        # Reset cred_idx when a new label row is hit (non-credential section)
        cred_idx = -1

        # --- Keyword matching for simple (single-cell) fields ---
        if "system name" in label_lc or "system id" in label_lc:
            _info_force_write(info_ws, row_idx, 3, system_name)
        elif label_lc == "url" or "application url" in label_lc:
            _info_force_write(info_ws, row_idx, 3, url_str)
        elif ("aws account" in label_lc or "account id" in label_lc
              or "cloud account" in label_lc or "subscription" in label_lc):
            _info_force_write(info_ws, row_idx, 3, aws_str or ip_str)
        elif "target ip" in label_lc:
            _info_force_write(info_ws, row_idx, 3, ip_str)
        elif "application description" in label_lc:
            _info_force_write(info_ws, row_idx, 3, scope_desc)
        elif "application usage" in label_lc:
            _info_force_write(info_ws, row_idx, 3, scope_desc)
        elif "start date" in label_lc:
            _info_force_write(info_ws, row_idx, 3, testing_start)
        elif "end date" in label_lc:
            _info_force_write(info_ws, row_idx, 3, testing_end)
        elif "date of retest" in label_lc:
            _info_force_write(info_ws, row_idx, 3, date_of_retest)
        elif "source ip" in label_lc or "source address" in label_lc:
            _info_force_write(info_ws, row_idx, 3, source_ip)
        elif "login credential" in label_lc or "tested account" in label_lc or label_lc == "credentials":
            # Start credential table section — integer rows that follow get cred data
            cred_idx = 0
        elif "client pic" in label_lc:
            next_is_client_pic = True
        elif "dt tester 2" in label_lc or "dt tester2" in label_lc or "retest tester" in label_lc:
            next_is_dt_tester_2 = True
        elif "dt tester" in label_lc:
            next_is_dt_tester = True


# ── Test Exe. Checklist sheet population ──────────────────────────────────

# Regex to extract the category prefix from an Objective cell, e.g.:
#   "API1:2023 - Broken Object Level Authorization"  → "API1"
#   "M1:2024 -  Improper Credential Usage"           → "M1"
#   "DA1:2021 - Injections"                          → "DA1"
#   "A01:2025 - Broken Access Control"               → "A01"
#   "A01:2021 - Broken Access Control"               → "A01"
_CHECKLIST_CAT_RE = re.compile(
    r"^((?:API|DA|M|A)\d{1,2})(?::\d{4})?",
    re.IGNORECASE,
)

# Same pattern used to extract the prefix from a finding's owasp_category
# field (e.g. "API4:2023" → "API4", "M3:2024" → "M3", "A01:2025" → "A01").
_OWASP_PREFIX_RE = re.compile(
    r"^((?:API|DA|M|A)0*\d{1,2})(?::\d{4})?",
    re.IGNORECASE,
)


def _owasp_prefix(category: str) -> str:
    """Extract the numeric prefix from an OWASP category string.

    Returns a normalised uppercase string, e.g.:
        "API1:2023"  → "API1"
        "M3:2024"    → "M3"
        "DA1:2021"   → "DA1"
        "A01:2025"   → "A01"
        "A01:2021"   → "A01"
    Returns "" for unrecognised/empty values.
    """
    if not category:
        return ""
    m = _OWASP_PREFIX_RE.match(category.strip())
    if not m:
        return ""
    return m.group(1).upper()


def populate_checklist_sheet(wb, covered_categories: set[str], findings: list[dict] | None = None) -> None:
    """Auto-populate the 'Test Exe. Checklist - *' sheet in *wb*.

    EVERY checklist row that names an OWASP objective is filled so the
    exported tracker is complete — no blank cells the consultant has to
    fill by hand:
        * Status       → "Covered"          (always)
        * Why not?     → "NA"               (always)
        * Evidence     → "NA"               (always)
    The Results / Observations pair then depends on whether that objective's
    prefix (e.g. "API1", "M1", "DA1", "A01") matched any finding:
        * matched  → Results "Vulnerabilities Identified",
                     Observations "Refer to Risk Register SN X: Title"
                     (one line per finding)
        * unmatched→ Results "No Vulnerability Identified",
                     Observations "No vulnerabilities found."

    All written cells are centre-aligned and inherit the font colour from the
    first template example row.

    *covered_categories* — set of normalised OWASP prefixes, e.g. {"API1"}.
                           May be empty, in which case every objective row is
                           filled as "No Vulnerability Identified".
    *findings*           — list of finding dicts (same structure as the Risk
                           Register rows); used to build per-finding Observations
                           references keyed by OWASP category prefix.
    """
    covered_categories = covered_categories or set()

    from copy import copy as _cp
    from openpyxl.styles import Alignment

    # Find the first sheet whose name contains "checklist" (case-insensitive).
    cl_ws = next(
        (wb[s] for s in wb.sheetnames if "checklist" in s.lower()),
        None,
    )
    if cl_ws is None:
        return

    # Locate the header row by scanning for "Objective" in column C (col 3).
    header_row = 0
    for row_idx in range(1, 30):
        cell_val = str(cl_ws.cell(row=row_idx, column=3).value or "").strip().lower()
        if cell_val == "objective":
            header_row = row_idx
            break
    if header_row == 0:
        return

    # Detect column indices dynamically from the header row.
    status_col:   int       = 7    # default col G
    why_col:      int       = 8    # default col H
    results_col:  int | None = None
    obs_col:      int | None = None
    evidence_col: int       = 10   # default col J

    for col_idx in range(1, 25):
        hdr = str(cl_ws.cell(row=header_row, column=col_idx).value or "").strip().lower()
        if not hdr:
            continue
        # NOTE: "Evidence to justify why it is not covered" contains BOTH
        # "evidence" AND "not covered", so it must be tested for "evidence"
        # BEFORE the why/not-covered branch — otherwise the evidence header
        # is misclassified as the why column and the real
        # "If it is not covered, why?" column (H) never gets written.
        if "status" in hdr:
            status_col = col_idx
        elif "evidence" in hdr:
            evidence_col = col_idx
        elif "not covered" in hdr or ("why" in hdr and "not" in hdr):
            why_col = col_idx
        elif "result" in hdr and "observation" not in hdr:
            results_col = col_idx
        elif "observation" in hdr:
            obs_col = col_idx

    # Build OWASP prefix → list of (sn, title) from findings for Observations.
    prefix_findings: dict[str, list[tuple[int, str]]] = {}
    if findings:
        for fd in findings:
            owasp  = fd.get("owasp_category") or ""
            prefix = _owasp_prefix(owasp)
            if prefix:
                sn    = fd.get("index", 0)
                title = fd.get("title", "")
                prefix_findings.setdefault(prefix, []).append((sn, title))

    # Read font style from the first example data row so covered rows inherit
    # the template's intended colours.
    first_data_row = header_row + 1
    example_fonts: dict[int, Any] = {}
    for col_idx in filter(None, [status_col, why_col, results_col, obs_col, evidence_col]):
        cell = cl_ws.cell(row=first_data_row, column=col_idx)
        if cell.font:
            example_fonts[col_idx] = _cp(cell.font)

    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _write_cell(row_i: int, col_i: int | None, value: str) -> None:
        """Write value with template colour and centre alignment."""
        if col_i is None:
            return
        c = cl_ws.cell(row=row_i, column=col_i)
        c.value = value
        c.alignment = _cp(center_align)
        if col_i in example_fonts:
            c.font = _cp(example_fonts[col_i])

    # Scan data rows below the header. 200 rows covers any realistic checklist.
    for row_idx in range(header_row + 1, header_row + 200):
        obj_val = str(cl_ws.cell(row=row_idx, column=3).value or "").strip()
        if not obj_val:
            continue

        m = _CHECKLIST_CAT_RE.match(obj_val)
        if not m:
            continue

        row_prefix = m.group(1).upper()
        is_covered = row_prefix in covered_categories

        # First three columns are standardised on EVERY objective row,
        # regardless of whether a finding mapped to it: the control area
        # was in scope / tested ("Covered"), so the not-covered rationale
        # and its evidence are "NA".
        _write_cell(row_idx, status_col,   "Covered")
        _write_cell(row_idx, why_col,      "NA")
        _write_cell(row_idx, evidence_col, "NA")

        if is_covered:
            if results_col:
                _write_cell(row_idx, results_col, "Vulnerabilities Identified")
            if obs_col:
                refs = prefix_findings.get(row_prefix, [])
                obs_text = (
                    "\n".join(
                        f"Refer to Risk Register SN {sn}: {title}"
                        for sn, title in refs
                    ) if refs else "Refer to Risk Register"
                )
                _write_cell(row_idx, obs_col, obs_text)
        else:
            # Objective tested but nothing found — say so explicitly rather
            # than leaving the row blank.
            if results_col:
                _write_cell(row_idx, results_col, "No Vulnerability Identified")
            if obs_col:
                _write_cell(row_idx, obs_col, "No vulnerabilities found.")


# ── Risk register write ────────────────────────────────────────────────────


def _apply_uniform_borders(ws, first_row: int, last_row: int,
                           n_cols: int) -> None:
    """Force a thin black border on all four sides of every cell in the
    rectangular region [first_row..last_row] x [1..n_cols].

    VibeDocs tracker templates ship their example data rows with slightly
    inconsistent borders — some cells are missing a bottom edge, so a long
    findings block renders with gaps in the gridlines (see the WAPT tracker
    where some recommendation/observation cells had no bottom border). The
    style-copy that propagates the template row to appended rows faithfully
    carries those gaps forward. This pass normalises every populated data
    cell to a uniform black box so the table reads cleanly. Fills, fonts and
    values are untouched — only the border is overwritten.
    """
    if last_row < first_row or n_cols < 1:
        return
    from openpyxl.styles import Border, Side
    side = Side(style="thin", color="FF000000")
    box = Border(left=side, right=side, top=side, bottom=side)
    for r in range(first_row, last_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).border = box


def write_risk_register(
        findings: Iterable[dict[str, Any]],
        *,
        template_path: Optional[Path] = None,
        output_path: Path,
        sn_col_index: Optional[int] = None,
        info_context: Optional[dict] = None,
        covered_categories: Optional[set] = None,
        group_sheets: Optional[list[dict]] = None,
        group_row_defaults: Optional[dict] = None,
        target_sheet: Optional[str] = None,
        set_observation_pointer: bool = True,
        rmm_enabled: bool = True,
) -> Path:
    """Render `findings` into an XLSX file at `output_path`.

    Two modes:

    1. **Template mode** (`template_path` set + file exists)
       The template is *copied* to `output_path` so EVERY sheet —
       Risk Register, Info, OWASP-CWE-SANS, Scoping Questionnaire,
       Test Exe. Checklist, Summary, etc. — and EVERY cell's
       formatting (colour, font, border, conditional rules, column
       widths, merged ranges, named styles) is preserved verbatim.

       We then walk the Risk Register sheet:
         * Find the header row (row 2 on VibeDocs trackers).
         * Find the S/N anchor column (auto-detected if
           `sn_col_index` is None).
         * For each finding, overwrite the cells of the next data row
           with the finding's values — using only `cell.value = …` so
           the style attached to the cell stays intact.
         * If the template shipped with *more* example data rows than
           findings we have, clear the *values* of the leftover rows
           but leave the styling alone.
         * If we have *more* findings than the template's data block,
           openpyxl appends new rows; we copy the immediately-preceding
           data row's style onto each new cell so the pattern continues
           visually.

       No `ws.delete_rows()` is ever called — that's what previously
       wrecked colours and conditional formatting.

    2. **Synthesise mode** (no template)
       Fall back to a single-sheet workbook using
       `DEFAULT_EXPORT_COLUMNS`.
    """
    from openpyxl import Workbook, load_workbook
    from copy import copy as _copy_style
    output_path.parent.mkdir(parents=True, exist_ok=True)

    if not template_path or not Path(template_path).exists():
        # ---------- Synthesise mode ----------
        wb = Workbook()
        ws = wb.active
        ws.title = "Risk Register"
        _write_header_row(ws, [c[0] for c in DEFAULT_EXPORT_COLUMNS])
        _write_data_rows(ws, 2,
                          [c[0] for c in DEFAULT_EXPORT_COLUMNS],
                          findings,
                          col_keys=[c[1] for c in DEFAULT_EXPORT_COLUMNS])
        wb.save(str(output_path))
        return output_path

    # ---------- Template mode ----------
    shutil.copyfile(str(template_path), str(output_path))
    wb = load_workbook(str(output_path))

    # Convert to list once — needed for Info/Checklist population and the
    # Risk Register write below.
    findings_list = list(findings)

    # Auto-populate the Info sheet before touching the Risk Register so
    # engagement details (dates, tester, scope) are written into the
    # correct tracker sheet on every export — no manual copy-paste needed.
    if info_context:
        try:
            populate_info_sheet(wb, info_context)
        except Exception as _ie:
            log.warning("Info sheet population failed — skipping: %s", _ie)

    # Always run — even with zero covered categories — so a tracker that
    # has a checklist sheet gets every objective row filled ("No
    # Vulnerability Identified" where nothing matched). populate_checklist_sheet
    # no-ops on trackers that don't carry a checklist sheet.
    try:
        populate_checklist_sheet(wb, covered_categories or set(), findings=findings_list)
    except Exception as _ce:
        log.warning("Checklist sheet population failed — skipping: %s", _ce)

    sheet_name = (target_sheet if (target_sheet and target_sheet in wb.sheetnames)
                  else _pick_sheet(wb.sheetnames))

    if not sheet_name:
        # Template has no Risk Register sheet — add one and write fresh.
        ws = wb.create_sheet("Risk Register")
        _write_header_row(ws, [c[0] for c in DEFAULT_EXPORT_COLUMNS])
        _write_data_rows(ws, 2,
                          [c[0] for c in DEFAULT_EXPORT_COLUMNS],
                          findings_list,
                          col_keys=[c[1] for c in DEFAULT_EXPORT_COLUMNS])
        wb.save(str(output_path)); return output_path

    ws = wb[sheet_name]
    header_row, headers = _find_header_row(ws)
    if header_row == 0:
        # No detectable header — write with defaults from row 1.
        _write_header_row(ws, [c[0] for c in DEFAULT_EXPORT_COLUMNS])
        _write_data_rows(ws, 2,
                          [c[0] for c in DEFAULT_EXPORT_COLUMNS],
                          findings_list,
                          col_keys=[c[1] for c in DEFAULT_EXPORT_COLUMNS])
        wb.save(str(output_path)); return output_path

    # Map field -> column index (0-based) using the same aliases as
    # the parser — this guarantees round-trip symmetry.
    col_map = _map_columns(headers)
    sn_col = sn_col_index if sn_col_index is not None else col_map.pop("_sn", None)
    col_map.pop("_sn", None)
    n_cols = len(headers)

    # Build the inverse: column_index -> field key (so we can iterate
    # left-to-right across the row).
    inv: dict[int, str] = {v: k for k, v in col_map.items()}

    # Detect how far the original data block goes by walking down from
    # header_row+1 until we hit a fully-blank row (or the natural end
    # of the sheet). This is the region whose values we'll overwrite
    # or clear — outside it we don't touch anything.
    data_start = header_row + 1
    last_data_row = data_start - 1
    for r in range(data_start, (ws.max_row or data_start) + 1):
        row_blank = True
        for c in range(1, min(n_cols, ws.max_column) + 1):
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                row_blank = False
                break
        if row_blank:
            break
        last_data_row = r

    # The "style template row" — when we have to append rows beyond
    # the original data block we use this row's cell styles as the
    # pattern so the table keeps a consistent look.
    style_template_row = data_start if last_data_row < data_start else data_start
    style_cells = [
        ws.cell(row=style_template_row, column=c) for c in range(1, n_cols + 1)
    ]

    # Write each finding. Overwrite cell.value only — keep cell.style
    # intact so colours / borders / number formats survive.
    for offset, f in enumerate(findings_list):
        target_row = data_start + offset
        # Append-with-style for rows beyond the original template block.
        if target_row > last_data_row:
            for c_idx, style_cell in enumerate(style_cells, start=1):
                new_cell = ws.cell(row=target_row, column=c_idx)
                if style_cell.has_style:
                    new_cell.font          = _copy_style(style_cell.font)
                    new_cell.fill          = _copy_style(style_cell.fill)
                    new_cell.border        = _copy_style(style_cell.border)
                    new_cell.alignment     = _copy_style(style_cell.alignment)
                    new_cell.protection    = _copy_style(style_cell.protection)
                    new_cell.number_format = style_cell.number_format
        # S/N column — auto-fill with 1-based index if we know which column.
        if sn_col is not None:
            ws.cell(row=target_row, column=sn_col + 1).value = offset + 1
        # Walk every column; only write the ones we have a mapping for.
        # `severity_written` is set when the severity column gets the
        # palette-paint treatment below; that branch handles its own
        # cell.value assignment so we skip the generic value-write.
        severity_col_idx = col_map.get("severity")
        screenshot_col_idx = col_map.get("_screenshot")
        # Accumulate base64 images pasted into rich-text fields (description,
        # impact, remediation etc.) so they get embedded in the Screenshot
        # column alongside the uploaded screenshots rather than being dropped.
        _inline_images_for_row: list[dict] = []
        for c_idx in range(1, n_cols + 1):
            field = inv.get(c_idx - 1)
            if not field or field == "_sn":
                continue
            # Severity gets special handling — colour fill + INFO display.
            if field == "severity":
                val = f.get(field) if isinstance(f, dict) else getattr(f, field, None)
                _apply_severity_styling(ws, c_idx, target_row, val)
                continue
            # Status gets traffic-light colour coding (Open=red, Closed=green…)
            # applied consistently on every row regardless of template styling.
            if field == "status":
                val = f.get(field) if isinstance(f, dict) else getattr(f, field, None)
                if hasattr(val, "value"):
                    val = val.value
                _apply_status_styling(ws, c_idx, target_row, val)
                continue
            # Screenshot / retest-screenshot columns — defer to the
            # image-embedding pass below.  Don't blank or text-write the
            # cell here so the embedded images land cleanly into a cell
            # that retains its template styling (border, fill).
            if field in ("_screenshot", "_retest_screenshot"):
                continue
            val = f.get(field) if isinstance(f, dict) else getattr(f, field, None)
            if val is None:
                # Don't blank a cell the template may have pre-filled —
                # only overwrite when we have a real value.
                # But also don't carry over the previous finding's value
                # (which would happen if a template data row already had
                # text in this cell). So write "" to clear text fields.
                ws.cell(row=target_row, column=c_idx).value = None
                continue
            if hasattr(val, "value"):
                val = val.value
            # Strip Quill HTML from rich-text fields before writing to Excel.
            # Without this, base64-encoded data-URL images (pasted screenshots)
            # spill into the cell as thousands of raw bytes, making the cell
            # unreadable.  We also extract any embedded images so they can be
            # embedded into the Screenshot column for this finding.
            if isinstance(val, str) and "<" in val:
                # Harvest pasted images from the field BEFORE stripping —
                # they'll be appended to the finding's screenshot list so they
                # end up in the Screenshot column alongside any uploaded shots.
                _field_imgs = _extract_base64_images(val)
                if _field_imgs and screenshot_col_idx is not None:
                    _inline_images_for_row.extend(
                        {"path": None, "_bytes": blob, "caption": f"{field}: pasted image"}
                        for blob in _field_imgs
                    )
                val = _strip_html_to_text(val)
            cell = ws.cell(row=target_row, column=c_idx)
            cell.value = val
            if _CENTER_ALIGN is not None:
                cell.alignment = _CENTER_ALIGN

        # ---- Embed finding screenshots into the Screenshot column ----
        # The finding dict may carry a `screenshots` key whose value is
        # a list of `{path, caption}` entries (or legacy plain strings).
        # We also include any base64 images pasted into rich-text fields
        # (description, impact, etc.) that were harvested above so pasted
        # screenshots render as actual images rather than raw bytes text.
        screenshots = (f.get("screenshots") if isinstance(f, dict)
                       else getattr(f, "screenshots", None)) or []
        all_screenshots = list(screenshots) + _inline_images_for_row
        if screenshot_col_idx is not None and all_screenshots:
            _embed_finding_screenshots(
                ws, target_row, screenshot_col_idx + 1, all_screenshots
            )

        # ---- Embed retest screenshots into Post Review Screenshot col ----
        # Populated for retest / report-update versions. Uses the same
        # compositor as primary screenshots so captions are baked in.
        retest_screenshot_col_idx = col_map.get("_retest_screenshot")
        retest_shots = (f.get("retest_screenshots") if isinstance(f, dict)
                        else getattr(f, "retest_evidence", None)) or []
        if retest_screenshot_col_idx is not None and retest_shots:
            _embed_finding_screenshots(
                ws, target_row, retest_screenshot_col_idx + 1, retest_shots
            )

    # Clear leftover template rows beyond our findings count — values
    # only, styling preserved.
    rows_written = len(findings_list)
    for r in range(data_start + rows_written, last_data_row + 1):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).value = None

    # Normalise every populated finding cell to a uniform thin black border.
    # Template example rows ship with inconsistent borders (some cells lack a
    # bottom edge); without this the gridlines render with gaps on long blocks.
    if rows_written:
        _apply_uniform_borders(
            ws, data_start, data_start + rows_written - 1, n_cols)

    # Infra grouped-findings: one extra sheet per group (e.g. "Outdated
    # patches", "SSL misconfigs"), each populated with the individual
    # per-host findings from that group's xlsx in the same tracker format.
    if group_sheets:
        try:
            add_grouped_findings_sheets(
                wb, sheet_name, group_sheets, defaults=group_row_defaults or {},
                set_observation_pointer=set_observation_pointer)
        except Exception as _ge:
            log.warning("Grouped-findings sheets skipped: %s", _ge)

    # GovTech CSG ICT RMM disabled -> drop the RMM Risk/Impact/Likelihood Rating
    # columns from every sheet (mirrors the Word report's RMM strip).
    if not rmm_enabled:
        try:
            _strip_rmm_columns(wb)
        except Exception as _re:
            log.warning("RMM column strip skipped: %s", _re)

    wb.save(str(output_path))
    return output_path


def _strip_rmm_columns(wb) -> None:
    """Delete the GovTech CSG ICT RMM columns ("… RMM Risk Rating",
    "… RMM Impact Rating", "… RMM Likelihood Rating") from every sheet whose
    header row contains them. Used when the report has the RMM methodology
    disabled."""
    for ws in wb.worksheets:
        # Locate the column-header row (VibeDocs trackers use row 2; the "S/N"
        # cell is a reliable anchor). Fall back to any row that mentions "rmm".
        hdr_row = 0
        max_c = ws.max_column or 0
        for r in range(1, min((ws.max_row or 1), 6) + 1):
            vals = [str(ws.cell(row=r, column=c).value or "").strip().lower()
                    for c in range(1, max_c + 1)]
            if "s/n" in vals or any("rmm" in v for v in vals):
                hdr_row = r
                break
        if not hdr_row:
            continue
        victims = []
        for c in range(1, max_c + 1):
            h = str(ws.cell(row=hdr_row, column=c).value or "").strip().lower()
            if "rmm" in h and "rating" in h:
                victims.append(c)
        # Delete right-to-left so earlier indices stay valid.
        for c in sorted(victims, reverse=True):
            ws.delete_cols(c, 1)


# ── Infra grouped-findings: per-group detail sheets ────────────────────────

# Attachment-xlsx (per-category) display column -> tracker finding field key.
_ATTACH_COL_TO_FIELD = {
    "finding name": "title", "name": "title", "plugin name": "title",
    "host": "affected_asset", "ip": "affected_asset", "ip address": "affected_asset",
    "port": "port", "service port": "port",
    "protocol": "protocol",
    "risk": "severity", "risk factor": "severity", "severity": "severity",
    "cvss3 score": "cvss_score", "cvss v3 score": "cvss_score",
    "cvss score": "cvss_score", "cvss": "cvss_score",
    "cvss3 vector": "cvss_vector", "cvss vector": "cvss_vector",
    "plugin output": "_observation", "synopsis": "_synopsis",
    "description": "description",
    "solution": "remediation",
    # Cloud (Steampipe/Prowler) per-service attachment — VibeDocs tracker columns.
    "cvss risk rating": "severity",
    "affected resource(s)/instance(s)": "affected_asset",
    "affected resource(s)/instances": "affected_asset",
    "affected resources/instances": "affected_asset",
    "issue title": "title",
    "benchmark": "benchmark",
    "benchmark clause": "benchmark_clauses",
    "benchmark clauses": "benchmark_clauses",
    "observation": "_observation",
    "implication": "impact", "implications": "impact",
    "recommendation": "remediation", "recommendations": "remediation",
}


def _aws_account_from_arn(value: Any) -> str:
    """Extract the 12-digit AWS account ID from an ARN / resource string.
    e.g. 'arn:aws:iam::632605670340:root' -> '632605670340'. Returns '' if none."""
    s = str(value or "")
    m = re.search(r"arn:aws[^:]*:[^:]*:[^:]*:(\d{12})", s)
    if m:
        return m.group(1)
    m = re.search(r"\b(\d{12})\b", s)
    return m.group(1) if m else ""


def _sanitize_sheet_name(name: str, existing: set) -> str:
    """Excel sheet name: <=31 chars, none of []:*?/\\ , unique."""
    s = re.sub(r"[\[\]:\*\?/\\]", " ", str(name or "Group")).strip()
    s = re.sub(r"\s+", " ", s)[:31] or "Group"
    base, n, out = s, 2, s
    while out in existing:
        suffix = f" ({n})"
        out = (base[: 31 - len(suffix)] + suffix)
        n += 1
    return out


def _read_attachment_findings(xlsx_path: Path) -> list[dict]:
    """Read a per-category attachment xlsx into tracker finding dicts."""
    from openpyxl import load_workbook as _lw
    wb = _lw(str(xlsx_path), data_only=True, read_only=True)
    try:
        ws = wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []
    # Header-row detection: the cloud per-service attachment uses a 2-row
    # VibeDocs header (row 0 = group labels "Risk/Findings/…", row 1 = column
    # names, data from row 2). The infra attachment uses a single header row.
    def _hdr(r):
        return [str(h or "").strip().lower() for h in (r or [])]
    hdr0 = _hdr(rows[0])
    hdr1 = _hdr(rows[1]) if len(rows) > 1 else []
    if "s/n" in hdr1 and "s/n" not in hdr0:
        headers = hdr1
        data_rows = rows[2:]
    else:
        headers = hdr0
        data_rows = rows[1:]
    out: list[dict] = []
    for raw in data_rows:
        if raw is None or all(v in (None, "") for v in raw):
            continue
        rec: dict = {}
        for ci, val in enumerate(raw):
            if ci >= len(headers):
                break
            field = _ATTACH_COL_TO_FIELD.get(headers[ci])
            if field and val not in (None, ""):
                rec[field] = val
        if not rec:
            continue
        # Observation = plugin output, falling back to synopsis.
        obs = str(rec.pop("_observation", "") or "").strip()
        syn = str(rec.pop("_synopsis", "") or "").strip()
        observation = obs or syn or str(rec.get("description", "") or "")
        # Normalise severity. Nessus "informational" findings (banners, CPE,
        # BIOS info, …) carry no Risk value — treat any empty/None/Info risk
        # as Informational: no CVSS, Status NA. Real risks keep their score.
        canonical = _normalize_severity_label(rec.get("severity"))
        is_info = canonical is None or canonical == "Informational"
        # AWS account ID from the resource ARN -> System column (cloud).
        acct = _aws_account_from_arn(rec.get("affected_asset"))
        out.append({
            "title": rec.get("title", ""),
            "severity": "Informational" if is_info else canonical,
            "cvss_score": "" if is_info else rec.get("cvss_score", ""),
            "cvss_vector": "" if is_info else rec.get("cvss_vector", ""),
            "affected_asset": rec.get("affected_asset", ""),
            "port": rec.get("port", ""),
            "protocol": rec.get("protocol", ""),
            "description": observation,          # -> Observation column
            "impact": rec.get("impact", ""),
            "remediation": rec.get("remediation", ""),
            "benchmark": rec.get("benchmark", ""),
            "benchmark_clauses": rec.get("benchmark_clauses", ""),
            "system": acct,                      # per-row AWS account ID
            "status": "NA" if is_info else "Open",
            "_is_info": is_info,
        })
    return out


def add_grouped_findings_sheets(wb, base_sheet_name: Optional[str],
                                groups: list[dict], defaults: dict | None = None,
                                set_observation_pointer: bool = True) -> list[str]:
    """For each group ({title, xlsx_path}), copy the Risk Register sheet into a
    new sheet named after the group and populate it with that group's
    per-host findings. Returns the created sheet names.
    """
    from copy import copy as _cp
    defaults = defaults or {}
    if not base_sheet_name or base_sheet_name not in wb.sheetnames:
        return []
    base_ws = wb[base_sheet_name]
    header_row, headers = _find_header_row(base_ws)
    if header_row == 0:
        return []
    n_cols = len(headers)
    col_map = _map_columns(headers)
    sn_col = col_map.pop("_sn", None)
    col_map.pop("_screenshot", None)
    col_map.pop("_retest_screenshot", None)
    created: list[str] = []
    title_to_sheet: dict[str, str] = {}

    for g in groups or []:
        title = (g or {}).get("title")
        xp = (g or {}).get("xlsx_path")
        if not (title and xp and Path(xp).exists()):
            continue
        try:
            gfindings = _read_attachment_findings(Path(xp))
        except Exception as e:
            log.warning("group sheet read failed for %r: %s", title, e)
            continue
        if not gfindings:
            continue

        # Severity override: if the consultant re-rated the grouped finding in
        # VibeDocs, stamp that severity onto EVERY row of this service's detail sheet
        # (and don't treat them as Info any more) so the sheet matches the Risk
        # Register row.
        ov_sev = _normalize_severity_label((g or {}).get("severity"))
        if ov_sev:
            for f in gfindings:
                f["severity"] = ov_sev
                f["_is_info"] = (ov_sev in ("Informational", "Info"))
                if f["_is_info"]:
                    f["status"] = "NA"

        new_ws = wb.copy_worksheet(base_ws)
        new_ws.title = _sanitize_sheet_name(title, set(wb.sheetnames))

        # Determine the template's data block so we can clear example rows.
        last = header_row
        for r in range(header_row + 1, (new_ws.max_row or header_row) + 1):
            if any(new_ws.cell(row=r, column=c).value not in (None, "")
                   for c in range(1, min(n_cols, new_ws.max_column) + 1)):
                last = r
        for r in range(header_row + 1, last + 1):
            for c in range(1, n_cols + 1):
                new_ws.cell(row=r, column=c).value = None

        # Style template row = first data row (keeps borders/fills on appends).
        style_cells = [new_ws.cell(row=header_row + 1, column=c) for c in range(1, n_cols + 1)]
        for offset, f in enumerate(gfindings):
            tr = header_row + 1 + offset
            if tr > last:
                for c_idx, sc in enumerate(style_cells, start=1):
                    nc = new_ws.cell(row=tr, column=c_idx)
                    if sc.has_style:
                        nc.font = _cp(sc.font)
                        nc.fill = _cp(sc.fill)
                        nc.border = _cp(sc.border)
                        nc.alignment = _cp(sc.alignment)
                        nc.number_format = sc.number_format
            if sn_col is not None:
                new_ws.cell(row=tr, column=sn_col + 1).value = offset + 1
            for field, ci in col_map.items():
                if field in ("severity", "status"):
                    continue
                val = f.get(field)
                if val in (None, ""):
                    # Per-row defaults (System / Area of Review / Date Raised…).
                    val = defaults.get(field)
                if val not in (None, ""):
                    new_ws.cell(row=tr, column=ci + 1).value = (
                        _no_formula(val) if isinstance(val, str) else val)
            # NOTE: _apply_*_styling take 1-BASED column indices; col_map is 0-based.
            if "severity" in col_map:
                sci = col_map["severity"] + 1
                _apply_severity_styling(new_ws, sci, tr, f.get("severity"))
                if f.get("_is_info"):
                    # House style: Info findings show "INFO" (matching the
                    # tracker's CVSS Risk Rating drop-down) on the Info-blue fill.
                    new_ws.cell(row=tr, column=sci).value = "INFO"
            if "status" in col_map:
                _apply_status_styling(new_ws, col_map["status"] + 1, tr, f.get("status") or "Open")

        # Clear any leftover red "example-row" fill on the empty rows below the
        # findings (template example rows we blanked but didn't restyle), so the
        # CVSS Risk Rating / Status columns don't show a stray colour block.
        from openpyxl.styles import PatternFill as _PF
        for r in range(header_row + 1 + len(gfindings), last + 1):
            for key in ("severity", "status"):
                if key in col_map:
                    new_ws.cell(row=r, column=col_map[key] + 1).fill = _PF(fill_type=None)

        # Uniform thin black border on every populated finding cell, matching
        # the main Risk Register sheet (closes template gridline gaps).
        if gfindings:
            _apply_uniform_borders(
                new_ws, header_row + 1, header_row + len(gfindings), n_cols)

        created.append(new_ws.title)
        title_to_sheet[str(title)] = new_ws.title

    # Point each grouped finding's Observation in the MAIN Risk Register sheet
    # at its dedicated sheet (infra trackers). The Cloud VAPT tracker instead
    # puts "Refer to <Service> Tab" in the Affected Resource / Benchmark /
    # Benchmark Clauses / Steps columns (done by the caller's dicts), so it
    # passes set_observation_pointer=False to keep the real Observation text.
    title_ci = col_map.get("title")
    obs_ci = col_map.get("description")
    if set_observation_pointer and title_to_sheet and title_ci is not None and obs_ci is not None:
        for r in range(header_row + 1, (base_ws.max_row or header_row) + 1):
            t = str(base_ws.cell(row=r, column=title_ci + 1).value or "").strip()
            sheet_nm = title_to_sheet.get(t)
            if sheet_nm:
                base_ws.cell(row=r, column=obs_ci + 1).value = (
                    f"Please refer to the sheet titled '{sheet_nm}' for the "
                    f"detailed findings.")
    return created


def tidy_unused_rows(xlsx_path: Path) -> int:
    """Strip cell formatting (fill + borders) from the empty rows BELOW the last
    populated data row of every Risk-Register-style sheet.

    The VibeDocs tracker template ships each Risk Register / per-service sheet
    with a block of pre-formatted (bordered) example rows. When a report uses
    fewer rows than the block, the leftover rows keep their borders and render as
    an empty bordered/black block under the findings. This clears those rows back
    to default so the sheet ends cleanly at the last finding.

    Sheets without a recognisable Risk-Register header row (e.g. the Info sheet
    or a pasted IAMActionHunter CSV sheet) are skipped untouched. Returns the
    number of cells cleared.
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Border

    no_fill = PatternFill(fill_type=None)
    no_border = Border()
    try:
        wb = openpyxl.load_workbook(str(xlsx_path))
    except Exception:
        return 0
    cleared = 0
    for ws in wb.worksheets:
        try:
            header_row, headers = _find_header_row(ws)
        except Exception:
            header_row = 0
        if not header_row:
            continue
        n_cols = max(len(headers or []), ws.max_column or 0)
        if n_cols <= 0:
            continue
        max_r = ws.max_row or header_row
        last_data = header_row
        for r in range(header_row + 1, max_r + 1):
            if any(ws.cell(row=r, column=c).value not in (None, "")
                   for c in range(1, n_cols + 1)):
                last_data = r
        for r in range(last_data + 1, max_r + 1):
            for c in range(1, n_cols + 1):
                cell = ws.cell(row=r, column=c)
                if cell.has_style:
                    cell.fill = no_fill
                    cell.border = no_border
                    cleared += 1
    if cleared:
        try:
            wb.save(str(xlsx_path))
        except Exception:
            return 0
    return cleared


def _embed_finding_screenshots(ws, row_idx: int, col_idx_1based: int,
                                screenshots: list) -> None:
    """Embed one or more screenshots into a single cell of the Risk
    Register, vertically stacked, with each caption rendered onto the
    bottom of its own image so caption + image travel as one unit.

    `screenshots` is a list of `{path, caption}` dicts (or legacy bare
    paths). Files that don't exist on disk are silently skipped so the
    export never errors out on missing assets — the consultant has
    already seen warnings in the UI for those.

    Why caption-into-image instead of cell text
    -------------------------------------------
    Previously we wrote all captions as the cell's plain-text value
    above the embedded images. Excel renders the cell text in the
    SAME drawing layer order regardless of caption position, and the
    embedded images sit on top — captions ended up hidden behind the
    screenshots. By compositing each caption into a strip at the
    bottom of its own image via PIL, the caption is part of the image
    pixels themselves and cannot be obscured. As a side effect, the
    cell's text value is cleared.

    Why TwoCellAnchor instead of OneCellAnchor
    ------------------------------------------
    `OneCellAnchor` pins only the top-left corner — Excel treats the
    image as a free-floating object that the user can drag anywhere
    on the sheet. `TwoCellAnchor` with ``editAs="oneCell"`` pins both
    corners to the target cell so sort/filter/row-insert operations
    move the image with its row. This is the closest behaviour to
    "image lives inside the cell" that the .xlsx format supports
    without the Excel 365 IMAGE()-formula API (which openpyxl can't
    emit and which doesn't accept embedded image bytes).
    """
    from io import BytesIO
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.drawing.spreadsheet_drawing import (
        TwoCellAnchor, AnchorMarker)
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.units import pixels_to_EMU
    from PIL import Image as PILImage, ImageDraw, ImageFont

    target_cell = ws.cell(row=row_idx, column=col_idx_1based)
    col_letter = get_column_letter(col_idx_1based)

    # Normalise to list[{path|_bytes, caption}]
    entries: list[dict] = []
    for x in screenshots:
        if isinstance(x, str):
            entries.append({"path": x, "caption": ""})
        elif isinstance(x, dict) and x.get("_bytes"):
            # Raw image bytes (pasted from Quill rich-text fields)
            entries.append({"_bytes": x["_bytes"],
                            "caption": str(x.get("caption") or ""),
                            "path": None})
        elif isinstance(x, dict) and x.get("path"):
            entries.append({"path": str(x["path"]),
                            "caption": str(x.get("caption") or "")})

    # Display dimensions: how large the image appears in the Excel cell at
    # 100% zoom. The embedded PNG bytes are kept at FULL native resolution
    # so the image remains sharp when the client zooms in — Excel has the
    # original pixels and renders them at higher fidelity as zoom increases.
    MAX_DISPLAY_W = 360
    MAX_DISPLAY_H = 220
    PADDING_PX    = 8
    CAPTION_PAD   = 6
    CAPTION_LINE  = 16   # px per caption line at display scale
    MIN_CAPTION_H = 22   # px — at least one line + padding

    # Load a small display-scale font for caption layout measurements.
    # The actual rendering font is scaled up proportionally per image.
    _FONT_PATHS = (
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
        "/usr/share/fonts/dejavu/DejaVuSans.ttf",
        "C:/Windows/Fonts/arial.ttf",
        "C:/Windows/Fonts/segoeui.ttf",
    )
    caption_font = None
    for ttf in _FONT_PATHS:
        try:
            caption_font = ImageFont.truetype(ttf, 12)
            break
        except (OSError, IOError):
            continue
    if caption_font is None:
        caption_font = ImageFont.load_default()

    # Build one full-resolution composite PNG per screenshot.
    # `composites` stores (png_bytes, display_w, display_h) — the PNG is
    # native resolution, but img.width / img.height are set to the display
    # dimensions so the cell stays a reasonable size at 100% zoom.
    composites: list[tuple[BytesIO, int, int]] = []   # (png_bytes, disp_w, disp_h)
    for i, entry in enumerate(entries):
        raw_bytes: bytes | None = entry.get("_bytes")
        if raw_bytes is None:
            p = Path(entry.get("path") or "")
            if not p.exists() or not p.is_file():
                continue
            try:
                with open(p, "rb") as _fh:
                    raw_bytes = _fh.read()
            except OSError:
                continue

        try:
            with PILImage.open(BytesIO(raw_bytes)) as im:
                im.load()
                w0, h0 = im.size
                if w0 <= 0 or h0 <= 0:
                    continue

                # Display scale: how much to shrink for the cell view.
                display_scale = min(MAX_DISPLAY_W / w0, MAX_DISPLAY_H / h0, 1.0)
                disp_w = int(w0 * display_scale)
                disp_h = int(h0 * display_scale)

                caption = (entry["caption"] or "").strip() or f"Screenshot {i+1}"

                # Caption height at display scale (drives row/col sizing).
                lines_disp = _wrap_text_to_pixel_width(
                    caption, caption_font, max(1, disp_w - 2 * CAPTION_PAD),
                )
                cap_disp_h = max(MIN_CAPTION_H,
                                 CAPTION_LINE * len(lines_disp) + 2 * CAPTION_PAD)

                # Caption geometry at native resolution so text is crisp.
                # native_scale is >= 1; capped so the font stays sane.
                native_scale = max(1.0, 1.0 / display_scale) if display_scale > 0 else 1.0
                native_font_size = min(max(int(12 * native_scale), 12), 72)
                native_cap_font = None
                for ttf in _FONT_PATHS:
                    try:
                        native_cap_font = ImageFont.truetype(ttf, native_font_size)
                        break
                    except (OSError, IOError):
                        continue
                if native_cap_font is None:
                    native_cap_font = caption_font

                native_pad    = max(CAPTION_PAD, int(CAPTION_PAD * native_scale))
                native_line_h = max(CAPTION_LINE, int(CAPTION_LINE * native_scale))
                native_min_h  = max(MIN_CAPTION_H, int(MIN_CAPTION_H * native_scale))

                lines_native = _wrap_text_to_pixel_width(
                    caption, native_cap_font, max(1, w0 - 2 * native_pad),
                )
                cap_native_h = max(native_min_h,
                                   native_line_h * len(lines_native) + 2 * native_pad)

                # Composite at full native resolution — no downscale of the image.
                full_native_h = h0 + cap_native_h
                composite = PILImage.new("RGB", (w0, full_native_h), (255, 255, 255))
                composite.paste(im.convert("RGB"), (0, 0))
                draw = ImageDraw.Draw(composite)
                draw.rectangle([(0, h0), (w0, full_native_h)], fill=(245, 245, 245))
                draw.line([(0, h0), (w0, h0)], fill=(200, 200, 200),
                          width=max(1, int(native_scale)))
                for j, line in enumerate(lines_native):
                    draw.text(
                        (native_pad, h0 + native_pad + j * native_line_h),
                        line, fill=(40, 40, 40), font=native_cap_font,
                    )

                buf = BytesIO()
                composite.save(buf, format="PNG")
                buf.seek(0)
                # Store DISPLAY dimensions — the embedded PNG is full-res.
                composites.append((buf, disp_w, disp_h + cap_disp_h))
        except Exception:
            continue

    if not composites:
        return

    # Captions are now baked into the images — clear any cell text so
    # nothing competes with the drawings.
    target_cell.value = None

    # Total height we need = sum of composite heights + outer padding.
    total_h = sum(h for (_, _, h) in composites) \
            + PADDING_PX * (len(composites) + 1)
    max_w   = max(w for (_, w, _) in composites)

    # Bump the row height so the stack fits. 1 row-height unit ≈ 1.33 px,
    # so multiply by 0.75 to convert px → pt.
    needed_pt = total_h * 0.75
    cur_pt = ws.row_dimensions[row_idx].height or 15.0
    if needed_pt > cur_pt:
        ws.row_dimensions[row_idx].height = needed_pt

    # Bump the column width so the widest image fits horizontally.
    # 1 column-width unit ≈ 7 px (default font).
    needed_cw = max_w / 7.0 + 2
    cur_cw = ws.column_dimensions[col_letter].width or 8.43
    if needed_cw > cur_cw:
        ws.column_dimensions[col_letter].width = needed_cw

    # Embed each composite as a separately-anchored drawing using
    # TwoCellAnchor with editAs="oneCell". Both anchors point at the
    # same target cell with progressive y-offsets so the images stack
    # neatly. Excel reads editAs="oneCell" as "move with cells but
    # don't size with cells" — the image is firmly associated with
    # the cell and travels with it on row sort / row insert / row
    # delete operations.
    y_offset_px = PADDING_PX
    for buf, w, h in composites:
        try:
            img = XLImage(buf)
            img.width = w
            img.height = h
            from_marker = AnchorMarker(
                col=col_idx_1based - 1,
                colOff=pixels_to_EMU(PADDING_PX),
                row=row_idx - 1,
                rowOff=pixels_to_EMU(y_offset_px),
            )
            to_marker = AnchorMarker(
                col=col_idx_1based - 1,
                colOff=pixels_to_EMU(PADDING_PX + w),
                row=row_idx - 1,
                rowOff=pixels_to_EMU(y_offset_px + h),
            )
            img.anchor = TwoCellAnchor(
                editAs="oneCell",
                _from=from_marker,
                to=to_marker,
            )
            ws.add_image(img)
            y_offset_px += h + PADDING_PX
        except Exception:                                    # pragma: no cover
            # Skip the bad image — the export is best-effort. The
            # consultant can re-upload if needed.
            continue


def _wrap_text_to_pixel_width(text: str, font, max_px: int) -> list[str]:
    """Greedy word-wrap `text` to fit within `max_px` pixels using
    `font` for measurement. Returns at least one line, even if the
    input is blank — keeps the caption strip a constant minimum
    height.

    Pillow ≥ 9.2 deprecated `ImageFont.getsize()` in favour of
    `.getbbox()`; we use whichever is available so the function works
    across the Pillow versions the project might pin.
    """
    if not text:
        return [""]

    def _measure(s: str) -> int:
        try:
            bbox = font.getbbox(s)
            return bbox[2] - bbox[0]
        except AttributeError:                              # pragma: no cover
            return font.getsize(s)[0]

    words = text.split()
    if not words:
        return [""]
    lines: list[str] = []
    current = words[0]
    for word in words[1:]:
        trial = f"{current} {word}"
        if _measure(trial) <= max_px:
            current = trial
        else:
            lines.append(current)
            current = word
    lines.append(current)
    return lines


_FORMULA_TRIGGERS = frozenset("=+-@|\t")


def _no_formula(val: Any) -> Any:
    """Prevent Excel formula injection.

    Strings that start with a formula-trigger character (`=`, `+`, `-`, `@`,
    `|`, tab) are prefixed with a single quote so Excel treats the cell as
    plain text rather than evaluating it as a formula. The leading `'` is
    invisible in Excel's display (it acts as a quote-prefix escape) but is
    stored in the XML so the cell is never executed.

    Non-string values are returned unchanged.
    """
    if isinstance(val, str) and val and val[0] in _FORMULA_TRIGGERS:
        return "'" + val
    return val


def _write_header_row(ws, headers: list[str]) -> None:
    for col_idx, label in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=label)


def _apply_na_status(ws, col_idx: int, row: int) -> None:
    """Write 'N/A' with a plain white fill and black font for Informational findings."""
    from openpyxl.styles import PatternFill, Font
    cell = ws.cell(row=row, column=col_idx)
    cell.value = "N/A"
    cell.fill = PatternFill("solid", fgColor="FFFFFF")
    cell.font = Font(color="000000")
    cell.alignment = _CENTER_ALIGN


def _write_data_rows(ws, start_row: int, headers, findings,
                      col_keys: list[Optional[str]]) -> None:
    for offset, f in enumerate(findings):
        row = start_row + offset
        for col_idx, key in enumerate(col_keys, start=1):
            if not key:
                continue
            val = f.get(key) if isinstance(f, dict) else getattr(f, key, None)
            if key == "severity":
                _apply_severity_styling(ws, col_idx, row, val)
                continue
            if key == "status":
                if hasattr(val, "value"):
                    val = val.value
                # Informational findings carry no remediation status — show N/A
                # with no background and black font instead of a traffic-light colour.
                sev = f.get("severity") if isinstance(f, dict) else getattr(f, "severity", None)
                if hasattr(sev, "value"):
                    sev = sev.value
                if str(sev or "").strip().lower() == "informational":
                    _apply_na_status(ws, col_idx, row)
                else:
                    _apply_status_styling(ws, col_idx, row, val)
                continue
            if val is None:
                continue
            if hasattr(val, "value"):
                val = val.value
            cell = ws.cell(row=row, column=col_idx)
            cell.value = _no_formula(val)
            if _CENTER_ALIGN is not None:
                cell.alignment = _CENTER_ALIGN
