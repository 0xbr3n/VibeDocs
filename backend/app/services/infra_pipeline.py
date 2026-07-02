"""Report-level Infra Scan Pipeline orchestrator.

Runs the existing toolkit pipelines (`services.tools.va_recurring`
and `services.tools.va_retest`) directly from the report editor —
the consultant uploads their CSV(s), picks a route (first-scan /
recurring / retest), and the result is materialised on the report
version as:

  * Three GROUPED `ReportFinding` rows, one per category:
        - "Outdated / Unsupported Software Versions (Grouped)"
        - "SSL / TLS Misconfigurations (Grouped)"
        - "Information Disclosure via Service Banners (Grouped)"
    Each carries the corresponding by-category .xlsx as a file
    attachment under `ReportFinding.attachments`. Affected Asset is
    auto-stamped to the team's standard pointer phrase so the Word
    + Excel-tracker output reads:
        "Please refer to applications and affected hosts listed in
         the attachment below."

  * The raw pipeline ZIP is persisted under
    `UPLOAD_DIR/infra_pipeline/{report_id}/{version}/run-{ts}.zip`
    so the consultant can re-download every artefact (the
    uncategorised + risk-accepted + audit sheets) for manual
    review.

Re-running the pipeline replaces the existing grouped attachments
in place — same finding row, new file. Individual rows from the
"uncategorised" bucket are NOT auto-created here; the consultant
opens the uncategorised .xlsx from the result ZIP, decides what
belongs where, and either re-uploads to fix the categorisation or
adds bespoke findings manually.
"""
from __future__ import annotations

import io
import logging
import shutil
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional

from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from ..config import settings
from ..models import (
    FindingLibrary, FindingStatus, ReportFinding, ReportVersion, Severity, User,
)

logger = logging.getLogger(__name__)

# Severity ranking for max-severity roll-up when multiple CSV rows map to
# the same (title, port) group. Informational is explicitly 0 (lowest).
_SEV_RANK: dict[str, int] = {
    "Critical": 4, "High": 3, "Medium": 2, "Low": 1, "Informational": 0,
}


# ============================================================
# Mapping: pipeline category name → (library title, attachment label)
# ============================================================
#
# The categoriser in `va_automater.categorize` emits the LHS strings
# verbatim; the LHS must match those names exactly. Library titles
# on the RHS must match the seed (`seed_findings_v2._findings_catalogue`)
# so the same row picks up new attachments on re-runs.

CATEGORY_TO_LIBRARY: dict[str, dict] = {
    "Outdated Software & Patches": {
        "library_title": "Outdated / Unsupported Software Versions (Grouped)",
        "attachment_filename": "outdated_software.xlsx",
        "attachment_label": "List of missing security patches and outdated versions",
        # Keyword used to match the by-category filename in the
        # pipeline output ZIP. Files are written as
        # `Outdated Software _ Patches.xlsx` (the `&` is sanitised
        # to `_`), so an exact-string compare to the category key
        # `"Outdated Software & Patches"` misses. Substring match
        # on a stable word from the start of the filename catches
        # every spelling variant.
        "filename_keyword": "outdated",
    },
    "SSL Misconfigurations": {
        "library_title": "SSL / TLS Misconfigurations (Grouped)",
        "attachment_filename": "ssl_misconfig.xlsx",
        "attachment_label": "List of SSL / TLS misconfigurations on the affected hosts",
        "filename_keyword": "ssl",
    },
    "Information Disclosure": {
        "library_title": "Information Disclosure via Service Banners (Grouped)",
        "attachment_filename": "info_disclosure.xlsx",
        "attachment_label": "List of information disclosure issues on the affected hosts",
        "filename_keyword": "information",
    },
    "Insecure Service Configurations": {
        "library_title": "Insecure Service Configurations (Grouped)",
        "attachment_filename": "insecure_service_config.xlsx",
        "attachment_label": "List of insecure service / OS configuration weaknesses on the affected hosts",
        # by_category writer sanitises the category name to a filename;
        # "insecure" is the stable leading word in
        # `Insecure Service Configurations.xlsx`.
        "filename_keyword": "insecure",
    },
}

# ============================================================
# Nessus → ReportFinding column mapping for the Uncategorised
# importer
# ============================================================
#
# Each entry maps a ReportFinding attribute → the list of Nessus
# / pipeline header strings it might appear under. Header matching
# is case-insensitive + substring (header.lower() in cell.lower()
# OR cell.lower() in header.lower()). The first match wins per row.
#
# `Host` typically holds a comma-joined IP list when the pipeline's
# `group_ips_in_by_category` option is on; we use it verbatim as
# `affected_asset` so the consultant sees every affected host in
# one cell.

_NESSUS_COLMAP: dict[str, tuple[str, ...]] = {
    "title":          ("name", "finding name", "title", "plugin name"),
    "affected_asset": ("host", "ip", "ips", "asset", "affected host"),
    "severity_text":  ("risk", "severity", "risk rating"),
    "description":    ("description", "synopsis"),
    "poc_steps":      ("plugin output", "plugin_output", "output",
                        "details", "evidence"),
    "remediation":    ("solution", "recommendation", "remediation",
                        "fix"),
    "references":     ("see also", "references", "see_also", "url"),
    "cvss_score":     ("cvss v3.0 base score", "cvss v3 base score",
                        "cvss3 base score", "cvss3 score", "cvss2 score",
                        "cvss base score", "cvss score"),
    "cvss_vector":    ("cvss v3.0 vector", "cvss v3 vector",
                        "cvss3 vector", "cvss2 vector", "cvss vector",
                        "vector"),
    "cwe":            ("cwe", "cwe id"),
    "port":           ("port", "ports"),
    "plugin_id":      ("plugin id", "plugin_id", "pluginid", "nessus id",
                        "plugin", "pid"),
}

# Risk → Severity enum. Nessus emits the LHS strings verbatim. We
# skip blank / "None" / "Informational" rows entirely so the
# consultant doesn't end up with hundreds of low-signal entries
# on every infra report.
_RISK_TO_SEVERITY: dict[str, str] = {
    "critical":      "Critical",
    "high":          "High",
    "medium":        "Medium",
    "low":           "Low",
}


# Categories whose by-category xlsx file the user has explicitly
# asked to OMIT from the served result ZIP (the consultant
# reviewed the output and decided these two buckets don't go on
# Infra reports today — they live in their own appendices, if at
# all). The underlying pipeline still produces them and uses them
# for the categorisation logic; we just strip them from the ZIP we
# stream back. Removing them keeps the consultant's download
# focused on the 4 files that map to grouped findings + the
# uncategorised audit.
_STRIPPED_CATEGORY_KEYWORDS: set[str] = {
    "default", "weak credentials",
    "web application",
}

# Substrings to keep when re-building the served ZIP from the
# pipeline output. Anything in `by_category/` whose filename
# doesn't contain one of these survives only if it's the
# `uncategorized` audit. Other roots (audit sheets, summary.txt,
# remaining_findings.xlsx) are dropped to keep the user-facing
# bundle tight — they were useful for the toolkit-mode debug
# walkthrough but redundant on a report where the 3 grouped
# findings already carry the per-category list.
_KEEP_BY_CATEGORY_KEYWORDS: set[str] = {
    "outdated", "ssl", "information disclosure", "insecure", "uncategor",
}


# Standard "Please refer to attachment" phrase the team uses for
# every grouped finding. Stamped on both Affected Asset (the
# {{ f.affected_asset }} placeholder in the Word template and the
# tracker exporter's "Affected Asset" column) so the deliverable
# reads consistently. Quoted as a module constant so it's the
# single source of truth — tests and the tracker exporter can
# import it instead of hardcoding the string.
ATTACHMENT_POINTER_PHRASE = (
    "Please refer to applications and affected hosts listed in the "
    "attachment below."
)


# ============================================================
# Pipeline route enum
# ============================================================

PIPELINE_FIRST_SCAN = "first_scan"
PIPELINE_RECURRING  = "recurring"
PIPELINE_RETEST     = "retest"

_VALID_PIPELINES = {PIPELINE_FIRST_SCAN, PIPELINE_RECURRING, PIPELINE_RETEST}


# ============================================================
# Public entry point
# ============================================================

def run_infra_pipeline(
    db: Session,
    rv: ReportVersion,
    user: User,
    *,
    pipeline: str,
    current_csvs: Iterable[tuple[str, bytes]],
    risk_accept: Optional[Iterable[tuple[str, bytes]]] = None,
    prev_tracker: Optional[Iterable[tuple[str, bytes]]] = None,
    original_tracker: Optional[tuple[str, bytes]] = None,
    new_ip_action: str = "include",
    enable_version_check: bool = True,
    custom_comment_col: str = "",
    custom_comment_default: str = "",
    group_ips_in_by_category: bool = False,
    include_informational: bool = False,
    skip_categories: Optional[set] = None,
) -> dict:
    """Run the chosen pipeline and persist the results onto `rv`.

    Returns a summary suitable for the UI:
        {
          "ok": True,
          "pipeline": "first_scan" | "recurring" | "retest",
          "result_zip_url": "/api/reports/...",
          "result_zip_name": "...",
          "summary": {...passthrough from underlying pipeline...},
          "groups_attached": [
              {"category": "...", "library_title": "...",
               "finding_id": 123, "attachment_filename": "..."},
              ...
          ],
        }
    """
    pipeline = (pipeline or "").lower().strip()
    if pipeline not in _VALID_PIPELINES:
        raise ValueError(
            f"pipeline must be one of {sorted(_VALID_PIPELINES)}, got {pipeline!r}"
        )

    current_csvs_list = list(current_csvs or [])
    # Recurring scans run without new CSVs — findings carry over from the
    # prior version automatically; only risk-accept and tracker diffs matter.
    if not current_csvs_list and pipeline != PIPELINE_RECURRING:
        raise ValueError("At least one current scan CSV is required.")

    # ---- Step 1: run the underlying pipeline ----
    # Recurring with no CSVs: use the existing grouped-finding xlsx attachments
    # from this version as the current scan data. The by-category xlsx files
    # written by the pipeline carry the same display column names recognised by
    # the va_automater loaders (Finding Name / Host / Port / Risk / Plugin ID),
    # so they round-trip cleanly back into `run_pipeline()` as CSV input.
    # This lets the consultant re-run with fresh risk-accept / tracker docs
    # without re-uploading the original Nessus CSVs.
    if pipeline == PIPELINE_RECURRING and not current_csvs_list:
        synthetic = _collect_existing_xlsx_as_csvs(rv)
        if not synthetic:
            # No grouped finding data on disk — nothing to process.
            return {
                "ok": True,
                "pipeline": pipeline,
                "result_zip_url": None,
                "result_zip_name": None,
                "summary": {
                    "note": (
                        "No scan CSVs provided and no existing categorised "
                        "findings were found on this report version. Upload "
                        "Nessus CSV(s) to run the pipeline."
                    ),
                },
                "groups_attached": [],
                "subtract_review_rows": 0,
                "subtract_review_filename": None,
            }
        current_csvs_list = synthetic
        logger.info(
            "recurring no-CSV: built %d synthetic CSV(s) from existing "
            "grouped findings on vid=%s",
            len(synthetic), rv.id,
        )

    if pipeline == PIPELINE_RETEST:
        if not original_tracker:
            raise ValueError(
                "Retest pipeline requires the original tracker file."
            )
        from .tools.va_retest import run_retest
        result = run_retest(
            current_csvs=current_csvs_list,
            original_tracker=original_tracker,
            new_ip_action=new_ip_action,
            enable_version_check=enable_version_check,
            custom_comment_col=custom_comment_col,
            custom_comment_default=custom_comment_default,
        )
    else:
        # First-scan and Recurring both call the same underlying
        # function. The only difference is whether risk-accept and
        # previous-tracker inputs are supplied — first-scan never
        # has them, recurring usually does.
        from .tools.va_recurring import run_pipeline
        result = run_pipeline(
            current_csvs=current_csvs_list,
            risk_accept=(list(risk_accept) if risk_accept else None),
            prev_tracker=(list(prev_tracker) if prev_tracker else None),
            custom_comment_col=custom_comment_col,
            custom_comment_default=custom_comment_default,
            group_ips_in_by_category=group_ips_in_by_category,
        )

    raw_zip_bytes = result.get("zip_bytes") or b""
    pipeline_summary = result.get("summary") or {}

    # Re-pack the ZIP so the user-facing download contains only the
    # checked grouped-finding xlsx files + Uncategorized.xlsx (with
    # skip_categories rows merged in) + summary.txt. The extractor
    # in step 4 still runs against the ORIGINAL bytes so nothing is
    # lost on the attach-to-findings path.
    zip_bytes = _build_filtered_zip(raw_zip_bytes, skip_categories=skip_categories)

    # ---- Step 2: persist the result ZIP server-side ----
    result_zip_name = result.get("zip_name") or _default_zip_name(pipeline)
    zip_dir = Path(settings.UPLOAD_DIR) / "infra_pipeline" / str(rv.report_id) / rv.version
    zip_dir.mkdir(parents=True, exist_ok=True)
    zip_disk_name = f"run-{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.zip"
    zip_disk_path = zip_dir / zip_disk_name
    zip_disk_path.write_bytes(zip_bytes)

    # ---- Step 3: for the retest pipeline there's no category
    # extraction — we drop the updated tracker as a single attachment
    # on the version's note attachment list, and we're done. Skip the
    # grouping logic below.
    if pipeline == PIPELINE_RETEST:
        # Retest path may still surface a subtract-review signal if the
        # tracker comparison flagged borderline matches. Pull it through
        # too so the UI shows the same triage callout when relevant.
        retest_review_rows = int(
            pipeline_summary.get("n_subtract_review_rows") or 0
        )
        return {
            "ok": True,
            "pipeline": pipeline,
            "result_zip_url": _zip_download_url(rv, zip_disk_name),
            "result_zip_name": result_zip_name,
            "summary": pipeline_summary,
            "groups_attached": [],
            "subtract_review_rows": retest_review_rows,
            "subtract_review_filename": "subtract_review.xlsx",
        }

    # ---- Step 4: extract the by-category xlsx files from the
    # ORIGINAL ZIP (the filtered ZIP we save to disk may have
    # already dropped categories the user opted to strip). ----
    category_files = _extract_category_files(raw_zip_bytes)

    # ---- Step 5: upsert one ReportFinding per category ----
    groups_attached = _upsert_grouped_findings(
        db, rv, user, category_files, pipeline_summary, zip_dir,
        skip_categories=skip_categories,
    )

    # ---- Step 6: import individual findings out of the
    # Uncategorised workbook. Reduces the consultant's manual
    # workload — every Low / Medium / High / Critical row gets a
    # ReportFinding stub auto-created with title / affected_asset /
    # description / poc_steps / remediation / severity / CVSS
    # filled from the Nessus row. Informational + blank-Risk rows
    # are skipped (noise on the deliverable). The consultant still
    # has to add screenshots + review/delete as the UI callout
    # tells them — see the `uncategorised_added` summary the API
    # returns. Idempotent on re-run: stale auto-imported rows from
    # a previous pipeline run are deleted before the fresh batch.
    #
    # Also import findings from any skip_categories xlsx files so
    # that findings routed to a category the user opted to skip
    # (e.g. "Insecure Service Configurations" when using only 3
    # groups) still appear as individual findings in the report.
    # Without this they would be silently lost — categorised into
    # a non-selected bucket and absent from both the Uncategorised
    # workbook AND the grouped finding rows.
    skipped_cat_xlsx: list[bytes] = []
    if skip_categories:
        for cat_name in (skip_categories or set()):
            cat_bytes = category_files.get(cat_name)
            if cat_bytes:
                skipped_cat_xlsx.append(cat_bytes)

    try:
        # Count what's about to be wiped so the API response can
        # report the swap.
        prev_count = (
            db.query(ReportFinding)
              .filter(ReportFinding.report_version_id == rv.id,
                      ReportFinding.source == "infra_pipeline_uncategorised")
              .count()
        )
        uncat_added = _import_uncategorised_findings(
            db, rv, user, raw_zip_bytes,
            include_informational=include_informational,
            extra_xlsx_list=skipped_cat_xlsx or None,
        )
    except Exception as e:                                  # pragma: no cover
        logger.warning(
            "uncategorised auto-import failed for vid=%s: %s", rv.id, e,
        )
        uncat_added = []
        prev_count = 0

    db.commit()

    # Number of rows the underlying matcher flagged for human review,
    # plus the filename inside the result ZIP. Pulled out as top-level
    # fields so the UI doesn't have to peek into pipeline_summary —
    # they drive a dedicated triage callout. Filename is hard-coded to
    # match `pipelines.OUT_SUBTRACT_REVIEW`; the file lives at the root
    # of the result ZIP, alongside summary.txt.
    subtract_review_rows = int(pipeline_summary.get("n_subtract_review_rows") or 0)
    partial_upgrade_rows = int(pipeline_summary.get("n_partial_upgrades") or 0)

    # Per-severity counts for the UI summary callout.
    sev_counts: dict[str, int] = {}
    for row in uncat_added:
        sev = row.get("severity", "Unknown")
        sev_counts[sev] = sev_counts.get(sev, 0) + 1

    return {
        "ok": True,
        "pipeline": pipeline,
        "result_zip_url": _zip_download_url(rv, zip_disk_name),
        "result_zip_name": result_zip_name,
        "summary": pipeline_summary,
        "groups_attached": groups_attached,
        "uncategorised_added_count": len(uncat_added),
        "uncategorised_added": uncat_added,
        # Per-severity breakdown: {"Critical": N, "High": N, "Medium": N, "Low": N}.
        # Used by the UI to show "3 High, 5 Medium, 2 Low" instead of just "10 findings".
        "uncategorised_severity_counts": sev_counts,
        "uncategorised_replaced_count": prev_count,
        "subtract_review_rows": subtract_review_rows,
        "subtract_review_filename": "subtract_review.xlsx",
        "partial_upgrade_rows": partial_upgrade_rows,
        "partial_upgrade_filename": "partial_upgrades.xlsx",
    }


# ============================================================
# Internals
# ============================================================

def _default_zip_name(pipeline: str) -> str:
    ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    return f"infra_{pipeline}_{ts}.zip"


def _zip_download_url(rv: ReportVersion, zip_disk_name: str) -> str:
    """Public download URL for the persisted result ZIP."""
    return (
        f"/api/reports/versions/{rv.id}/infra-pipeline/"
        f"download?file={zip_disk_name}"
    )


def _extract_category_files(zip_bytes: bytes) -> dict[str, bytes]:
    """Pull every `by_category/*.xlsx` member out of the pipeline
    output ZIP. Returns a `{category_name: xlsx_bytes}` dict where
    `category_name` matches the keys of `CATEGORY_TO_LIBRARY`.

    Matching is KEYWORD-based rather than exact-string because the
    pipeline sanitises filenames when writing them to disk —
    `Outdated Software & Patches` becomes `Outdated Software _ Patches.xlsx`,
    `SSL/TLS Misconfigurations` becomes `SSL_TLS Misconfigurations.xlsx`,
    etc. Keyword matching survives every such substitution.
    """
    out: dict[str, bytes] = {}
    if not zip_bytes:
        return out
    try:
        zf = zipfile.ZipFile(io.BytesIO(zip_bytes))
    except zipfile.BadZipFile:
        logger.warning("pipeline returned a non-ZIP payload — skipping extraction")
        return out
    try:
        for info in zf.infolist():
            name = info.filename
            if "/" not in name or not name.lower().endswith(".xlsx"):
                continue
            head, _, leaf = name.rpartition("/")
            if "by_category" not in head:
                continue
            stem = leaf.rsplit(".", 1)[0].lower()
            for category_name, spec in CATEGORY_TO_LIBRARY.items():
                kw = (spec.get("filename_keyword") or category_name).lower()
                if kw in stem:
                    out[category_name] = zf.read(info)
                    break
    finally:
        zf.close()
    return out


def _merge_xlsx_bytes(
    base_xlsx: Optional[bytes],
    extra_list: list[bytes],
    *,
    source_labels: Optional[list[Optional[str]]] = None,
) -> bytes:
    """Append data rows from extra xlsx files into base_xlsx.

    Reads the header from the first available source, then collects all
    data rows from every source (skipping each source's own header row).
    Returns new xlsx bytes. Falls back to base_xlsx on any error.

    Args:
        base_xlsx: Base xlsx bytes (e.g. Uncategorized.xlsx). May be None.
        extra_list: Additional xlsx files to append. Must be non-empty.
        source_labels: Optional list of one label per source (index 0 for
            base_xlsx, indices 1+ for extra_list entries). When provided,
            a "Source Category" column is appended to every data row so
            the consultant can see which category each finding came from.
            A None label means the row came from the base Uncategorized
            workbook (column is left blank for those rows).
    """
    if not extra_list:
        return base_xlsx or b""
    try:
        from openpyxl import load_workbook, Workbook
    except ImportError:
        return base_xlsx or b""

    SOURCE_CAT_COL = "Source Category"
    inject_labels = bool(source_labels)

    # All sources in order: base first, then extras.
    # Build parallel list of (bytes, label) pairs.
    sources_with_labels: list[tuple[Optional[bytes], Optional[str]]] = []
    if base_xlsx:
        lbl = source_labels[0] if source_labels else None
        sources_with_labels.append((base_xlsx, lbl))
    elif source_labels:
        # base_xlsx is None but labels were provided — skip index 0
        pass
    sources_with_labels.extend(
        (xb, source_labels[1 + i] if source_labels and (1 + i) < len(source_labels) else None)
        for i, xb in enumerate(extra_list)
    )

    all_rows: list[list] = []
    header: Optional[list] = None

    for src_idx, (xlsx_bytes, label) in enumerate(sources_with_labels):
        if not xlsx_bytes:
            continue
        try:
            wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
            ws = wb.active
            if ws is None:
                wb.close()
                continue
            first_row = True
            for row in ws.iter_rows(values_only=True):
                if first_row:
                    if header is None:
                        header = [c if c is not None else "" for c in row]
                    first_row = False
                    continue
                if row and any(c for c in row if c is not None):
                    data_row = [c if c is not None else "" for c in row]
                    if inject_labels:
                        data_row.append(label or "")
                    all_rows.append(data_row)
            wb.close()
        except Exception as e:
            logger.warning(
                "infra_pipeline: _merge_xlsx_bytes source[%d] read failed: %s",
                src_idx, e,
            )

    out_wb = Workbook()
    out_ws = out_wb.active
    if header:
        if inject_labels:
            header = list(header) + [SOURCE_CAT_COL]
        out_ws.append(header)
    for row in all_rows:
        out_ws.append(row)
    buf = io.BytesIO()
    out_wb.save(buf)
    return buf.getvalue()


def _build_filtered_zip(
    original_zip_bytes: bytes,
    skip_categories: Optional[set] = None,
) -> bytes:
    """Re-pack the pipeline output ZIP for the user-facing download.

    Kept: the checked grouped-finding by_category xlsx files +
    Uncategorized.xlsx (with skipped-category rows merged in) + summary.txt.
    Dropped: debug sheets, remaining_findings, match_preview, and any
    by_category file whose category is in skip_categories.

    When skip_categories is provided, rows from those category xlsx files
    are appended into Uncategorized.xlsx so the consultant sees all
    un-grouped findings in a single workbook.

    Returns NEW zip bytes. Falls back to original bytes on parse error.
    """
    if not original_zip_bytes:
        return original_zip_bytes
    try:
        src = zipfile.ZipFile(io.BytesIO(original_zip_bytes))
    except zipfile.BadZipFile:
        return original_zip_bytes

    # Build a set of filename keywords for skip_categories so we can
    # identify their xlsx files by partial name match.
    skip_kws: set[str] = set()
    if skip_categories:
        for cat_name in skip_categories:
            spec = CATEGORY_TO_LIBRARY.get(cat_name)
            if spec:
                kw = (spec.get("filename_keyword") or cat_name).lower()
                skip_kws.add(kw)

    # Build a reverse map: filename_keyword → category_name so we can
    # annotate merged rows with which skipped category they came from.
    skip_kw_to_cat: dict[str, str] = {}
    if skip_categories:
        for cat_name in skip_categories:
            spec = CATEGORY_TO_LIBRARY.get(cat_name)
            if spec:
                kw = (spec.get("filename_keyword") or cat_name).lower()
                skip_kw_to_cat[kw] = cat_name

    # First pass: read Uncategorized.xlsx bytes and skip_category xlsx
    # bytes so we can merge them before writing the output ZIP.
    # Track (bytes, category_name) pairs so we can stamp a Source Category
    # column on each row so the consultant knows where the finding came from.
    uncategorised_bytes: Optional[bytes] = None
    skip_xlsx_info: list[tuple[bytes, str]] = []  # (bytes, category_name)

    for info in src.infolist():
        if info.is_dir():
            continue
        name = info.filename
        head, _, leaf = name.rpartition("/")
        lleaf = leaf.lower()
        if "by_category" not in head or not lleaf.endswith(".xlsx"):
            continue
        if "uncategor" in lleaf:
            uncategorised_bytes = src.read(info)
        elif skip_kws and any(k in lleaf for k in skip_kws):
            # Determine which category this file belongs to.
            cat_label = next(
                (cat for kw, cat in skip_kw_to_cat.items() if kw in lleaf),
                "Uncategorized",
            )
            skip_xlsx_info.append((src.read(info), cat_label))

    # Merge skipped-category rows into Uncategorized.xlsx.
    # Pass source_labels so a "Source Category" column is injected:
    #   - index 0  → None  (base Uncategorized.xlsx rows, left blank)
    #   - index 1+ → category names for skip_categories files
    merged_uncat: Optional[bytes] = None
    if skip_xlsx_info:
        skip_xlsx_bytes = [x[0] for x in skip_xlsx_info]
        skip_xlsx_labels = [x[1] for x in skip_xlsx_info]
        # Label for the base Uncategorized rows is None (blank column).
        source_labels = [None] + skip_xlsx_labels
        merged_uncat = _merge_xlsx_bytes(
            uncategorised_bytes, skip_xlsx_bytes, source_labels=source_labels,
        )

    # Second pass: write the filtered ZIP.
    buf = io.BytesIO()
    uncategorised_written = False
    try:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as dst:
            for info in src.infolist():
                if info.is_dir():
                    continue
                name = info.filename
                lname = name.lower()
                head, _, leaf = name.rpartition("/")
                lleaf = leaf.lower()

                # Keep summary.txt at the top level.
                if lname.endswith("summary.txt") and "/" not in name:
                    dst.writestr(info, src.read(info))
                    continue

                if "by_category" in head and lleaf.endswith(".xlsx"):
                    # Drop explicitly stripped categories (Default, Weak Creds, Web App).
                    if any(s in lleaf for s in _STRIPPED_CATEGORY_KEYWORDS):
                        continue
                    # Drop skip_categories files — rows are in Uncategorized now.
                    if skip_kws and any(k in lleaf for k in skip_kws):
                        continue
                    # Write Uncategorized.xlsx (merged if applicable).
                    if "uncategor" in lleaf:
                        data = merged_uncat if merged_uncat else src.read(info)
                        dst.writestr(info, data)
                        uncategorised_written = True
                        continue
                    # Keep other checked-category xlsx files.
                    if any(k in lleaf for k in _KEEP_BY_CATEGORY_KEYWORDS):
                        dst.writestr(info, src.read(info))
                        continue
                    # Drop anything else (future/unrecognised categories).
                    continue

            # If the original ZIP had no Uncategorized.xlsx but skip_categories
            # produced rows, write them as a new entry so nothing is lost.
            if merged_uncat and not uncategorised_written:
                dst.writestr("by_category/Uncategorized.xlsx", merged_uncat)
    finally:
        src.close()
    return buf.getvalue()


def _upsert_grouped_findings(
    db: Session,
    rv: ReportVersion,
    user: User,
    category_files: dict[str, bytes],
    pipeline_summary: dict,
    attachments_dir: Path,
    skip_categories: Optional[set] = None,
) -> list[dict]:
    """For every category we have an xlsx for, create / update the
    matching ReportFinding row on this version.

    Idempotent — re-running the pipeline replaces the previously
    attached xlsx in place rather than appending a new attachment.
    Categories listed in `skip_categories` are skipped entirely —
    the pipeline still produces their xlsx (in the result ZIP) but
    no grouped finding row is created/updated for them.
    """
    attachments_dir.mkdir(parents=True, exist_ok=True)

    # Pre-load the 3 library rows by title once so we can copy the
    # canonical description / impact / remediation / references
    # without 3 separate DB lookups inside the loop.
    library_by_title: dict[str, FindingLibrary] = {}
    titles_we_want = [spec["library_title"] for spec in CATEGORY_TO_LIBRARY.values()]
    for row in db.query(FindingLibrary).filter(
            FindingLibrary.title.in_(titles_we_want)).all():
        library_by_title[row.title] = row

    out: list[dict] = []

    for category, xlsx_bytes in category_files.items():
        if not xlsx_bytes:
            continue
        if skip_categories and category in skip_categories:
            logger.debug("Skipping grouped finding for category %r (user opted out)", category)
            continue
        spec = CATEGORY_TO_LIBRARY.get(category)
        if not spec:
            continue
        library = library_by_title.get(spec["library_title"])

        # Persist the xlsx attachment under a stable filename so
        # re-runs overwrite the previous bytes cleanly.
        att_filename = spec["attachment_filename"]
        att_disk_name = f"{rv.report_id}_{rv.version}_{att_filename}"
        att_disk_path = attachments_dir / att_disk_name
        att_disk_path.write_bytes(xlsx_bytes)

        # Find an existing finding on this version with the same
        # title — re-runs of the pipeline target THAT row rather
        # than appending a new one each time.
        existing = next(
            (f for f in rv.findings if f.title == spec["library_title"]),
            None,
        )

        if existing is None:
            # Create new ReportFinding from the library row.
            severity = (library.default_severity if library
                         else Severity.medium)
            f = ReportFinding(
                report_version_id=rv.id,
                library_id=library.id if library else None,
                title=spec["library_title"],
                description=(library.description if library else ""),
                impact=(library.impact if library else ""),
                remediation=(library.remediation if library else ""),
                references=(library.references if library else ""),
                cwe=(library.cwe if library else None),
                severity=severity,
                cvss_score=(library.default_cvss_score if library else None),
                cvss_vector=(library.default_cvss_vector if library else ""),
                affected_asset=ATTACHMENT_POINTER_PHRASE,
                added_by_id=user.id,
                source="infra_pipeline",
                source_ref=category,
                attachments=[_attachment_entry(
                    att_filename, att_disk_path, spec["attachment_label"], user,
                )],
            )
            db.add(f)
            db.flush()
            created = True
        else:
            # Re-attach: drop any existing attachment with the same
            # filename, append the new one. Other fields stay as the
            # consultant left them (they may have hand-edited the
            # description or CVSS — don't trample).
            atts = list(existing.attachments or [])
            atts = [a for a in atts
                    if (a.get("filename") if isinstance(a, dict) else None)
                       != att_filename]
            atts.append(_attachment_entry(
                att_filename, att_disk_path, spec["attachment_label"], user,
            ))
            existing.attachments = atts
            flag_modified(existing, "attachments")
            # Ensure affected_asset still points at the attachment —
            # protects against the consultant clearing it accidentally.
            if not (existing.affected_asset or "").strip():
                existing.affected_asset = ATTACHMENT_POINTER_PHRASE
            f = existing
            created = False

        out.append({
            "category": category,
            "library_title": spec["library_title"],
            "finding_id": f.id,
            "attachment_filename": att_filename,
            "created": created,
        })

    return out


def _parse_xlsx_into_groups(
    xlsx_bytes: bytes,
    *,
    include_informational: bool = False,
) -> dict[tuple[str, str], dict]:
    """Parse one pipeline output xlsx into a groups dict.

    Returns ``{(title_lower, port_lower): {field: value, ...}}`` suitable
    for merging with the Uncategorised workbook groups.  Groups are keyed
    by (title.lower(), port.lower()) so same-finding rows collapse into one.

    The caller is responsible for determining which xlsx files to pass:
      - Checked categories → their xlsx is attached to a grouped finding;
        do not pass them here (no individual imports needed).
      - Unchecked categories (skip_categories) → pass as extra_xlsx_list;
        ALL Low/Medium/High/Critical findings are imported as individual rows.
      - Uncategorized.xlsx → always passed; all L/M/H/C rows imported.

    Args:
        xlsx_bytes: Raw bytes of any ``format_for_output``-produced xlsx.
        include_informational: When True, rows whose Risk column is blank /
            NaN / None / Informational are included (as Informational
            severity). Default False → only High/Medium/Low/Critical rows.
    """
    _INFO_RISKS = frozenset({"informational", "info", "none", "nan", "n/a"})
    groups: dict[tuple[str, str], dict] = {}

    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    except Exception as e:
        logger.warning("infra_pipeline: failed to open xlsx for grouping: %s", e)
        return groups

    ws = wb.active
    if ws is None or ws.max_row < 2:
        wb.close()
        return groups

    # Detect header row (first non-empty row, max 5).
    header_row = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=True):
        if row and any(c for c in row if c):
            header_row = row
            break
    if header_row is None:
        wb.close()
        return groups

    # Build column-index map.
    col_map: dict[str, int] = {}
    for idx, raw in enumerate(header_row):
        if raw is None:
            continue
        hdr = str(raw).strip().lower()
        if not hdr:
            continue
        for field_name, aliases in _NESSUS_COLMAP.items():
            if field_name in col_map:
                continue
            for alias in aliases:
                a = alias.lower()
                if a == hdr or a in hdr or hdr in a:
                    col_map[field_name] = idx
                    break

    if "severity_text" not in col_map:
        wb.close()
        return groups

    def _cell(row_tuple, key: str) -> str:
        i = col_map.get(key)
        if i is None or i >= len(row_tuple):
            return ""
        v = row_tuple[i]
        return "" if v is None else str(v).strip()

    def _normalise_host_cell(raw: str) -> list[str]:
        if not raw:
            return []
        text = raw.strip().strip('"').strip("'")
        for sep in (";", "\n", "\r", "\t"):
            text = text.replace(sep, ",")
        out: list[str] = []
        for tok in text.split(","):
            tok = tok.strip().strip('"').strip("'")
            if not tok or tok.lower() in ("n/a", "none", "null", "-"):
                continue
            out.append(tok)
        return out

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(c for c in row if c):
            continue
        risk = _cell(row, "severity_text").lower()
        if not risk or risk in _INFO_RISKS:
            if not include_informational:
                continue
            sev_value = "Informational"
        else:
            sev_value = _RISK_TO_SEVERITY.get(risk)
            if sev_value is None:
                continue

        title = _cell(row, "title")
        if not title:
            continue

        port  = _cell(row, "port")
        key   = (title.lower(), port.lower())
        hosts = _normalise_host_cell(_cell(row, "affected_asset"))

        raw_cvss = _cell(row, "cvss_score")
        cvss_score: Optional[float] = None
        if raw_cvss:
            try:
                cvss_score = float(raw_cvss)
            except (TypeError, ValueError):
                # Handle cells like "7.5 (AV:N/AC:L/Au:N/C:P/I:P/A:P)"
                # or "7.5 / 8.8" — extract the first numeric value.
                import re as _re
                _m = _re.match(r"(\d+(?:\.\d+)?)", raw_cvss.strip())
                if _m:
                    try:
                        cvss_score = float(_m.group(1))
                    except (TypeError, ValueError):
                        pass

        if key not in groups:
            groups[key] = {
                "title":       title[:500],
                "port":        port,
                "hosts":       list(hosts),
                "hosts_seen":  set(hosts),
                "description": _cell(row, "description"),
                "poc_steps":   _cell(row, "poc_steps"),
                "remediation": _cell(row, "remediation"),
                "references":  _cell(row, "references"),
                "cwe":         _cell(row, "cwe")[:255] or None,
                "severity":    sev_value,
                "cvss_score":  cvss_score,
                "cvss_vector": _cell(row, "cvss_vector"),
            }
        else:
            bucket = groups[key]
            for h in hosts:
                if h not in bucket["hosts_seen"]:
                    bucket["hosts_seen"].add(h)
                    bucket["hosts"].append(h)
            if _SEV_RANK.get(sev_value, 0) > _SEV_RANK.get(bucket["severity"], 0):
                bucket["severity"] = sev_value
            if raw_cvss and cvss_score is not None:
                if bucket["cvss_score"] is None or cvss_score > bucket["cvss_score"]:
                    bucket["cvss_score"] = cvss_score

    wb.close()
    return groups


def _import_uncategorised_findings(
    db: Session, rv: ReportVersion, user: User,
    raw_zip_bytes: bytes,
    *,
    include_informational: bool = False,
    extra_xlsx_list: list[bytes] | None = None,
) -> list[dict]:
    """Walk the Uncategorized.xlsx workbook from the pipeline ZIP
    and create a ``ReportFinding`` row for every H/M/L/C row.

    When ``include_informational`` is True, rows whose Risk is blank,
    NaN, None, or Informational are also imported with status=N/A —
    they are documented but clearly not actioned. By default those rows
    are skipped entirely (only Low / Medium / High / Critical imported).

    ``extra_xlsx_list`` accepts additional xlsx bytes (e.g. from skip_categories
    workbooks).  Their rows are merged into the same groups dict so that
    findings routed to a category the user opted to skip (e.g. "Insecure
    Service Configurations" when using only 3 grouped findings) are still
    auto-imported as individual findings.

    Returns a list of summary dicts (one per row created).
    """
    if not raw_zip_bytes and not extra_xlsx_list:
        return []

    xlsx_bytes_main: Optional[bytes] = None
    if raw_zip_bytes:
        try:
            zf = zipfile.ZipFile(io.BytesIO(raw_zip_bytes))
        except zipfile.BadZipFile:
            zf = None

        if zf is not None:
            # Find the Uncategorized.xlsx member — must live under
            # by_category/, start with "uncategor", not be an Office lock.
            candidates_under_by_category: list[str] = []
            candidates_elsewhere: list[str] = []
            for info in zf.infolist():
                if info.is_dir():
                    continue
                name   = info.filename
                leaf   = name.rsplit("/", 1)[-1]
                leaf_lc = leaf.lower()
                if leaf.startswith("~$") or not leaf_lc.endswith(".xlsx"):
                    continue
                if not leaf_lc.startswith("uncategor"):
                    continue
                if "by_category/" in name.lower():
                    candidates_under_by_category.append(name)
                else:
                    candidates_elsewhere.append(name)

            def _pick_best(names: list[str]) -> Optional[str]:
                for n in names:
                    if n.rsplit("/", 1)[-1].lower() in ("uncategorized.xlsx",
                                                        "uncategorised.xlsx"):
                        return n
                return names[0] if names else None

            target_name = (_pick_best(candidates_under_by_category)
                           or _pick_best(candidates_elsewhere))
            if target_name:
                logger.info(
                    "uncategorised auto-import: reading %r from result ZIP",
                    target_name,
                )
                try:
                    xlsx_bytes_main = zf.read(target_name)
                except Exception:
                    pass
            zf.close()

    # Build merged groups from all xlsx sources.
    groups: dict[tuple[str, str], dict] = {}
    if xlsx_bytes_main:
        groups = _parse_xlsx_into_groups(
            xlsx_bytes_main, include_informational=include_informational,
        )
    for extra_bytes in (extra_xlsx_list or []):
        extra_groups = _parse_xlsx_into_groups(
            extra_bytes, include_informational=include_informational,
        )
        for key, bucket in extra_groups.items():
            if key not in groups:
                groups[key] = bucket
            else:
                existing = groups[key]
                for h in bucket["hosts"]:
                    if h not in existing["hosts_seen"]:
                        existing["hosts_seen"].add(h)
                        existing["hosts"].append(h)
                if _SEV_RANK.get(bucket["severity"], 0) > _SEV_RANK.get(
                    existing["severity"], 0
                ):
                    existing["severity"] = bucket["severity"]
                if bucket["cvss_score"] is not None:
                    if (existing["cvss_score"] is None
                            or bucket["cvss_score"] > existing["cvss_score"]):
                        existing["cvss_score"] = bucket["cvss_score"]

    if not groups:
        return []

    # IDEMPOTENT RE-RUN: remove stale auto-imported rows from prior runs.
    previous_auto = (
        db.query(ReportFinding)
          .filter(ReportFinding.report_version_id == rv.id,
                  ReportFinding.source == "infra_pipeline_uncategorised")
          .all()
    )
    if previous_auto:
        prev_titles = [f.title for f in previous_auto]
        logger.info(
            "uncategorised auto-import: clearing %d stale row(s) (titles: %s%s)",
            len(previous_auto),
            ", ".join(prev_titles[:5]),
            "…" if len(prev_titles) > 5 else "",
        )
        for f in previous_auto:
            db.delete(f)
        db.flush()

    # Materialise one ReportFinding per group.
    # Informational / NaN / None risk findings are only imported when the
    # consultant explicitly opted in via include_informational=True, in which
    # case they get status=N/A. By default they are skipped entirely so scanner
    # noise (e.g. "Nessus SYN scanner", "ICMP timestamp") doesn't flood the VibeDocs.
    added: list[dict] = []
    for bucket in groups.values():
        sev_rank = _SEV_RANK.get(bucket.get("severity", ""), 0)
        if sev_rank == 0:
            if not include_informational:
                continue  # skip Informational / NaN / unknown severity by default
            finding_status = FindingStatus.not_applicable
        else:
            finding_status = FindingStatus.open

        hosts_str = ", ".join(bucket["hosts"])
        port      = bucket["port"]
        affected  = (f"{hosts_str}  (Port: {port})" if port and hosts_str
                     else hosts_str)

        finding = ReportFinding(
            report_version_id=rv.id,
            title=bucket["title"],
            affected_asset=affected,
            description=bucket["description"],
            poc_steps=bucket["poc_steps"],
            remediation=bucket["remediation"],
            references=bucket["references"],
            cwe=bucket["cwe"],
            severity=Severity(bucket["severity"]),
            cvss_score=bucket["cvss_score"],
            cvss_vector=bucket["cvss_vector"],
            added_by_id=user.id,
            status=finding_status,
            source="infra_pipeline_uncategorised",
            source_ref="uncategorised",
            attachments=[],
            screenshots=[],
        )
        db.add(finding)
        added.append({
            "title":          bucket["title"],
            "severity":       bucket["severity"],
            "affected_asset": affected,
            "host_count":     len(bucket["hosts"]),
        })

    if added:
        db.flush()
    return added


def _collect_existing_xlsx_as_csvs(rv: ReportVersion) -> list[tuple[str, bytes]]:
    """Convert existing grouped-finding xlsx attachments to synthetic CSV bytes.

    Finds every ``source == "infra_pipeline"`` finding on ``rv``, reads the
    first xlsx attachment for each one from disk, and returns a list of
    ``(filename.csv, csv_bytes)`` tuples suitable for passing directly to
    ``va_recurring.run_pipeline()`` as ``current_csvs``.

    The by-category xlsx files carry display column names ("Finding Name",
    "Host", "Port", "Risk", etc.) that are defined in ``COL_ALIASES`` and
    therefore recognised by ``load_nessus_folder()`` — they round-trip
    cleanly without any column renaming.

    Returns an empty list when no usable xlsx files are found on disk.
    """
    try:
        import pandas as pd
    except ImportError:
        logger.warning("pandas not available — cannot build synthetic CSVs")
        return []

    out: list[tuple[str, bytes]] = []
    seen_paths: set[str] = set()

    grouped = [f for f in rv.findings if (f.source or "").startswith("infra_pipeline")
               and f.source != "infra_pipeline_uncategorised"]

    for finding in grouped:
        for att in (finding.attachments or []):
            if not isinstance(att, dict):
                continue
            raw_path = att.get("path") or ""
            if not raw_path or raw_path in seen_paths:
                continue
            xlsx_path = Path(raw_path)
            if not xlsx_path.exists() or xlsx_path.suffix.lower() != ".xlsx":
                continue
            try:
                df = pd.read_excel(xlsx_path, dtype=str)
                if df.empty:
                    continue
                # Drop pipeline-internal columns that confuse the loader's
                # column-mapping heuristics. Status / Comments are write-only
                # audit columns, not scan fields. Category is absent from
                # by-category files already but may appear in remaining_findings.
                drop_cols = [c for c in df.columns
                             if c.lower() in ("status", "comments",
                                              "category", "source file",
                                              "source row",
                                              "ip in current scan",
                                              "near-miss on same host",
                                              "near-miss in current scan",
                                              "ip in accepted tracker",
                                              "near-miss on same host (accepted)",
                                              "near-miss in accepted tracker",
                                              "review reason", "review action",
                                              "source category",
                                              "category score")]
                df = df.drop(columns=drop_cols, errors="ignore")
                csv_bytes = df.to_csv(index=False).encode("utf-8")
                # Name after the finding title so the output ZIP's source-file
                # column helps with debugging; sanitise to a safe filename.
                safe = "".join(
                    c if c.isalnum() or c in "-_" else "_"
                    for c in (finding.title or xlsx_path.stem)
                ).strip("_")[:60] or "scan"
                out.append((f"{safe}.csv", csv_bytes))
                seen_paths.add(raw_path)
                break  # one attachment per grouped finding is all we need
            except Exception as exc:
                logger.warning(
                    "infra_pipeline: could not read xlsx %s for vid=%s: %s",
                    xlsx_path, rv.id, exc,
                )

    return out


def _attachment_entry(filename: str, disk_path: Path, label: str,
                       user: User) -> dict:
    """Serialise one attachment to the dict shape stored on the JSON
    column.
    """
    return {
        "filename": filename,
        "path": str(disk_path),
        "kind": "xlsx",
        "label": label,
        "uploaded_at": datetime.utcnow().isoformat() + "Z",
        "uploaded_by": user.username,
        # Identifier the UI uses for download / replace URLs. Keep
        # it filename-keyed so the consultant doesn't have to know
        # an internal id.
        "key": filename,
    }


# ============================================================
# Re-upload helper — used by the per-finding attachment endpoint
# ============================================================

def replace_attachment(
    finding: ReportFinding, key: str, new_bytes: bytes, user: User,
) -> dict:
    """Replace the attachment whose `key` matches `key`. Returns the
    updated entry. Raises ValueError if no attachment matches the
    key.
    """
    atts = list(finding.attachments or [])
    target_idx = None
    for i, a in enumerate(atts):
        if isinstance(a, dict) and (a.get("key") or a.get("filename")) == key:
            target_idx = i
            break
    if target_idx is None:
        raise ValueError(f"Attachment {key!r} not found on this finding.")

    existing = atts[target_idx]
    disk_path = Path(existing.get("path") or "")
    if not disk_path:
        # Path was never set — derive one under the report's pipeline
        # directory. Defence for legacy rows; new rows always carry
        # the path.
        disk_path = (Path(settings.UPLOAD_DIR) / "infra_pipeline"
                     / "ad_hoc" / f"{uuid.uuid4().hex}_{key}")
        disk_path.parent.mkdir(parents=True, exist_ok=True)
    disk_path.write_bytes(new_bytes)

    updated = dict(existing)
    updated["uploaded_at"] = datetime.utcnow().isoformat() + "Z"
    updated["uploaded_by"] = user.username
    updated["path"] = str(disk_path)
    atts[target_idx] = updated
    finding.attachments = atts
    flag_modified(finding, "attachments")
    return updated
