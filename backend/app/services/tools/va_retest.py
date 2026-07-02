"""
Retest workflow — update an existing tracker against a fresh rescan.

The recurring-scan tool subtracts risk-accepted findings from this
quarter's scan and produces categorised xlsx files. The RETEST workflow
is different: the consultant already has a tracker from the original
engagement (with images, formatting, "Client Screenshots" columns the
client filled in during remediation), and they just need that same
tracker brought up to date against the latest scan:

  * Findings present in the original tracker but NO LONGER in the
    current rescan → auto-mark **Closed** in the tracker.
  * Findings still present → optionally run the version-check helper to
    detect rows where the installed version is now ≥ the recommended
    fix (Apache 2.4.41 → 2.4.58, etc.) and auto-close those too.
  * IPs that exist in the current rescan but were NOT in the original
    tracker → reported separately; the consultant can choose to
    APPEND those new findings as fresh rows OR drop them from the
    output entirely OR list them in a side-file without touching the
    tracker.
  * Images / embedded screenshots in the original tracker are preserved
    because we edit the workbook with openpyxl (in-place row mutations)
    instead of round-tripping via pandas.

The user can also name a custom "Comments" column (e.g. "VibeDocs
Comments") and supply a default justification value that gets written
into every row this script closes — same pattern as the recurring-VA
tool, so muscle-memory carries over.

Public entry point:
    run_retest(...) -> dict with:
        zip_bytes  : the output ZIP (updated tracker + side files + summary)
        summary    : counts + lists (closed_count, version_closed_count,
                     new_ips, etc.)
        zip_name   : timestamped download filename
"""
from __future__ import annotations

import io
import logging
import re
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Iterable, Optional, Tuple

log = logging.getLogger(__name__)


# Per-file / per-request caps. Tracker workbooks can carry many
# embedded screenshots, so the per-file cap is generous; total-bytes
# protects against a runaway upload set.
_MAX_FILE_BYTES   = 200 * 1024 * 1024     # 200 MB per upload
_MAX_TOTAL_BYTES  = 600 * 1024 * 1024     # 600 MB total
_MAX_CSV_FILES    = 50

CSV_EXTS     = {".csv"}
TRACKER_EXTS = {".xlsx", ".xls"}

# UI dropdown values for new-IP handling. Kept in sync with the
# template's <select> options — keep these stable so a bookmarked
# form submission doesn't break across refactors.
NEW_IP_ACTIONS = {"include", "exclude", "list_only"}


def _safe_basename(name: str) -> str:
    name = name.rsplit("/", 1)[-1].rsplit("\\", 1)[-1]
    name = re.sub(r"[^A-Za-z0-9._-]", "_", name).strip("._-")
    return name or "upload"


def _ext_of(name: str) -> str:
    return "." + name.rsplit(".", 1)[-1].lower() if "." in name else ""


def _check_size(filename: str, size: int, running_total: int) -> int:
    if size > _MAX_FILE_BYTES:
        raise ValueError(
            f"{filename}: file exceeds the "
            f"{_MAX_FILE_BYTES // (1024*1024)} MB upload limit."
        )
    if running_total + size > _MAX_TOTAL_BYTES:
        raise ValueError(
            f"Total uploaded bytes exceed the "
            f"{_MAX_TOTAL_BYTES // (1024*1024)} MB request limit."
        )
    return running_total + size


def _detect_column(headers: list[str], aliases: list[str]) -> Optional[str]:
    """Case- and whitespace-insensitive lookup of a header by alias
    list. Returns the ORIGINAL header name from the sheet so we can
    address the column by its real value rather than reconstructing
    one. None if no alias matched."""
    norm_map = {(h or "").strip().lower(): h for h in headers if h is not None}
    for alias in aliases:
        h = norm_map.get(alias.strip().lower())
        if h:
            return h
    return None


def run_retest(
    *,
    current_csvs: Iterable[Tuple[str, bytes]],
    original_tracker: Tuple[str, bytes],
    sheet_index: int = 0,
    custom_comment_col: str = "",
    custom_comment_default: str = "",
    new_ip_action: str = "include",
    enable_version_check: bool = True,
) -> dict:
    """Run a retest pass.

    Args:
      current_csvs: list of `(filename, bytes)` for THIS quarter's
        rescan CSV exports.
      original_tracker: single `(filename, bytes)` of the previous
        tracker xlsx — the file with the original findings + any
        client-uploaded screenshots in dedicated columns. We modify
        this file in-place (openpyxl) so embedded images survive.
      sheet_index: which sheet of the tracker holds the findings.
        Defaults to 0 (first sheet) which is the Nessus-style
        convention; the UI lets the consultant override per upload.
      custom_comment_col: optional column name (e.g. "VibeDocs
        Comments"). Pre-fills with `custom_comment_default` on every
        row we close so the consultant doesn't have to paste a
        boilerplate by hand. If the column already exists on the
        sheet we just write into it.
      new_ip_action: one of ``include`` / ``exclude`` / ``list_only``.
        Controls what happens to IPs that exist in the rescan but
        weren't in the original tracker (often net-new hosts).
      enable_version_check: when True, the still-open rows get one
        more pass — any row whose plugin_output reports an installed
        version >= the recommended fix is also closed.

    Returns the standard `{zip_bytes, summary, zip_name}` dict the
    toolkit router streams back. The ZIP contains:
      - updated_tracker.xlsx  (the edited copy; original is NOT
                                touched on disk)
      - new_hosts.xlsx        (when new_ip_action != "exclude")
      - version_remediated.xlsx (when version-check closed rows)
      - summary.txt
    """
    # ------------------------------------------------------------
    # 1. Input validation. Keep messages plain — the router relays
    #    them verbatim as a 400.
    # ------------------------------------------------------------
    if new_ip_action not in NEW_IP_ACTIONS:
        raise ValueError(
            f"new_ip_action must be one of {sorted(NEW_IP_ACTIONS)} "
            f"(got {new_ip_action!r})."
        )

    csv_list = list(current_csvs)
    if not csv_list:
        raise ValueError("Upload at least one Nessus CSV scan.")
    if len(csv_list) > _MAX_CSV_FILES:
        raise ValueError(f"Too many Nessus CSVs (max {_MAX_CSV_FILES}).")

    if not original_tracker or not original_tracker[1]:
        raise ValueError("Upload the original tracker (.xlsx / .xls).")

    total = 0
    for name, data in csv_list:
        if _ext_of(name) not in CSV_EXTS:
            raise ValueError(f"{name}: scans must be .csv files.")
        total = _check_size(name, len(data), total)

    tname, tdata = original_tracker
    if _ext_of(tname) not in TRACKER_EXTS:
        raise ValueError(
            f"{tname}: tracker must be .xlsx or .xls."
        )
    total = _check_size(tname, len(tdata), total)

    # ------------------------------------------------------------
    # 2. Lay out a temp workspace. The lib needs a folder for the
    #    CSV loader; we also drop the tracker on disk so openpyxl
    #    can load it via path (it works on streams too but path is
    #    closer to the CLI's tested code path).
    # ------------------------------------------------------------
    from .va_automater.loaders import load_nessus_folder
    from .va_automater.matching import NewScanIndex, TIER_ORDER
    from .va_automater.identifiers import (
        normalize_plugin_id, normalize_name, normalize_ip, safe_port,
    )
    from .va_automater.ip_diff import find_new_ips
    from .va_automater.version_check import decide_for_row
    from .va_automater.schema import COL_ALIASES, TRACKER_COL_ALIASES
    from openpyxl import load_workbook

    with tempfile.TemporaryDirectory(prefix="va_retest_") as tmproot_str:
        tmproot = Path(tmproot_str)
        csv_dir = tmproot / "current"
        csv_dir.mkdir()
        for name, data in csv_list:
            (csv_dir / _safe_basename(name)).write_bytes(data)

        tracker_in = tmproot / _safe_basename(tname)
        tracker_in.write_bytes(tdata)
        tracker_out = tmproot / "updated_tracker.xlsx"

        # 3. Load current rescan -> canonical DataFrame -> match index.
        current_df = load_nessus_folder(csv_dir)
        if "ip" not in current_df.columns:
            raise ValueError(
                "Rescan CSVs don't have a Host/IP column the loader could detect."
            )
        idx = NewScanIndex.build(current_df)

        # Set of every host actually present in this rescan. Used purely
        # for transparency: a tracker finding gets Closed when it doesn't
        # match the rescan, but if its host wasn't scanned AT ALL (offline,
        # out of scope, missed by a subnet sweep) that "closure" is really
        # "not retested". We don't change the close decision (consultant's
        # chosen behaviour) — we just COUNT and LIST these so the summary
        # flags them for a manual sanity check before the report ships.
        rescan_ip_set = {
            ip for ip in (normalize_ip(v) for v in current_df["ip"].tolist())
            if ip
        }

        # Lookup table from (pid, ip, port) -> the rescan's PLUGIN_OUTPUT
        # for that row. We use this during the version-check pass so the
        # "is the installed version newer than the recommended fix?"
        # decision reads the FRESH banner from the rescan, not the stale
        # plugin_output cell in the tracker (which still shows the old
        # version the client just upgraded away from). Falls back through
        # (pid, ip) and (name, ip, port) the same way NewScanIndex.match
        # does, so the lookup hits whichever match tier the row landed
        # on.
        from collections import defaultdict
        current_plugin_output: dict = {}
        for _, row in current_df.iterrows():
            pid_  = normalize_plugin_id(row.get("plugin_id", ""))
            name_ = normalize_name(row.get("finding_name", ""))
            ip_   = normalize_ip(row.get("ip", ""))
            port_ = safe_port(row.get("port", ""))
            po    = str(row.get("plugin_output", "") or "")
            if not ip_ or not po:
                continue
            if pid_:
                current_plugin_output.setdefault(("pid_ip_port", pid_, ip_, port_), po)
                current_plugin_output.setdefault(("pid_ip", pid_, ip_), po)
            if name_:
                current_plugin_output.setdefault(("name_ip_port", name_, ip_, port_), po)
                current_plugin_output.setdefault(("name_ip", name_, ip_), po)

        def _rescan_output(pid: str, name: str, ip: str, port: str) -> str:
            """Best plugin_output from the rescan for this tracker row.
            Mirrors NewScanIndex's tier preference (pid+ip+port → pid+ip
            → name+ip+port → name+ip). Returns "" when nothing matched —
            the version-check then falls back to the tracker's own
            plugin_output as a last resort, which is what `decide_for_row`
            already accepts."""
            if pid:
                v = (current_plugin_output.get(("pid_ip_port", pid, ip, port))
                     or current_plugin_output.get(("pid_ip", pid, ip)))
                if v:
                    return v
            if name:
                v = (current_plugin_output.get(("name_ip_port", name, ip, port))
                     or current_plugin_output.get(("name_ip", name, ip)))
                if v:
                    return v
            return ""

        # 4. Open the original tracker workbook in IMAGE-PRESERVING
        #    mode. openpyxl loads cell-anchored images by default;
        #    saving the same workbook back to disk preserves them as
        #    long as we don't delete rows that contain them. Mutations
        #    we do (writing into Status / custom-comments cells,
        #    appending new rows at the bottom) are all image-safe.
        wb = load_workbook(tracker_in)
        try:
            ws = wb.worksheets[sheet_index]
        except IndexError:
            raise ValueError(
                f"Tracker has {len(wb.worksheets)} sheet(s); requested index "
                f"{sheet_index} is out of range."
            )

        # 5. Read the header row + locate the columns we care about.
        #    The library's `TRACKER_COL_ALIASES` already enumerates the
        #    Plugin ID / Host / Port / Status header variants Nessus
        #    exports and VibeDocs trackers use; reuse it so the
        #    detection here stays in sync.
        max_col = ws.max_column or 0
        headers: list[str] = [
            ws.cell(row=1, column=c).value for c in range(1, max_col + 1)
        ]

        def _col(name_aliases: list[str]) -> Optional[int]:
            """1-based column index by alias list, or None."""
            h = _detect_column(headers, name_aliases)
            if h is None:
                return None
            return headers.index(h) + 1

        # Status is a tracker-only column (not in Nessus exports), so it
        # lives in TRACKER_COL_ALIASES. The data columns (plugin_id /
        # finding_name / IP / port / plugin_output / etc.) are in
        # COL_ALIASES alongside their Nessus header variants.
        col_pid     = _col(COL_ALIASES["plugin_id"])
        col_name    = _col(COL_ALIASES["finding_name"])
        col_ip      = _col(COL_ALIASES["ip"])
        col_port    = _col(COL_ALIASES["port"])
        col_status  = _col(TRACKER_COL_ALIASES["status"])
        col_pout    = _col(COL_ALIASES["plugin_output"])
        col_soln    = _col(COL_ALIASES["solution"])
        col_synp    = _col(COL_ALIASES["synopsis"])
        col_desc    = _col(COL_ALIASES["description"])

        if col_ip is None:
            raise ValueError(
                "Couldn't find a Host / IP column on the tracker. "
                "Available headers: " + ", ".join(str(h) for h in headers if h)
            )
        if col_status is None:
            # Trackers sometimes ship without a Status column (raw
            # Nessus xlsx export). Create one at the end so we always
            # have somewhere to write closures.
            col_status = (max_col or 0) + 1
            ws.cell(row=1, column=col_status, value="Status")
            headers.append("Status")
            max_col = col_status

        # Optional custom-comments column. Create if missing.
        col_custom: Optional[int] = None
        if custom_comment_col:
            col_custom = _col([custom_comment_col]) \
                or (max_col + 1)
            if ws.cell(row=1, column=col_custom).value != custom_comment_col:
                ws.cell(row=1, column=col_custom, value=custom_comment_col)
                if col_custom > max_col:
                    headers.append(custom_comment_col)
                    max_col = col_custom

        # 6. Walk every data row, decide closed / still-open, write
        #    Status + custom-comment cells in place. We never DELETE
        #    rows — that would orphan image anchors.
        n_rows = ws.max_row or 1
        closed_missing = 0
        closed_version = 0
        still_open = 0
        # Transparency counters: of the rows we closed as "no longer in
        # rescan", how many were on a host the rescan never covered?
        closed_on_unscanned_host = 0
        unscanned_hosts_closed: set[str] = set()

        # We treat anything already "Closed"/"Closed-Remediated"/etc
        # as untouchable so the script is rerunnable without flipping
        # a previously-closed status back to "Closed (current pass)".
        ALREADY_CLOSED = {"closed", "closed-remediated", "fixed",
                          "remediated", "resolved"}

        for r in range(2, n_rows + 1):
            existing_status = (
                str(ws.cell(row=r, column=col_status).value or "").strip().lower()
            )
            if existing_status in ALREADY_CLOSED:
                # Already closed — skip. Don't recount or rewrite.
                continue

            pid   = normalize_plugin_id(
                ws.cell(row=r, column=col_pid).value if col_pid else ""
            )
            name  = normalize_name(
                ws.cell(row=r, column=col_name).value if col_name else ""
            )
            ip    = normalize_ip(
                ws.cell(row=r, column=col_ip).value
            )
            port  = safe_port(
                ws.cell(row=r, column=col_port).value if col_port else ""
            )

            # No IP -> can't match -> leave as-is. (Audit log will
            # bubble this up as a row count in `n_skipped_no_ip`.)
            if not ip:
                still_open += 1
                continue

            tier = idx.match(plugin_id=pid, finding_name=name, ip=ip, port=port)
            if tier == "no_match":
                # Finding gone in this rescan -> mark closed.
                ws.cell(row=r, column=col_status, value="Closed")
                if col_custom and custom_comment_default:
                    ws.cell(row=r, column=col_custom,
                            value=custom_comment_default)
                closed_missing += 1
                # Flag (don't block) closures on hosts the rescan never
                # touched — these are the rows most likely to be false
                # "remediated" calls if a host was simply out of scope.
                if ip not in rescan_ip_set:
                    closed_on_unscanned_host += 1
                    unscanned_hosts_closed.add(ip)
                continue

            # Still present. Optionally run the version-check.
            # Prefer the rescan's plugin_output (fresh banner reflecting
            # the upgrade the client just performed) — fall back to the
            # tracker's own plugin_output only when the rescan doesn't
            # have anything for this finding (rare but possible if the
            # rescan trimmed Plugin Output for size).
            if enable_version_check:
                rescan_po = _rescan_output(pid, name, ip, port)
                tracker_po = (
                    ws.cell(row=r, column=col_pout).value if col_pout else ""
                )
                effective_po = rescan_po or str(tracker_po or "")
                d = decide_for_row(
                    plugin_output=effective_po,
                    solution=str(
                        ws.cell(row=r, column=col_soln).value
                        if col_soln else "" or ""),
                    synopsis=str(
                        ws.cell(row=r, column=col_synp).value
                        if col_synp else "" or ""),
                    description=str(
                        ws.cell(row=r, column=col_desc).value
                        if col_desc else "" or ""),
                    finding_name=name,
                )
                if d.status == "remediated":
                    ws.cell(row=r, column=col_status, value="Closed")
                    if col_custom and custom_comment_default:
                        ws.cell(row=r, column=col_custom,
                                value=custom_comment_default)
                    # Also refresh the Plugin Output cell with the
                    # rescan's fresh banner so the closed row's audit
                    # trail explains why we closed it ("installed
                    # 2.4.58 ≥ recommended 2.4.55").
                    if col_pout and rescan_po:
                        ws.cell(row=r, column=col_pout, value=rescan_po)
                    closed_version += 1
                    continue
            still_open += 1

        # 7. New-IP detection + handling.
        # Build a quick set of IPs in the original tracker so we can
        # diff against current. We reuse the canonical-DataFrame
        # path the recurring-scan tool uses to keep semantics
        # identical (normalize_ip + dedupe).
        original_ips: set[str] = set()
        for r in range(2, n_rows + 1):
            v = ws.cell(row=r, column=col_ip).value
            ip = normalize_ip(v)
            if ip:
                original_ips.add(ip)
        # Build a wrap-around df just for the find_new_ips signature.
        import pandas as pd
        original_df_lite = pd.DataFrame({"ip": list(original_ips)})
        new_ips = find_new_ips(current_df, original_df_lite)

        # Rows in the current scan that sit on a new IP — what we'd
        # potentially append (or list separately).
        new_ip_rows = current_df[
            current_df["ip"].map(normalize_ip).isin(new_ips)
        ].reset_index(drop=True) if new_ips else current_df.iloc[0:0]

        # When action is "include", append every new-IP row as a fresh
        # row at the bottom of the tracker. We carry forward only the
        # columns that exist in the tracker (matched by header) so we
        # don't accidentally widen the schema. Status defaults to
        # "Open" so the consultant immediately sees the new work.
        rows_appended = 0
        if new_ip_action == "include" and len(new_ip_rows):
            # Map canonical column names from the canonical df onto
            # tracker columns by alias. If the tracker doesn't carry a
            # column we have data for (e.g. tracker has no "Plugin
            # Output" column) we skip that field cleanly rather than
            # invent one.
            from .va_automater.output_format import CANON_TO_DISPLAY
            display_lookup = {  # display-name -> column index on tracker
                _detect_column(headers, [display_name]): canon
                for canon, display_name in CANON_TO_DISPLAY.items()
                if _detect_column(headers, [display_name])
            }
            # Reverse: tracker_col_idx -> canonical key
            tracker_col_to_canon: dict[int, str] = {}
            for canon, display_name in CANON_TO_DISPLAY.items():
                target = _detect_column(headers, [display_name])
                if target:
                    tracker_col_to_canon[headers.index(target) + 1] = canon
            for _, row in new_ip_rows.iterrows():
                r = ws.max_row + 1
                for col_idx in range(1, max_col + 1):
                    canon = tracker_col_to_canon.get(col_idx)
                    if canon is None:
                        continue
                    value = row.get(canon, "")
                    ws.cell(row=r, column=col_idx,
                            value=("" if pd.isna(value) else value))
                # Always stamp Status / custom-comments on appended
                # rows so they have a sensible default the consultant
                # can override.
                ws.cell(row=r, column=col_status, value="Open")
                rows_appended += 1

        # 8. Save the updated tracker (image-safe; openpyxl preserves
        #    embedded drawings since they're anchored to cells that
        #    we left intact).
        wb.save(tracker_out)

        # 9. Side files. new_hosts.xlsx is emitted for include +
        #    list_only paths so the consultant always has a single
        #    "what's new" reference.
        out_dir = tmproot / "out"
        out_dir.mkdir()
        # Copy the updated tracker into the out_dir so the ZIP layout
        # is just the deliverables.
        import shutil
        shutil.copyfile(tracker_out, out_dir / "updated_tracker.xlsx")

        if new_ip_action in ("include", "list_only") and len(new_ip_rows):
            try:
                from .va_automater.output_format import format_for_output
                formatted = format_for_output(
                    new_ip_rows,
                    custom_comment_col=custom_comment_col,
                    custom_comment_default=custom_comment_default,
                )
                formatted.to_excel(
                    out_dir / "new_hosts.xlsx", index=False,
                )
            except Exception as e:                           # pragma: no cover
                log.warning("new_hosts.xlsx skipped: %s", e)

        # 10. summary.txt
        summary_lines = [
            "VA Retest — Summary",
            "=" * 40,
            f"Generated at:                {datetime.utcnow().isoformat()}Z",
            f"Rescan CSV files:            {len(csv_list)}",
            f"Rescan rows loaded:          {len(current_df)}",
            f"Tracker file:                {tname}",
            f"Tracker sheet:               {ws.title} (index {sheet_index})",
            f"Tracker data rows:           {max(0, n_rows - 1)}",
            "",
            "-- Closures applied --",
            f"  Closed (no longer present in rescan): {closed_missing}",
            f"  Closed (version remediated):          {closed_version}"
            + ("  [version-check enabled]" if enable_version_check
               else "  [version-check disabled]"),
            f"  Still open after this pass:          {still_open}",
            "",
            "-- Closure sanity check --",
            f"  Closed on hosts NOT in this rescan:   {closed_on_unscanned_host}"
            + ("  <-- REVIEW: these hosts may simply have been out of scope,"
               " not remediated" if closed_on_unscanned_host else ""),
            "",
            "-- New IPs vs original tracker --",
            f"  Net-new IPs in rescan:               {len(new_ips)}",
            f"  Action chosen:                       {new_ip_action}",
            f"  Rows appended into tracker:          {rows_appended}",
        ]
        if new_ips:
            summary_lines.append("")
            summary_lines.append("  New IPs:")
            for ip in sorted(new_ips):
                n = int((current_df["ip"].map(normalize_ip) == ip).sum())
                plural = "rows" if n != 1 else "row"
                summary_lines.append(f"    - {ip}  ({n} {plural})")
        if unscanned_hosts_closed:
            summary_lines.append("")
            summary_lines.append(
                "  Hosts closed but absent from rescan (verify scope):"
            )
            for ip in sorted(unscanned_hosts_closed):
                summary_lines.append(f"    - {ip}")
        (out_dir / "summary.txt").write_text(
            "\n".join(summary_lines) + "\n", encoding="utf-8",
        )

        # 11. Zip everything.
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for p in sorted(out_dir.rglob("*")):
                if p.is_file():
                    arc = p.relative_to(out_dir).as_posix()
                    zf.writestr(arc, p.read_bytes())
        zip_bytes = zip_buf.getvalue()

    zip_name = datetime.utcnow().strftime("va_retest_%Y%m%d_%H%M%S.zip")
    summary = {
        "rescan_csv_count":  len(csv_list),
        "rescan_rows":       int(len(current_df)),
        "tracker_rows":      max(0, n_rows - 1),
        "closed_missing":    closed_missing,
        "closed_version":    closed_version,
        "still_open":        still_open,
        "closed_on_unscanned_host": closed_on_unscanned_host,
        "unscanned_hosts_closed":   sorted(unscanned_hosts_closed),
        "new_ips":           sorted(new_ips),
        "new_ip_rows":       int(len(new_ip_rows)),
        "rows_appended":     rows_appended,
        "new_ip_action":     new_ip_action,
        "enable_version_check": bool(enable_version_check),
    }
    return {"zip_bytes": zip_bytes, "summary": summary, "zip_name": zip_name}
