"""
Best-effort text extraction from uploaded engagement documents (Purchase
Orders / Requests For Information).

Two formats supported:
  * PDF (.pdf) — extracted via pypdf if installed, otherwise we just store
    the file and return an empty string.
  * Excel (.xlsx, .xls) — iterate sheets/cells with openpyxl. Captures one
    cell per line so the consultant can grep through it.

Returned text is plain-text, capped to ~250 KB so a malicious upload can't
balloon our DB row. The original file is always saved at full fidelity —
this routine only generates a *preview* string for search and quick review.
"""
from __future__ import annotations
import logging
from pathlib import Path

log = logging.getLogger(__name__)

_MAX_PREVIEW_CHARS = 256 * 1024


def extract_text(path: Path) -> str:
    """Return up to _MAX_PREVIEW_CHARS of plain text from `path`. Never raises."""
    suffix = path.suffix.lower()
    try:
        if suffix == ".pdf":
            return _extract_pdf(path)
        if suffix in (".xlsx", ".xls", ".xlsm"):
            return _extract_xlsx(path)
        # Plain text fallback
        try:
            return path.read_text(encoding="utf-8", errors="replace")[:_MAX_PREVIEW_CHARS]
        except Exception:
            return ""
    except Exception as e:
        log.warning("doc_extract failed for %s: %s", path, e)
        return ""


def _extract_pdf(path: Path) -> str:
    try:
        from pypdf import PdfReader  # pip install pypdf
    except ImportError:
        log.info("pypdf not installed; skipping PDF text extraction for %s", path)
        return ""
    out: list[str] = []
    total = 0
    reader = PdfReader(str(path))
    for page in reader.pages:
        try:
            t = page.extract_text() or ""
        except Exception:
            t = ""
        if not t:
            continue
        out.append(t)
        total += len(t)
        if total >= _MAX_PREVIEW_CHARS:
            break
    return "\n\n".join(out)[:_MAX_PREVIEW_CHARS]


def _extract_xlsx(path: Path) -> str:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return ""
    out: list[str] = []
    total = 0
    wb = load_workbook(str(path), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            out.append(f"=== Sheet: {ws.title} ===")
            for row in ws.iter_rows(values_only=True):
                line = " | ".join("" if c is None else str(c) for c in row)
                if not line.strip():
                    continue
                out.append(line)
                total += len(line) + 1
                if total >= _MAX_PREVIEW_CHARS:
                    return "\n".join(out)[:_MAX_PREVIEW_CHARS]
    finally:
        wb.close()
    return "\n".join(out)[:_MAX_PREVIEW_CHARS]
