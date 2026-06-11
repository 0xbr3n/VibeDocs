"""
Excel "Risk Register" tracker — interactive import + export.

Three endpoints back the upload UI:

  POST /api/reports/versions/{vid}/tracker/preview
       multipart: file (.xlsx)
       Parses the tracker, returns the detected column mapping + the first
       N rows so the consultant can sanity-check before committing. Stores
       the upload server-side under a one-shot preview id so commit can
       reuse it without re-uploading. Returns 400 with a structured error
       if the tracker can't be parsed at all.

  POST /api/reports/versions/{vid}/tracker/commit
       JSON: { preview_id, column_map: {field -> col_index}, sn_col? }
       Reads the previously uploaded file using the user-confirmed mapping
       and inserts findings. Returns the same summary the old one-shot
       import returned.

  GET  /api/reports/versions/{vid}/tracker/export
       Unchanged — emits an XLSX preserving the template's other sheets.
"""
from __future__ import annotations
import json
import secrets
import shutil
import tempfile
from datetime import datetime
from pathlib import Path
from typing import Optional

from fastapi import APIRouter, Body, Depends, File, HTTPException, UploadFile
from fastapi.responses import FileResponse
from pydantic import BaseModel
from sqlalchemy.orm import Session, selectinload
from sqlalchemy.orm.attributes import flag_modified

from ..database import get_db
from ..models import (
    User, Report, Project, ReportVersion, ReportFinding, AccessLevel,
    AuditLog, Severity, FindingStatus, ScanImport,
)
import logging as _logging
from ..auth import get_current_user
from ..config import settings

_log = _logging.getLogger(__name__)
from .permissions import require_access
from ..services import risk_register as rr_svc
from ..services import tracker_templates as _tpl_picker
from ..services.upload_utils import stream_save as _stream_save

_MAX_TRACKER_BYTES = 20 * 1024 * 1024   # 20 MB — Excel trackers


def _save_tracker_images(images: list[dict], findings_by_row: dict,
                          screenshot_col: Optional[int],
                          retest_col: Optional[int],
                          report_id: int) -> int:
    """Attach extracted embedded screenshots to findings.

    Strategy:
      * Each image carries `row` + `col` (1-based sheet coords).
      * `findings_by_row` maps `sheet_row -> ReportFinding`.
      * If the image's column matches the Screenshot column, attach
        to `finding.screenshots`. If it matches the Post-Review
        Screenshot column, attach to `finding.retest_evidence`. If
        neither, fall back to `screenshots` so we don't silently drop
        evidence the consultant pasted off-grid.

    Returns the number of images that were saved + attached.
    """
    if not images:
        return 0
    out_dir = Path(settings.UPLOAD_DIR) / "screenshots" / str(report_id)
    out_dir.mkdir(parents=True, exist_ok=True)
    saved = 0
    for entry in images:
        rf = findings_by_row.get(entry["row"])
        if rf is None:
            continue
        # File on disk + path string in the JSON column.
        import uuid as _uuid
        out_name = f"tracker_{_uuid.uuid4().hex}{entry['ext']}"
        out_path = out_dir / out_name
        try:
            out_path.write_bytes(entry["bytes"])
        except Exception:                                 # pragma: no cover
            continue
        col = entry["col"]
        # Compare 1-based sheet column with 0-based stored mapping.
        if retest_col is not None and col == retest_col + 1:
            evid = list(rf.retest_evidence or [])
            evid.append(str(out_path))
            rf.retest_evidence = evid
            flag_modified(rf, "retest_evidence")
        else:
            shots = list(rf.screenshots or [])
            shots.append(str(out_path))
            rf.screenshots = shots
            flag_modified(rf, "screenshots")
        saved += 1
    return saved


router = APIRouter(tags=["tracker"])

# Preview state — in-memory map of preview_id -> {path, version_id, owner_id, created_at}.
# The actual file lives on disk so commit can run even if the worker restarts;
# this dict just records the binding. Stale entries are pruned lazily.
_PREVIEWS: dict[str, dict] = {}
_PREVIEW_TTL_SECONDS = 60 * 30   # 30 minutes


def _mgmt_comments_text(f) -> str:
    """Build the Management Comments cell from a finding's list of dated
    comments — one "[DD-MM-YYYY] <text>" block per entry (newline-separated).
    Falls back to the legacy single comment + date."""
    import datetime as _dt
    stmts = list(getattr(f, "client_statements", None) or [])
    if not stmts:
        txt = (getattr(f, "client_statement", "") or "").strip()
        if txt:
            stmts = [{"date": getattr(f, "client_statement_date", "") or "", "text": txt}]
    blocks = []
    for s in stmts:
        if not isinstance(s, dict):
            continue
        text = str(s.get("text") or "").strip()
        if not text:
            continue
        d = str(s.get("date") or "").strip()
        if d:
            try:
                d = _dt.datetime.strptime(d, "%Y-%m-%d").strftime("%d-%m-%Y")
            except ValueError:
                pass
            blocks.append(f"[{d}] {text}")
        else:
            blocks.append(text)
    return "\n".join(blocks)


def _version_with_access(db: Session, vid: int, user: User,
                          need: AccessLevel = AccessLevel.view) -> ReportVersion:
    v = db.get(ReportVersion, vid)
    if not v:
        raise HTTPException(404, "Report version not found")
    r = db.get(Report, v.report_id)
    if not r:
        raise HTTPException(404, "Parent report not found")
    require_access(db, user, r, need=need)
    return v


def _prune_previews() -> None:
    now = datetime.utcnow().timestamp()
    dead = [pid for pid, p in _PREVIEWS.items()
             if now - p["created_at"] > _PREVIEW_TTL_SECONDS]
    for pid in dead:
        try:
            Path(_PREVIEWS[pid]["path"]).unlink(missing_ok=True)
        except Exception:
            pass
        _PREVIEWS.pop(pid, None)
    # Also prune stale disk sidecars so they don't accumulate.
    tracker_root = Path(settings.UPLOAD_DIR) / "trackers"
    for meta_file in tracker_root.glob("**/preview_meta_*.json"):
        try:
            meta = json.loads(meta_file.read_text())
            if now - meta.get("created_at", 0) > _PREVIEW_TTL_SECONDS:
                Path(meta.get("path", "")).unlink(missing_ok=True)
                meta_file.unlink(missing_ok=True)
        except Exception:
            pass


def _preview_meta_path(report_id: int, preview_id: str) -> Path:
    return (Path(settings.UPLOAD_DIR) / "trackers" / str(report_id)
            / f"preview_meta_{preview_id}.json")


def _load_preview(preview_id: str, report_id: int | None = None) -> dict | None:
    """Return preview record from in-memory cache or disk sidecar.

    Disk sidecars survive worker restarts and cross-worker routing (the two
    uvicorn workers each have their own _PREVIEWS dict so a preview uploaded
    to worker-1 would fail on a commit routed to worker-2 without the disk
    fallback).
    """
    rec = _PREVIEWS.get(preview_id)
    if rec:
        return rec
    if report_id is not None:
        meta_path = _preview_meta_path(report_id, preview_id)
        if meta_path.exists():
            try:
                rec = json.loads(meta_path.read_text())
                _PREVIEWS[preview_id] = rec   # warm the local cache
                return rec
            except Exception:
                pass
    return None


# ============================================================
# Preview
# ============================================================

@router.post("/api/reports/versions/{vid}/tracker/preview")
def tracker_preview(vid: int,
                     file: UploadFile = File(...),
                     db: Session = Depends(get_db),
                     user: User = Depends(get_current_user)):
    """First step of the import walkthrough. Returns the detected column
    mapping + a sample of rows the user can eyeball; the actual write only
    happens at /commit.

    The uploaded file is staged on the report's tracker dir so /commit can
    pick it up without a second upload. Stale preview files are pruned
    after `_PREVIEW_TTL_SECONDS`.
    """
    v = _version_with_access(db, vid, user, need=AccessLevel.edit)
    if not (file.filename or "").lower().endswith((".xlsx", ".xls", ".xlsm")):
        raise HTTPException(400, "Tracker must be an Excel file (.xlsx / .xls / .xlsm)")

    _prune_previews()
    track_dir = Path(settings.UPLOAD_DIR) / "trackers" / str(v.report_id)
    track_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%S")
    preview_id = secrets.token_urlsafe(12)
    stored = track_dir / f"preview_{stamp}_{preview_id}.xlsx"
    _stream_save(file.file, stored, max_bytes=_MAX_TRACKER_BYTES)
    meta = {
        "path": str(stored),
        "version_id": v.id,
        "owner_id": user.id,
        "created_at": datetime.utcnow().timestamp(),
        "original_name": file.filename,
    }
    _PREVIEWS[preview_id] = meta
    # Write disk sidecar so commit survives worker restarts / cross-worker routing.
    try:
        _preview_meta_path(v.report_id, preview_id).write_text(json.dumps(meta))
    except Exception:
        pass

    # First pass: read the workbook ONCE with no override so the auto-detected
    # column_map and headers come back. We do a second cheap pass to also
    # surface RAW cell values for every sample row, indexed by column number.
    # The client uses that raw payload to re-render the preview locally when
    # the user changes a column mapping — no round-trip needed.
    try:
        parsed = rr_svc.parse_risk_register(stored)
    except Exception as e:
        _PREVIEWS.pop(preview_id, None)
        try: stored.unlink(missing_ok=True)
        except Exception: pass
        _log.warning("Tracker parse failed for upload: %s", e)
        raise HTTPException(400, "Could not read the tracker file. Please ensure it is a valid .xlsx workbook.")

    SAMPLE_LIMIT = 25
    CELL_CAP = 240

    # Pull raw rows from the workbook for the preview, keyed by 0-based column
    # index. This is what enables live-remap on the client.
    try:
        raw_sample = _read_raw_sample(stored, parsed.get("sheet"),
                                       parsed.get("header_row") or 0,
                                       SAMPLE_LIMIT, CELL_CAP)
    except Exception:
        raw_sample = []

    return {
        "preview_id": preview_id,
        "sheet": parsed["sheet"],
        "header_row": parsed["header_row"],
        "headers": parsed["headers"],
        "column_map": parsed["column_map"],
        "sn_col": parsed.get("sn_col"),
        "row_count": len(parsed["rows"]),
        # Raw cell values per data row. Each entry is
        #   { "_sheet_row": <int>, "cells": [<col0>, <col1>, ...] }
        # The client maps cells[col_map[<field>]] to the right preview column
        # when the user adjusts a dropdown.
        "raw_sample": raw_sample,
        "expires_in_seconds": _PREVIEW_TTL_SECONDS,
        "field_options": [
            "title", "severity", "cvss_score", "cvss_vector",
            "description", "impact", "remediation", "references",
            "affected_asset", "poc_steps", "status", "retest_notes",
            "client_statement", "owasp_category", "cwe",
        ],
    }


def _read_raw_sample(path: Path, sheet_name: str, header_row: int,
                      limit: int, cell_cap: int) -> list[dict]:
    """Return up to `limit` rows below `header_row` as raw cell-arrays.

    We deliberately do NOT apply column-mapping or finding-field naming here —
    the caller wants the unprojected data so the UI can re-render previews
    against whatever mapping the user picks next."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    try:
        if not sheet_name or sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        out = []
        for row_idx, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, values_only=True),
                start=header_row + 1):
            if all(c is None or str(c).strip() == "" for c in row):
                # Tracker data block ended.
                if out: break
                continue
            cells = []
            for c in row:
                if c is None:
                    cells.append(None)
                else:
                    s = str(c)
                    cells.append(s if len(s) <= cell_cap else s[:cell_cap] + "…")
            out.append({"_sheet_row": row_idx, "cells": cells})
            if len(out) >= limit:
                break
        return out
    finally:
        wb.close()


# ============================================================
# Commit
# ============================================================

class CommitBody(BaseModel):
    preview_id: str
    column_map: dict[str, int] | None = None
    sn_col: int | None = None


@router.post("/api/reports/versions/{vid}/tracker/commit")
def tracker_commit(vid: int,
                    body: CommitBody,
                    db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    """Second step of the import walkthrough. Uses the file staged at
    /preview together with the user-confirmed column mapping. Idempotency
    is left to the caller — a re-commit will create duplicate rows.
    """
    v = _version_with_access(db, vid, user, need=AccessLevel.edit)
    _prune_previews()
    rec = _load_preview(body.preview_id, report_id=v.report_id)
    if not rec:
        raise HTTPException(404, "Preview expired or unknown. Re-upload the tracker.")
    if rec["version_id"] != v.id:
        raise HTTPException(403, "Preview was uploaded under a different report version")
    if rec["owner_id"] != user.id:
        raise HTTPException(403, "Preview belongs to a different user")

    stored = Path(rec["path"])
    if not stored.exists():
        _PREVIEWS.pop(body.preview_id, None)
        raise HTTPException(410, "Preview file is missing on disk. Re-upload.")

    # The override mapping the user supplied. _sn lives on `sn_col` if any.
    override: dict[str, int] = dict(body.column_map or {})
    if body.sn_col is not None:
        override["_sn"] = int(body.sn_col)

    try:
        parsed = rr_svc.parse_risk_register(stored, override_mapping=override or None)
    except Exception as e:
        _log.warning("Tracker parse failed with custom mapping: %s", e)
        raise HTTPException(400, "Could not read the tracker file with that column mapping.")

    # Build a lookup of existing findings in this version keyed by
    # normalised title so re-imports update rather than duplicate.
    existing_findings = db.query(ReportFinding).filter(
        ReportFinding.report_version_id == v.id
    ).all()
    existing_by_title: dict[str, ReportFinding] = {
        (rf.title or "").strip().lower(): rf for rf in existing_findings
    }

    created: list[int] = []
    updated: list[int] = []
    skipped: list[dict] = []
    # Map sheet-row -> ReportFinding so the image extractor below can
    # attach embedded screenshots to the correct row's finding.
    findings_by_row: dict[int, ReportFinding] = {}
    # Collect distinct DT Tester names across rows so we can store them
    # at the report-level (report.details.tester_names) — the docx
    # generator uses that list, and the report-edit page renders it in
    # the "Tester names" field. Per-row tester_name itself never lands
    # on ReportFinding because there's no column for it; collecting
    # is the only way to round-trip the value.
    tester_seen: list[str] = []
    for f in parsed["rows"]:
        title = f.get("title")
        if not title:
            skipped.append({"row": f.get("_sheet_row"), "reason": "no title"})
            continue
        sev_val = f.get("severity") or "Medium"
        try: severity = Severity(sev_val)
        except ValueError: severity = Severity.medium
        status_val = (f.get("status") or "Open").strip()
        try: status = FindingStatus(status_val)
        except ValueError: status = FindingStatus.open
        cvss_score = f.get("cvss_score")
        try: cvss_score = float(cvss_score) if cvss_score is not None else None
        except (ValueError, TypeError): cvss_score = None
        tn = f.get("tester_name")
        if isinstance(tn, str) and tn and tn not in tester_seen:
            tester_seen.append(tn)

        title_key = str(title).strip().lower()
        existing_rf = existing_by_title.get(title_key)

        if existing_rf:
            # UPDATE the matching finding — keep any engagement-specific
            # fields that were blank in the tracker (don't clobber with None).
            existing_rf.severity      = severity
            existing_rf.status        = status
            if f.get("description"):    existing_rf.description     = f["description"]
            if f.get("impact"):         existing_rf.impact          = f["impact"]
            if f.get("remediation"):    existing_rf.remediation     = f["remediation"]
            if f.get("references"):     existing_rf.references      = f["references"]
            if f.get("affected_asset"): existing_rf.affected_asset  = f["affected_asset"]
            if f.get("poc_steps"):      existing_rf.poc_steps       = f["poc_steps"]
            if f.get("cvss_vector"):    existing_rf.cvss_vector     = f["cvss_vector"]
            if cvss_score is not None:  existing_rf.cvss_score      = cvss_score
            if f.get("retest_notes"):   existing_rf.retest_notes    = f["retest_notes"]
            if f.get("client_statement"): existing_rf.client_statement = f["client_statement"]
            if f.get("cwe"):            existing_rf.cwe             = f["cwe"]
            existing_rf.source_ref = stored.name
            db.flush()
            rf = existing_rf
            updated.append(rf.id)
        else:
            rf = ReportFinding(
                report_version_id=v.id,
                title=str(title)[:500],
                description=f.get("description") or None,
                impact=f.get("impact") or None,
                remediation=f.get("remediation") or None,
                references=f.get("references") or None,
                affected_asset=f.get("affected_asset") or None,
                poc_steps=f.get("poc_steps") or None,
                severity=severity,
                cvss_vector=f.get("cvss_vector") or None,
                cvss_score=cvss_score,
                status=status,
                retest_notes=f.get("retest_notes") or None,
                client_statement=f.get("client_statement") or None,
                cwe=f.get("cwe") or None,
                added_by_id=user.id,
                source="tracker",
                source_ref=stored.name,
            )
            db.add(rf); db.flush()
            created.append(rf.id)

        sheet_row = f.get("_sheet_row")
        if sheet_row is not None:
            findings_by_row[int(sheet_row)] = rf

    # Pull every screenshot embedded in the Risk Register sheet and
    # attach it to the matching finding's screenshots / retest evidence
    # list. The image extractor opens the workbook in full mode (the
    # parser uses read-only which drops embedded images), so this is a
    # second pass over the same file.
    images_saved = 0
    try:
        imgs = rr_svc.extract_risk_register_images(stored)
        if imgs:
            images_saved = _save_tracker_images(
                imgs, findings_by_row,
                screenshot_col=parsed.get("screenshot_col"),
                retest_col=parsed.get("retest_screenshot_col"),
                report_id=v.report_id,
            )
    except Exception as e:                                # pragma: no cover
        # Don't fail the whole import on an image-extraction hiccup —
        # findings are already written; surface in audit but proceed.
        import logging
        logging.getLogger(__name__).warning(
            "tracker image extraction failed: %s", e)

    # Stamp the report so the next /export uses this XLSX as the template
    # (preserves the consultant's other sheets).
    report = db.get(Report, v.report_id)
    details = dict(report.details or {})
    details["tracker_template_path"] = str(stored)
    details["tracker_imported_at"] = datetime.utcnow().isoformat() + "Z"
    details["tracker_imported_by_id"] = user.id
    # Merge any newly-seen tester names with the existing list so a
    # subsequent import doesn't blow away manually-added testers.
    if tester_seen:
        existing_testers = list(details.get("tester_names") or [])
        for name in tester_seen:
            if name not in existing_testers:
                existing_testers.append(name)
        details["tester_names"] = existing_testers
    report.details = details
    flag_modified(report, "details")

    db.add(AuditLog(actor_id=user.id, action="tracker.commit",
                    object_type="report_version", object_id=v.id,
                    detail={"created": len(created),
                            "updated": len(updated),
                            "skipped": len(skipped),
                            "preview_id": body.preview_id,
                            "column_map": override,
                            "testers_captured": tester_seen,
                            "images_attached": images_saved}))
    db.commit()

    # Burn the preview record (file is now the canonical tracker template).
    # Remove both in-memory entry and disk sidecar.
    _PREVIEWS.pop(body.preview_id, None)
    try:
        _preview_meta_path(v.report_id, body.preview_id).unlink(missing_ok=True)
    except Exception:
        pass

    return {
        "ok": True,
        "version_id": v.id,
        "created_count": len(created),
        "updated_count": len(updated),
        "created_finding_ids": created,
        "updated_finding_ids": updated,
        "skipped": skipped,
    }


# ============================================================
# Legacy one-shot import (kept for backwards compat with prior UI)
# ============================================================

@router.post("/api/reports/versions/{vid}/tracker/import")
def import_tracker_legacy(vid: int,
                           file: UploadFile = File(...),
                           db: Session = Depends(get_db),
                           user: User = Depends(get_current_user)):
    """One-shot import: parse with auto-detected mapping and commit
    immediately. Equivalent to calling /preview then /commit with no
    overrides. Retained so older clients keep working."""
    v = _version_with_access(db, vid, user, need=AccessLevel.edit)
    if not (file.filename or "").lower().endswith((".xlsx", ".xls", ".xlsm")):
        raise HTTPException(400, "Tracker must be an Excel file (.xlsx / .xls / .xlsm)")

    track_dir = Path(settings.UPLOAD_DIR) / "trackers" / str(v.report_id)
    track_dir.mkdir(parents=True, exist_ok=True)
    stamp = datetime.utcnow().strftime("%Y%m%dT%H%M%S")
    stored = track_dir / f"tracker_{stamp}.xlsx"
    _stream_save(file.file, stored, max_bytes=_MAX_TRACKER_BYTES)

    try:
        parsed = rr_svc.parse_risk_register(stored)
    except Exception as e:
        _log.warning("Tracker import parse failed: %s", e)
        raise HTTPException(400, "Could not read the tracker file. Please ensure it is a valid .xlsx workbook.")

    created: list[int] = []
    skipped: list[dict] = []
    tester_seen: list[str] = []
    findings_by_row: dict[int, ReportFinding] = {}
    for f in parsed["rows"]:
        title = f.get("title")
        if not title:
            skipped.append({"row": f.get("_sheet_row"), "reason": "no title"})
            continue
        sev_val = f.get("severity") or "Medium"
        try: severity = Severity(sev_val)
        except ValueError: severity = Severity.medium
        status_val = (f.get("status") or "Open").strip()
        try: status = FindingStatus(status_val)
        except ValueError: status = FindingStatus.open
        cvss_score = f.get("cvss_score")
        try: cvss_score = float(cvss_score) if cvss_score is not None else None
        except (ValueError, TypeError): cvss_score = None
        tn = f.get("tester_name")
        if isinstance(tn, str) and tn and tn not in tester_seen:
            tester_seen.append(tn)

        rf = ReportFinding(
            report_version_id=v.id,
            title=str(title)[:500],
            description=f.get("description") or None,
            impact=f.get("impact") or None,
            remediation=f.get("remediation") or None,
            references=f.get("references") or None,
            affected_asset=f.get("affected_asset") or None,
            poc_steps=f.get("poc_steps") or None,
            severity=severity,
            cvss_vector=f.get("cvss_vector") or None,
            cvss_score=cvss_score,
            status=status,
            retest_notes=f.get("retest_notes") or None,
            client_statement=f.get("client_statement") or None,
            added_by_id=user.id,
            source="tracker",
            source_ref=stored.name,
        )
        db.add(rf); db.flush()
        created.append(rf.id)
        sheet_row = f.get("_sheet_row")
        if sheet_row is not None:
            findings_by_row[int(sheet_row)] = rf

    # Same embedded-screenshot pass as the walkthrough commit.
    images_saved = 0
    try:
        imgs = rr_svc.extract_risk_register_images(stored)
        if imgs:
            images_saved = _save_tracker_images(
                imgs, findings_by_row,
                screenshot_col=parsed.get("screenshot_col"),
                retest_col=parsed.get("retest_screenshot_col"),
                report_id=v.report_id,
            )
    except Exception as e:                                # pragma: no cover
        import logging
        logging.getLogger(__name__).warning(
            "tracker image extraction failed: %s", e)

    report = db.get(Report, v.report_id)
    details = dict(report.details or {})
    details["tracker_template_path"] = str(stored)
    details["tracker_imported_at"] = datetime.utcnow().isoformat() + "Z"
    details["tracker_imported_by_id"] = user.id
    if tester_seen:
        existing_testers = list(details.get("tester_names") or [])
        for name in tester_seen:
            if name not in existing_testers:
                existing_testers.append(name)
        details["tester_names"] = existing_testers
    report.details = details
    flag_modified(report, "details")

    db.add(AuditLog(actor_id=user.id, action="tracker.import",
                    object_type="report_version", object_id=v.id,
                    detail={"created": len(created), "skipped": len(skipped),
                            "stored_at": str(stored),
                            "testers_captured": tester_seen,
                            "images_attached": images_saved}))
    db.commit()
    return {
        "ok": True,
        "version_id": v.id,
        "created_count": len(created),
        "created_finding_ids": created,
        "skipped": skipped,
        "tracker_path": str(stored),
        "images_attached": images_saved,
    }


# ============================================================
# Export (unchanged)
# ============================================================

@router.get("/api/reports/versions/{vid}/tracker/export")
def export_tracker(vid: int,
                    db: Session = Depends(get_db),
                    user: User = Depends(get_current_user)):
    """Generate an Excel Risk Register tracker from this version.

    Template-selection priority:
      1. The XLSX the consultant uploaded into this report (stored under
         report.details.tracker_template_path) — preferred because it
         already carries their client-specific tweaks (sheet adds,
         Scoping Questionnaire answers, etc.).
      2. The bundled VibeDocs template that matches the report's
         template code (Web / API / Cloud / Network / Mobile / SCR /
         Thick Client) under
         `services/tracker_templates.pick_tracker_template(code)`.
      3. None — `write_risk_register` falls back to a synthesised
         single-sheet workbook so the caller still gets *something*.

    Output filename follows the team's convention:
        "<report name> <Type> Tracking List v<version>.xlsx"
    e.g. "test API VAPT Tracking List v0.1.xlsx".

    Tester names from the report (`details.tester_names`) are written
    into every row's DT Tester column on the way out — the reverse of
    the import-time aggregation.
    """
    v = _version_with_access(db, vid, user, need=AccessLevel.view)
    # Eagerly load the library relationship so owasp_category lookups
    # below (for OWASP column + checklist auto-populate) don't each fire
    # a separate SELECT per finding (N+1 problem on large reports).
    from ..models import ReportFinding as _RF
    db.query(_RF).filter(_RF.report_version_id == v.id).options(
        selectinload(_RF.library)
    ).all()
    report = db.get(Report, v.report_id)
    template_code = (report.template.code if report and report.template else "") or ""
    details = report.details or {}
    tester_names = list(details.get("tester_names") or [])
    # Comma-separate so the DT Tester cell reads naturally with multiple
    # consultants on the engagement.
    tester_label = ", ".join(t for t in tester_names if t) or None
    # Client Owner / POC. Precedence:
    #   1. report.details.client_owner (set on the Report-details edit page)
    #   2. project.details.client_poc.name (engagement-level POC)
    #   3. "" (column stays blank — nothing to fill)
    client_owner = (details.get("client_owner") or "").strip()
    if not client_owner and report and report.project:
        poc = (getattr(report.project, "details", None) or {}).get("client_poc") or {}
        if isinstance(poc, dict):
            client_owner = (poc.get("name") or "").strip()

    # "Date Raised" — the date the consultant marks as when this report
    # was first issued to the client. Stored on report.details.report_date
    # as a plain "YYYY-MM-DD" string (HTML5 <input type="date"> format).
    # Repeated into every tracker row's Date Raised column so the
    # downstream Risk Register reads one consistent value per export.
    # If nothing is set yet we leave the column blank rather than
    # guessing, so the consultant notices and fills it in.
    report_date = (details.get("report_date") or "").strip()

    # Application name — written into the per-row "System" column so every
    # tracker row identifies the application under test without the consultant
    # having to fill it in manually.
    application_name = (details.get("application_name") or "").strip()

    # Review type label — maps the report template code to the short text
    # that appears in the "Area of Review" column ("Web PT", "API VAPT", …).
    _REVIEW_TYPE_LABELS: dict[str, str] = {
        "web_vapt":           "Web VAPT",
        "api_vapt":           "API VAPT",
        "infra_vapt":         "Network VAPT",
        "infra_va":           "Network VA",
        "mobile_pt":          "Mobile VAPT",
        "thick_client_pt":    "Thick Client VAPT",
        "aws_cloud_vapt":     "Cloud VAPT",
        "azure_cloud_vapt":   "Cloud VAPT",
        "source_code_review": "Source Code Review",
        "kiosk_pt":           "Kiosk PT",
        "wifi_pt":            "Wi-Fi PT",
        "ot_vapt":            "OT VAPT",
    }
    review_type_label = _REVIEW_TYPE_LABELS.get(template_code.lower(), "")

    # Determine if this is a retest or report-update version.
    # Prefer the explicit report_type field set in Report Details; fall back
    # to the version string heuristic (anything beyond 0.x is a retest).
    version_str = v.version or "0.1"
    report_type_val = (details.get("report_type") or "").strip()
    is_retest = (
        report_type_val in ("Retest Report", "Report Update")
        or version_str not in ("0.1", "v0.1")
    )
    # DT Tester2 — same tester label as the primary, written only on
    # retest/update versions so the initial-report export leaves the
    # column blank.
    retest_tester_label = tester_label if is_retest else ""
    # Date Follow-Up — the report date stamped into the follow-up date column
    # for retest / update versions.
    date_follow_up = report_date if is_retest else ""

    sev_rank = {Severity.critical: 0, Severity.high: 1, Severity.medium: 2,
                Severity.low: 3, Severity.informational: 4}
    findings = sorted(v.findings,
                       key=lambda f: (sev_rank.get(f.severity, 9),
                                      (f.title or "").lower()))
    dicts = []
    # Grouped infra findings (Outdated patches / SSL misconfigs / …) each get
    # a dedicated tracker sheet built from their categorised xlsx attachment.
    group_sheets: list[dict] = []
    for idx, f in enumerate(findings, start=1):
        # Per-finding attachments (Infra Scan Pipeline categorised
        # workbooks). When the finding carries one, the description
        # is suffixed with a clear pointer line so the consultant
        # reading the tracker knows where to find the per-host data.
        # The Word renderer already shows the attachment as an
        # inline icon in the Observations section; the tracker row
        # mirrors that hint in prose.
        atts = list(getattr(f, "attachments", None) or [])
        att_suffix = ""
        if atts:
            names = [a.get("filename") for a in atts
                     if isinstance(a, dict) and a.get("filename")]
            if names:
                att_suffix = (
                    "\n\nRefer to the attached file"
                    + ("s" if len(names) > 1 else "")
                    + ": " + ", ".join(names)
                )
            # Collect the per-host xlsx so the exporter builds a sheet for it.
            for a in atts:
                if not isinstance(a, dict):
                    continue
                p = a.get("path")
                fn = str(a.get("filename") or "")
                if p and Path(p).exists() and fn.lower().endswith(".xlsx"):
                    # Pass the grouped finding's CURRENT severity so that if the
                    # consultant re-rated it in VibeDocs, every row of the per-service
                    # detail sheet reflects the new severity (not the stale
                    # attachment value).
                    group_sheets.append({
                        "title": f.title,
                        "xlsx_path": p,
                        "severity": (f.severity.value if getattr(f, "severity", None)
                                     else None),
                    })
                    break

        # Informational findings are advisory, not defects — their tracker
        # Status reads "NA" (mirrors the Word report). Severity stays the
        # canonical "Informational" so the SEVERITY_DISPLAY ("Info") + fill
        # lookups still resolve; only the displayed STATUS changes.
        _is_info = bool(f.severity and f.severity.value == "Informational")
        dicts.append({
            "index": idx,
            "title": f.title,
            "severity": f.severity.value if f.severity else "Informational",
            "cvss_score": f.cvss_score,
            "cvss_vector": f.cvss_vector or "",
            "status": "NA" if _is_info else (f.status.value if f.status else "Open"),
            "affected_asset": f.affected_asset or "",
            "description": (f.description or "") + att_suffix,
            "impact": f.impact or "",
            "poc_steps": f.poc_steps or "",
            "remediation": f.remediation or "",
            "references": f.references or "",
            # CWE classification of the finding — emitted into the CWE ID
            # column of the Risk Register. Inherits from the library
            # finding's CWE on add but can be overridden per-report on
            # the finding card.
            "cwe": getattr(f, "cwe", None) or "",
            # OWASP category from the linked library finding — written into
            # the "OWASP Top 10" column of the Risk Register. Findings added
            # manually (no library link) leave the column blank.
            "owasp_category": (
                (getattr(f.library, "owasp_category", None) or "")
                if getattr(f, "library", None) else ""
            ),
            # Per-finding screenshots embedded into the Screenshot
            # column of the Risk Register. Each entry is either a
            # legacy string path or a `{path, caption}` dict — the
            # writer accepts both. Files that don't exist on disk are
            # silently skipped so a missing asset never blocks the
            # export. Captions render as plain text in the same cell.
            "screenshots": f.screenshots or [],
            "retest_notes": f.retest_notes or "",
            "client_statement": _mgmt_comments_text(f),
            # Populated per-row so the writer fills the DT Tester column
            # with the engagement tester(s). Falls back to the row's
            # `added_by_id` user lookup below if no report-level testers
            # are configured.
            "tester_name": "" if is_retest else (tester_label or _tester_for_finding(db, f) or ""),
            # Client Owner / POC name — the same value on every row.
            # Resolved above from report.details.client_owner or the
            # project POC fallback.
            "client_owner": client_owner,
            # Same date stamped on every row. Empty string when the
            # consultant hasn't filled the field in yet — that's nicer
            # than an "1970-01-01" or "None" leaking into the deliverable.
            "date_raised": report_date,
            # Application name — written into the per-row "System" column.
            "system": application_name,
            # Review type — e.g. "Web PT" — fills the "Area of Review" col.
            "review_type": review_type_label,
            # DT Tester2 — retest tester name; blank for initial reports.
            "retest_tester_name": retest_tester_label,
            # Date Follow-Up — report date for retest/update versions.
            "date_follow_up": date_follow_up,
            # Retest evidence embedded into Post Review Screenshot column.
            "retest_screenshots": list(f.retest_evidence or []),
        })

    # ---- Resolve which .xlsx template to layout-source from ----
    uploaded_tmpl = details.get("tracker_template_path")
    uploaded_tmpl = Path(uploaded_tmpl) if uploaded_tmpl else None
    if uploaded_tmpl and uploaded_tmpl.exists():
        template_path: Optional[Path] = uploaded_tmpl
        template_source = "uploaded"
    else:
        bundled = _tpl_picker.pick_tracker_template(template_code)
        template_path = bundled
        template_source = "bundled" if bundled else "synthesised"
        if not bundled:
            # Loud warning — falling back to the synthesised flat
            # layout is almost always a deployment problem (the
            # `report-templates/` directory wasn't mounted into the
            # container). The previous version silently used
            # synthesis, which made the deliverable look wrong to
            # consultants without any log signal pointing at the
            # cause.
            import logging
            logging.getLogger(__name__).warning(
                "Tracker export for report_version=%s falling back to synthesised "
                "layout — no bundled template matched code=%r. "
                "Hit GET /api/admin/tracker-templates/diagnose for details.",
                v.id, template_code,
            )

    # ---- Output filename + path ----
    out_dir = Path(settings.REPORT_DIR) / str(v.report_id) / "trackers"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_name = _tpl_picker.output_filename(
        report.name or f"report_{v.report_id}",
        template_code,
        v.version,
    )
    out_path = out_dir / out_name

    # ---- Build Info-sheet context from project / report / version ----
    project = report.project if report else None
    proj_details = (getattr(project, "details", None) or {}) if project else {}
    scope_targets = list(getattr(project, "scope_targets", None) or [])
    urls, ips = rr_svc._classify_scope_targets(scope_targets)

    # Date formatting helper — produces "DD-Mon-YYYY" (e.g. "01-Jan-2025")
    # which reads naturally in an Excel cell regardless of locale.
    def _fmt_date(dt) -> str:
        if dt is None:
            return ""
        if isinstance(dt, str):
            # Try to reformat ISO date strings (YYYY-MM-DD) to DD-Mon-YYYY
            # so VibeDocs-entered dates match the same display style as project dates.
            s = dt.strip()
            try:
                from datetime import datetime as _dt
                return _dt.strptime(s, "%Y-%m-%d").strftime("%d-%b-%Y")
            except ValueError:
                pass
            return s
        try:
            return dt.strftime("%d-%b-%Y")
        except Exception:
            return str(dt)

    # Parse the VibeDocs testing window ("YYYY-MM-DD to YYYY-MM-DD") into separate
    # start / end strings. Fall back to project-level dates when not set.
    _tw = (details.get("testing_window") or "").strip()
    _tw_parts = [p.strip() for p in _tw.split(" to ")] if " to " in _tw else []
    testing_start = _fmt_date(_tw_parts[0]) if len(_tw_parts) >= 1 and _tw_parts[0] else _fmt_date(getattr(project, "testing_start", None))
    testing_end   = _fmt_date(_tw_parts[1]) if len(_tw_parts) >= 2 and _tw_parts[1] else _fmt_date(getattr(project, "testing_end", None))

    # Date of Retest: only meaningful for reports beyond the initial draft
    # (v0.1). `is_retest` and `version_str` are already resolved above.
    date_of_retest = _fmt_date(v.created_at) if is_retest else ""

    # Tester email: use the report's creator as primary (they're the lead
    # tester who generated the report), fall back to the exporting user.
    report_creator = (
        db.get(User, report.created_by_id)
        if report and report.created_by_id else None
    )
    tester_email = (
        (report_creator.email if report_creator else None)
        or user.email
        or ""
    )
    # Reporter contact number — profile-level field on the report's creator
    # (falls back to the exporting user). Writes into the Info sheet's DT
    # Tester "Contact No" column.
    tester_phone = (
        (getattr(report_creator, "phone", None) if report_creator else None)
        or getattr(user, "phone", None)
        or ""
    )

    # Client PIC email from project.details.client_poc
    poc = proj_details.get("client_poc") or {}
    client_pic_email = (poc.get("email") or "").strip() if isinstance(poc, dict) else ""

    info_context = {
        "system_name":            (details.get("application_name") or getattr(project, "name", None) or report.name or ""),
        "scope_description":      (getattr(project, "scope_description", None) or ""),
        "urls":                   urls,
        "ips":                    ips,
        "testing_start":          testing_start,
        "testing_end":            testing_end,
        "date_of_retest":         date_of_retest,
        "client_pic_name":        client_owner,   # already resolved above
        "client_pic_email":       client_pic_email,
        "tester_name":            tester_label or (user.full_name or user.username),
        "tester_email":           tester_email,
        "tester_phone":           tester_phone,
        # DT Tester 2 — separate from tester_name so initial reports leave it blank
        "retest_tester_name":     retest_tester_label,
        "retest_tester_email":    tester_email if is_retest else "",
        "retest_tester_phone":    tester_phone if is_retest else "",
        # Source IP address of the tester's machine
        "source_ip":              (details.get("source_ip") or "").strip(),
        # Login credentials table: list of {role, username, password} dicts
        "login_credentials":      list(details.get("login_credentials") or []),
        # AWS / cloud account IDs (Cloud VAPT) — fill the Info sheet's
        # "Target IP / Account" field with the scanned account IDs.
        "aws_account_ids":        list(details.get("aws_account_ids") or []),
    }

    # Build the set of covered OWASP category prefixes from the report's
    # findings (e.g. {"API1", "API3"} from "API1:2023" and "API3:2023").
    # These are used to auto-populate the Test Exe. Checklist sheet.
    covered_categories: set[str] = set()
    for f in v.findings:
        # owasp_category lives on the FindingLibrary FK, not on ReportFinding.
        lf = getattr(f, "library", None)
        raw = getattr(lf, "owasp_category", None) or ""
        prefix = rr_svc._owasp_prefix(raw)
        if prefix:
            covered_categories.add(prefix)

    try:
        rr_svc.write_risk_register(
            dicts,
            template_path=template_path,
            output_path=out_path,
            info_context=info_context,
            covered_categories=covered_categories or None,
            group_sheets=group_sheets or None,
            group_row_defaults={
                "system": application_name,
                "review_type": review_type_label,
                "date_raised": report_date,
                "client_owner": client_owner,
            },
            rmm_enabled=bool(details.get("rmm_enabled", True)),
        )
    except Exception as e:
        _log.exception("Failed to write tracker for version %s: %s", vid, e)
        raise HTTPException(500, "Failed to generate the tracker file. Please try again.")

    db.add(AuditLog(actor_id=user.id, action="tracker.export",
                    object_type="report_version", object_id=v.id,
                    detail={"findings": len(dicts), "out": str(out_path),
                            "template": str(template_path) if template_path else None,
                            "template_source": template_source,
                            "report_template_code": template_code,
                            "testers_written": tester_names}))
    db.commit()

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_path.name,
    )


@router.get("/api/reports/versions/{vid}/tracker/export/cloud")
def export_cloud_tracker(
    vid: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Cloud VA/VAPT Excel tracker, built on the bundled Cloud VAPT template.

    * Pipeline (VA) findings  -> "Risk Register (VA)" sheet, one per AWS service,
      each with a dedicated per-service detail sheet (IAM / S3 / ALB / ...). The
      Affected Resource(s), Benchmark, Benchmark Clauses and Steps to Replicate
      columns read "Refer to <Service> Tab".
    * Manual (PT) findings    -> "Risk Register (PT)" sheet (all columns typed in
      VibeDocs by the consultant).
    * System column           -> the affected AWS account ID (per-row from the
      resource ARN in the service sheets; the account-ID list on the RR rows).
    * Info sheet              -> the AWS account IDs (Report Details field).
    """
    import os as _os, tempfile as _tf
    from ..services.cloud_pipeline import CLOUD_SOURCE as _CLOUD_SOURCE

    v       = _version_with_access(db, vid, user, need=AccessLevel.view)
    report  = db.get(Report, v.report_id)
    project = db.get(Project, report.project_id) if report else None
    details = report.details or {}
    template_code = (report.template.code if report and report.template else "") or ""

    findings = list(v.findings)
    va = [f for f in findings if (getattr(f, "source", "") or "") == _CLOUD_SOURCE]
    pt = [f for f in findings if (getattr(f, "source", "") or "") != _CLOUD_SOURCE]

    aws_ids  = [str(a).strip() for a in (details.get("aws_account_ids") or []) if str(a).strip()]
    aws_join = "\n".join(aws_ids)

    application  = (details.get("application_name") or (getattr(project, "name", None) or "")
                    or report.name or "")
    review_label = "Cloud VAPT"
    report_date  = (details.get("report_date") or "").strip()
    client_owner = (details.get("client_owner") or "").strip()
    tester_label = ", ".join(t for t in (details.get("tester_names") or []) if t)

    def _svc(f) -> str:
        s = (getattr(f, "source_ref", "") or "").strip()
        if s:
            return s
        t = (f.title or "").strip()
        for marker in (" Misconfigurations", " Misconfiguration", " Findings"):
            if marker in t:
                return t.split(marker)[0].strip()
        return t or "Service"

    def _common(f, idx) -> dict:
        return {
            "index":        idx,
            "title":        f.title or "",
            "severity":     (f.severity.value if getattr(f, "severity", None) else "Informational"),
            "cvss_score":   getattr(f, "cvss_score", None),
            "cvss_vector":  getattr(f, "cvss_vector", "") or "",
            "status":       (f.status.value if getattr(f, "status", None) else "Open"),
            "description":  f.description or "",
            "impact":       getattr(f, "impact", "") or "",
            "remediation":  getattr(f, "remediation", "") or "",
            "client_statement": _mgmt_comments_text(f),
            "system":       aws_join,
            "review_type":  review_label,
            "date_raised":  report_date,
            "client_owner": client_owner,
            "tester_name":  tester_label,
        }

    va_dicts: list[dict] = []
    group_sheets: list[dict] = []
    for i, f in enumerate(va, start=1):
        d   = _common(f, i)
        svc = _svc(f)
        atts = list(getattr(f, "attachments", None) or [])
        xlsx = None
        for a in atts:
            if not isinstance(a, dict):
                continue
            pth = a.get("path")
            if (pth and str(a.get("filename") or "").lower().endswith(".xlsx")
                    and Path(pth).exists()):
                xlsx = pth
                break
        if xlsx:
            refer = f"Refer to {svc} Tab"
            d.update({
                "affected_asset":    refer,
                "benchmark":         refer,
                "benchmark_clauses": refer,
                "poc_steps":         refer,
            })
            group_sheets.append({"title": svc, "xlsx_path": xlsx, "severity": d["severity"]})
        else:
            d["affected_asset"] = getattr(f, "affected_asset", "") or ""
            d["poc_steps"]      = getattr(f, "poc_steps", "") or ""
        va_dicts.append(d)

    pt_dicts: list[dict] = []
    for i, f in enumerate(pt, start=1):
        d = _common(f, i)
        d["affected_asset"] = getattr(f, "affected_asset", "") or ""
        d["poc_steps"]      = getattr(f, "poc_steps", "") or ""
        pt_dicts.append(d)

    info_context = {
        "system_name":       application,
        "scope_description": (getattr(project, "scope_description", None) or ""),
        "aws_account_ids":   aws_ids,
        "tester_name":       tester_label or (user.full_name or user.username),
        "client_pic_name":   client_owner,
        "testing_start":     (details.get("testing_window") or "").strip(),
    }

    template_path = _tpl_picker.pick_tracker_template(template_code)
    if not template_path or not Path(template_path).exists():
        raise HTTPException(
            500,
            "Cloud VAPT tracker template not found in report-templates/. "
            "Expected 'XXX Cloud VAPT Tracking List v0.1*.xlsx'.",
        )

    out_dir = Path(settings.REPORT_DIR) / str(v.report_id) / "trackers"
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / _tpl_picker.output_filename(
        report.name or f"report_{v.report_id}", template_code, v.version)

    defaults = {
        "system":       aws_join,
        "review_type":  review_label,
        "date_raised":  report_date,
        "client_owner": client_owner,
    }

    _fd, _tmp = _tf.mkstemp(suffix=".xlsx"); _os.close(_fd)
    _tmp = Path(_tmp)
    try:
        rr_svc.write_risk_register(
            va_dicts,
            template_path=template_path,
            output_path=_tmp,
            info_context=info_context,
            group_sheets=group_sheets or None,
            group_row_defaults=defaults,
            target_sheet="Risk Register (VA)",
            set_observation_pointer=False,
        )
        rr_svc.write_risk_register(
            pt_dicts,
            template_path=_tmp,
            output_path=out_path,
            group_row_defaults=defaults,
            target_sheet="Risk Register (PT)",
            # Strip RMM columns on the final pass (covers VA + PT + per-service).
            rmm_enabled=bool(details.get("rmm_enabled", True)),
        )
    except Exception as e:
        _log.exception("Failed to write cloud tracker for version %s: %s", vid, e)
        raise HTTPException(500, "Failed to generate the cloud tracker file. Please try again.")
    finally:
        try:
            _tmp.unlink()
        except OSError:
            pass

    # Paste any uploaded IAMActionHunter privilege-escalation CSV(s) verbatim
    # into their own new worksheet(s). Uses the most recent cloud_iam import for
    # this version. Best-effort — a failure here must not block the tracker.
    iam_sheets = 0
    try:
        _iam_si = (db.query(ScanImport)
                     .filter(ScanImport.report_version_id == v.id,
                             ScanImport.scan_type == "cloud_iam")
                     .order_by(ScanImport.id.desc())
                     .first())
        iam_csvs = ((_iam_si.parsed_data or {}).get("csvs") if _iam_si else None) or []
        if iam_csvs:
            from ..services.tracker_iam import append_iam_csv_sheets
            iam_sheets = append_iam_csv_sheets(out_path, iam_csvs)
    except Exception as e:                                  # pragma: no cover
        _log.warning("IAMActionHunter sheet paste skipped for version %s: %s", vid, e)

    try:
        from ..services.risk_register import tidy_unused_rows
        _cleared = tidy_unused_rows(out_path)
        if _cleared:
            _log.info("Cleared %d unused formatted cells in cloud tracker (v%s)", _cleared, vid)
    except Exception as e:                                  # pragma: no cover
        _log.warning("tidy_unused_rows skipped: %s", e)

    db.add(AuditLog(actor_id=user.id, action="tracker.export.cloud",
                    object_type="report_version", object_id=v.id,
                    detail={"va_findings": len(va_dicts), "pt_findings": len(pt_dicts),
                            "service_sheets": len(group_sheets),
                            "iam_sheets": iam_sheets, "out": str(out_path)}))
    db.commit()

    return FileResponse(
        out_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=out_path.name,
    )

def _tester_for_finding(db: Session, f) -> Optional[str]:
    """Per-finding fallback for the DT Tester column when the report
    doesn't have an aggregated `tester_names` list yet — looks up the
    user who first added this finding.
    """
    if not getattr(f, "added_by_id", None):
        return None
    u = db.get(User, f.added_by_id)
    if not u:
        return None
    return u.full_name or u.username


@router.get("/api/admin/tracker-templates/diagnose")
def diagnose_tracker_templates(user: User = Depends(get_current_user)):
    """Diagnostic snapshot of the tracker-template configuration.

    Returns the configured folder, whether it exists, which files
    were found, and how every known ReportTemplate code resolves to
    a file on disk. Designed to make the "why is my export coming
    out as the synthesised flat layout instead of the VibeDocs
    template?" question answerable from the browser — no shell
    access needed.

    Restricted to admin/senior users since the response leaks the
    full server-side filesystem path to the templates directory.
    """
    from ..models import Role
    if user.role not in (Role.admin, Role.senior):
        raise HTTPException(403, "Admin or senior role required")
    return _tpl_picker.diagnose()
