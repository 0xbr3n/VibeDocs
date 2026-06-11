"""
Cloud VA/VAPT Pipeline — groups parsed CloudFindings by AWS/Azure service and
materialises one ReportFinding per service onto the report version.

Each grouped finding carries:
  * title         — "{Service} Misconfigurations"
  * severity/CVSS — highest severity finding in the group
  * description   — summary paragraph + numbered check list
  * remediation   — top-5 unique remediation steps from the group
  * impact        — severity-scaled business impact paragraph
  * affected_asset— comma-joined unique resources (capped at 20)
  * attachments   — per-service XLSX with every individual check row
  * source        — "cloud_pipeline"
  * source_ref    — service name (e.g. "S3", "IAM")

Re-running is idempotent: existing cloud_pipeline findings for the same
service are updated in place (attachment refreshed, content regenerated).
"""
from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter as _gcl
from sqlalchemy.orm import Session
from sqlalchemy.orm.attributes import flag_modified

from ..config import settings
from ..models import FindingStatus, ReportFinding, ReportVersion, Severity, User
from .cloud_parsers import CloudFinding, SEV_RANK, best_severity, group_by_service

logger = logging.getLogger(__name__)

CLOUD_SOURCE      = "cloud_pipeline"
ATTACHMENT_POINTER = (
    "Please refer to the per-service misconfiguration list in the attachment below."
)

_SEV_MAP: dict[str, Severity] = {
    "Critical":      Severity.critical,
    "High":          Severity.high,
    "Medium":        Severity.medium,
    "Low":           Severity.low,
    "Informational": Severity.informational,
}


# ─────────────────────────────────────────────────────────────────────────────
# VibeDocs Cloud VAPT Tracking List — XLSX builder (21-column format)
# Colors matched exactly to the bundled Cloud VAPT tracker template
# ─────────────────────────────────────────────────────────────────────────────

# Row 1 group-label fills
_GRP_FILL    = PatternFill("solid", fgColor="1F497D")   # dark navy (matches template row 1)
_OWN_FILL    = PatternFill("solid", fgColor="FF0000")   # bright red (Ownership group)
# Row 2 column-header fills
_COL_FILL    = PatternFill("solid", fgColor="538DD5")   # medium blue (matches template row 2)
_OWN_COL_FILL = PatternFill("solid", fgColor="FF6969")  # salmon/pink (Ownership cols in row 2)

_ALT_FILL    = PatternFill("solid", fgColor="F2F4F7")
_HDR_FONT    = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
_OWN_COL_FONT = Font(name="Calibri", bold=True, color="000000", size=10)  # black on pink
_DATA_FONT   = Font(name="Calibri", size=9)
_CENTER      = Alignment(horizontal="center", vertical="top", wrap_text=True)
_LEFT        = Alignment(horizontal="left",   vertical="top", wrap_text=True)
_THIN        = Side(style="thin", color="CCCCCC")
_BORDER      = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_CVSS_FILLS: dict[str, PatternFill] = {
    "Critical":      PatternFill("solid", fgColor="C00000"),
    "High":          PatternFill("solid", fgColor="FF0000"),
    "Medium":        PatternFill("solid", fgColor="FFC000"),
    "Low":           PatternFill("solid", fgColor="FFFF00"),
    "Informational": PatternFill("solid", fgColor="92D050"),
}


def _cvss_cell_font(severity: str) -> Font:
    light = severity in ("Critical", "High")
    return Font(name="Calibri", bold=True,
                color="FFFFFF" if light else "000000", size=9)


# 21-column VibeDocs Cloud VAPT Tracking List format (Steps to Reproduce and
# References removed; Ownership now spans Date Raised / DT Tester / Client Owner)
# (header, col_width, center_align)
_VIBEDOCS_COLS: list[tuple[str, int, bool]] = [
    ("S/N",                                   5, True),   # A  1
    ("System",                               20, False),  # B  2
    ("CVSS Risk Rating",                     14, True),   # C  3  ┐ Risk
    ("CVSS Score",                           10, True),   # D  4  │
    ("CVSS Vector",                          22, False),  # E  5  ┘
    ("Affected Resource(s)/Instance(s)",     36, False),  # F  6
    ("Issue Title",                          38, False),  # G  7
    ("Benchmark",                            24, False),  # H  8
    ("Benchmark Clause",                     40, False),  # I  9
    ("Observation",                          50, False),  # J  10
    ("Implication",                          40, False),  # K  11
    ("Recommendation",                       50, False),  # L  12
    ("Management Comments",                  30, False),  # M  13
    ("Date Raised",                          14, True),   # N  14  ┐ Ownership
    ("DT Tester",                            18, False),  # O  15  │
    ("Client Owner",                         18, False),  # P  16  ┘
    ("Status",                               12, True),   # Q  17
    ("Date Follow-Up",                       14, True),   # R  18
    ("DT Tester2",                           18, False),  # S  19
    ("Post Review Observations",             30, False),  # T  20  ┐ Follow-Up
    ("Post Review Screenshot",               22, False),  # U  21  ┘
]

# Row 1 group spans: (start_col, end_col, label, fill)
_GROUP_SPANS = [
    (3,  5,  "Risk",       _GRP_FILL),
    (14, 16, "Ownership",  _OWN_FILL),
    (20, 21, "Follow-Up",  _GRP_FILL),
]

# Columns in row 2 that get the pink Ownership fill (0-based set of 1-based col indices)
_OWN_COLS: frozenset[int] = frozenset({14, 15, 16})


def _benchmark_name(f: CloudFinding) -> str:
    """Return a human-readable CIS benchmark name for the Benchmark column."""
    svc  = (f.service or "").lower()
    comp = (f.compliance or "").lower()
    if "azure" in svc or "azure" in comp:
        return "CIS Microsoft Azure Foundations Benchmark v1.5.0"
    return "CIS AWS Foundations Benchmark v1.4.0"


def _write_vibedocs_header(ws) -> None:
    """Write the 2-row VibeDocs header.

    Row 1 = group span labels (Risk / Ownership / Follow-Up) on dark navy.
    Row 2 = individual column headers on medium blue; Ownership cols on salmon.
    Data starts at row 3; caller must set freeze_panes = 'A3'.
    """
    nc = len(_VIBEDOCS_COLS)

    # Row 1: fill every cell dark navy, then overlay group span labels
    for ci in range(1, nc + 1):
        c = ws.cell(row=1, column=ci)
        c.fill   = _GRP_FILL
        c.border = _BORDER

    for sc, ec, label, fill in _GROUP_SPANS:
        ws.merge_cells(start_row=1, start_column=sc, end_row=1, end_column=ec)
        cell = ws.cell(row=1, column=sc, value=label)
        cell.font      = _HDR_FONT
        cell.fill      = fill
        cell.alignment = _CENTER
        cell.border    = _BORDER

    ws.row_dimensions[1].height = 18

    # Row 2: column headers — medium blue; Ownership cols get salmon fill + black text
    for ci, (hdr, width, _) in enumerate(_VIBEDOCS_COLS, start=1):
        own  = ci in _OWN_COLS
        cell = ws.cell(row=2, column=ci, value=hdr)
        cell.font      = _OWN_COL_FONT if own else _HDR_FONT
        cell.fill      = _OWN_COL_FILL if own else _COL_FILL
        cell.alignment = _CENTER
        cell.border    = _BORDER
        ws.column_dimensions[_gcl(ci)].width = width

    ws.row_dimensions[2].height = 28


def _build_service_xlsx(service: str, findings: list[CloudFinding]) -> bytes:
    """Build a per-service attachment XLSX using the VibeDocs 21-column format."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = service[:31]

    _write_vibedocs_header(ws)

    for ri, f in enumerate(findings, start=3):
        alt  = ri % 2 == 1
        fill = _ALT_FILL if alt else PatternFill()
        values = [
            ri - 2,                  # A: S/N
            "",                      # B: System (blank in attachment)
            f.severity,              # C: CVSS Risk Rating
            f.cvss_score,            # D: CVSS Score
            "",                      # E: CVSS Vector
            f.resource,              # F: Affected Resource(s)/Instance(s)
            f.issue_title,           # G: Issue Title
            _benchmark_name(f),      # H: Benchmark
            f.title,                 # I: Benchmark Clause
            f.description,           # J: Observation
            f.risk,                  # K: Implication
            f.remediation or (
                "Refer to the CIS benchmark and vendor security documentation "
                "for detailed remediation guidance."
            ),                       # L: Recommendation
            "",                      # M: Management Comments
            "",                      # N: Date Raised
            "",                      # O: DT Tester
            "",                      # P: Client Owner
            "Open",                  # Q: Status
            "",                      # R: Date Follow-Up
            "",                      # S: DT Tester2
            "",                      # T: Post Review Observations
            "",                      # U: Post Review Screenshot
        ]
        for ci, (val, (_, _, center)) in enumerate(zip(values, _VIBEDOCS_COLS), start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font      = _DATA_FONT
            cell.fill      = fill
            cell.alignment = _CENTER if center else _LEFT
            cell.border    = _BORDER

        # Color-code CVSS Risk Rating cell (col C = col 3)
        c_cell = ws.cell(row=ri, column=3)
        c_cell.fill = _CVSS_FILLS.get(f.severity, PatternFill())
        c_cell.font = _cvss_cell_font(f.severity)

    ws.freeze_panes = "A3"
    ws.auto_filter.ref = f"A2:{_gcl(len(_VIBEDOCS_COLS))}2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Grouped finding content builders
# ─────────────────────────────────────────────────────────────────────────────

def _build_description(service: str, findings: list[CloudFinding]) -> str:
    sev_counts: dict[str, int] = {}
    for f in findings:
        sev_counts[f.severity] = sev_counts.get(f.severity, 0) + 1

    order = ["Critical", "High", "Medium", "Low", "Informational"]
    summary_parts = [
        f"{sev_counts[s]} {s}"
        for s in order
        if s in sev_counts
    ]
    header = (
        f"During the cloud security assessment, {len(findings)} misconfiguration(s) "
        f"were identified within the {service} service "
        f"({', '.join(summary_parts)}).\n\n"
        "The following individual checks failed:\n"
    )
    lines = [f"{i}. [{f.check_id}] {f.title}" for i, f in enumerate(findings, 1)]
    return header + "\n".join(lines)


def _build_remediation(findings: list[CloudFinding]) -> str:
    seen: list[str] = []
    for f in findings:
        r = (f.remediation or "").strip()
        if r and r not in seen:
            seen.append(r)
    if not seen:
        return (
            "Review the attached per-service misconfiguration list and apply the "
            "recommended remediations for each failed check. Refer to the CIS AWS "
            "Foundations Benchmark or the vendor security documentation for detailed "
            "step-by-step guidance."
        )
    top = seen[:5]
    lines = [f"{i}. {r}" for i, r in enumerate(top, 1)]
    if len(seen) > 5:
        lines.append(f"\n(+{len(seen) - 5} additional remediation(s) — see attached XLSX.)")
    return "\n".join(lines)


def _build_impact(service: str, findings: list[CloudFinding]) -> str:
    worst_sev, _ = best_severity(findings)
    return {
        "Critical": (
            f"Critical misconfigurations in the {service} service could allow "
            "unauthenticated or low-privileged attackers to access, modify, or destroy "
            "sensitive data and cloud resources, potentially leading to a full "
            "compromise of the environment."
        ),
        "High": (
            f"High-severity misconfigurations in {service} may expose sensitive data to "
            "unauthorised parties or enable privilege escalation within the cloud "
            "environment."
        ),
        "Medium": (
            f"Medium-severity misconfigurations in {service} increase the overall attack "
            "surface and may facilitate lateral movement or data exfiltration when "
            "combined with other weaknesses."
        ),
        "Low": (
            f"Low-severity misconfigurations in {service} represent deviations from "
            "security best practices that marginally increase risk and may be leveraged "
            "as part of a multi-step attack chain."
        ),
        "Informational": (
            f"Informational observations in {service} indicate configurations that "
            "deviate from best practices but do not present an immediate exploitable risk."
        ),
    }.get(worst_sev, "")


# ─────────────────────────────────────────────────────────────────────────────
# Public entry point
# ─────────────────────────────────────────────────────────────────────────────

def run_cloud_pipeline(
    db: Session,
    rv: ReportVersion,
    user: User,
    all_findings: list[CloudFinding],
) -> dict:
    """Group cloud findings by service and upsert one ReportFinding per service.

    Returns:
        {
          "ok": True,
          "total_findings": int,
          "services": [
              {
                "service": str,
                "count": int,
                "severity": str,
                "cvss_score": float,
                "finding_id": int,
              }
          ],
          "groups_created": int,
          "groups_updated": int,
        }
    """
    if not all_findings:
        return {
            "ok": True,
            "total_findings": 0,
            "services": [],
            "groups_created": 0,
            "groups_updated": 0,
        }

    groups    = group_by_service(all_findings)
    attach_dir = (
        Path(settings.UPLOAD_DIR)
        / "cloud_pipeline"
        / str(rv.report_id)
        / rv.version
    )
    attach_dir.mkdir(parents=True, exist_ok=True)

    results   = []
    created   = 0
    updated   = 0

    for service, svc_findings in groups.items():
        sev_str, cvss = best_severity(svc_findings)
        title         = f"{service} Misconfigurations"

        resources = list(dict.fromkeys(
            f.resource for f in svc_findings if f.resource
        ))[:20]
        affected = ", ".join(resources) if resources else ATTACHMENT_POINTER

        description = _build_description(service, svc_findings)
        remediation = _build_remediation(svc_findings)
        impact      = _build_impact(service, svc_findings)

        # Build per-service XLSX attachment
        xlsx_bytes    = _build_service_xlsx(service, svc_findings)
        safe_svc      = service.replace(" ", "_").replace("/", "_")
        xlsx_filename = f"cloud_{safe_svc.lower()}_misconfigs.xlsx"
        xlsx_path     = attach_dir / f"{uuid.uuid4().hex}__{xlsx_filename}"
        xlsx_path.write_bytes(xlsx_bytes)

        attachment = {
            "filename":    xlsx_filename,
            "path":        str(xlsx_path),
            "kind":        "xlsx",
            "label":       (
                f"{service} misconfiguration list "
                f"({len(svc_findings)} individual check(s))"
            ),
            "uploaded_at": datetime.utcnow().isoformat(),
            "uploaded_by": user.username,
            "key":         f"cloud_{safe_svc.lower()}",
        }

        existing: ReportFinding | None = (
            db.query(ReportFinding)
              .filter(
                  ReportFinding.report_version_id == rv.id,
                  ReportFinding.source == CLOUD_SOURCE,
                  ReportFinding.source_ref == service,
              )
              .first()
        )

        if existing:
            att_key  = attachment["key"]
            old_atts = [
                a for a in (existing.attachments or [])
                if not (isinstance(a, dict) and a.get("key") == att_key)
            ]
            existing.attachments  = old_atts + [attachment]
            existing.title        = title
            existing.description  = description
            existing.impact       = impact
            existing.remediation  = remediation
            existing.affected_asset = affected
            existing.severity     = _SEV_MAP.get(sev_str, Severity.medium)
            existing.cvss_score   = cvss
            flag_modified(existing, "attachments")
            updated    += 1
            finding_id  = existing.id
        else:
            rf = ReportFinding(
                report_version_id = rv.id,
                title             = title,
                description       = description,
                impact            = impact,
                remediation       = remediation,
                affected_asset    = affected,
                severity          = _SEV_MAP.get(sev_str, Severity.medium),
                cvss_score        = cvss,
                status            = FindingStatus.open,
                added_by_id       = user.id,
                source            = CLOUD_SOURCE,
                source_ref        = service,
                attachments       = [attachment],
            )
            db.add(rf)
            db.flush()
            created    += 1
            finding_id  = rf.id

        results.append({
            "service":    service,
            "count":      len(svc_findings),
            "severity":   sev_str,
            "cvss_score": cvss,
            "finding_id": finding_id,
        })

    db.commit()
    logger.info(
        "Cloud pipeline vid=%s: %d findings → %d services (%d created, %d updated)",
        rv.id, len(all_findings), len(groups), created, updated,
    )
    return {
        "ok":             True,
        "total_findings": len(all_findings),
        "services":       results,
        "groups_created": created,
        "groups_updated": updated,
    }
