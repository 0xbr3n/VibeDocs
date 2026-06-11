"""
Nessus Policy Compliance / CIS Host Configuration Review → Excel workbook.

This is a server-side port of the standalone
`nessus_compliance_to_excel.py` consultant CLI tool. It strips out the
argparse / stdout side of the CLI and exposes two clean entry points:

  * ``parse_uploads(uploads)`` — accepts an iterable of
    ``(filename, raw_bytes)`` tuples (typically straight from FastAPI's
    ``UploadFile.file.read()``) and returns the list of
    ``ComplianceRow`` records extracted from every input.
  * ``build_workbook_bytes(rows)`` — converts a row list into the
    styled .xlsx bytes the user downloads.

The XML parsing + workbook layout logic is intentionally copied
verbatim from the user's reference script so the Excel output looks
identical regardless of whether the consultant runs the CLI locally or
uploads the .nessus via the in-app toolkit page.

Why a separate module under ``services/tools/`` rather than
``routers/``: the routing layer should stay thin; the underlying
domain logic gets its own home so we can (a) unit-test it in
isolation, (b) reuse the parser later (e.g. a report-version
ingestion path that pulls compliance rows straight into a findings
table). The module has no FastAPI dependency.
"""
from __future__ import annotations

import io
import re
from collections import Counter, defaultdict
from dataclasses import dataclass, asdict
from datetime import datetime
from typing import Dict, Iterable, List, Optional, Tuple
import defusedxml.ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


SUPPORTED_EXTENSIONS = {".nessus", ".xml"}
RESULT_ORDER = ["FAILED", "PASSED", "WARNING", "ERROR", "INFO", "UNKNOWN"]


@dataclass
class ComplianceRow:
    source_file: str
    host_name: str
    ip_address: str
    fqdn: str
    operating_system: str
    policy_name: str
    plugin_id: str
    plugin_name: str
    benchmark_level: str
    benchmark_profile: str
    policy_setting: str
    description_of_requirement: str
    solution: str
    result: str
    system_value_or_error: str
    compliance_requirement: str
    severity: str
    service_name: str
    port: str
    protocol: str
    see_also: str
    synopsis: str
    description: str


# ---------- XML helpers ----------

def _strip_tag(tag: str) -> str:
    if not tag:
        return ""
    if "}" in tag:
        tag = tag.split("}", 1)[1]
    return tag


def _child_text(elem: Optional[ET.Element], name: str, default: str = "") -> str:
    if elem is None:
        return default
    for child in list(elem):
        if _strip_tag(child.tag) == name:
            return _normalize_text(child.text or "")
    return default


def _all_children(elem: Optional[ET.Element], name: str) -> List[ET.Element]:
    if elem is None:
        return []
    return [child for child in list(elem) if _strip_tag(child.tag) == name]


def _normalize_text(value: Optional[str]) -> str:
    if value is None:
        return ""
    value = value.replace("\r", " ").replace("\n", " ")
    value = re.sub(r"\s+", " ", value)
    return value.strip()


def _parse_host_properties(report_host: ET.Element) -> Dict[str, str]:
    props: Dict[str, str] = {}
    host_properties = None
    for child in list(report_host):
        if _strip_tag(child.tag) == "HostProperties":
            host_properties = child
            break

    if host_properties is None:
        return props

    for tag in _all_children(host_properties, "tag"):
        key = tag.attrib.get("name", "")
        props[key] = _normalize_text(tag.text or "")
    return props


# ---------- compliance extraction ----------

def _is_compliance_item(report_item: ET.Element) -> bool:
    plugin_family = report_item.attrib.get("pluginFamily", "")
    if plugin_family == "Policy Compliance":
        return True
    names = {_strip_tag(c.tag) for c in list(report_item)}
    return any(
        name in names
        for name in {
            "compliance-result",
            "compliance-check-name",
            "compliance-info",
            "compliance-solution",
            "compliance-actual-value",
            "compliance-policy-value",
            "cm_compliance-result",
            "cm_compliance-check-name",
            "cm_compliance-info",
            "cm_compliance-solution",
            "cm_compliance-actual-value",
            "cm_compliance-policy-value",
        }
    )


def _extract_benchmark_level(*texts: str) -> str:
    haystack = " ".join(t for t in texts if t).upper()
    if re.search(r"\bL1\b", haystack) or "LEVEL 1" in haystack:
        return "L1"
    if re.search(r"\bL2\b", haystack) or "LEVEL 2" in haystack:
        return "L2"
    return "Unknown"


def _extract_benchmark_profile(
    plugin_name: str, check_name: str, info: str, description: str
) -> str:
    haystack = " | ".join([plugin_name, check_name, info, description])
    m = re.search(r"(CIS[^|]{0,140}?Benchmark)", haystack, flags=re.IGNORECASE)
    if m:
        return _normalize_text(m.group(1))
    return _normalize_text(plugin_name) or "Policy Compliance"


def _split_policy_setting(check_name: str) -> str:
    parts = [p.strip() for p in check_name.split(" - ") if p.strip()]
    if len(parts) >= 2:
        return parts[-1]
    parts = [p.strip() for p in check_name.split(":") if p.strip()]
    if len(parts) >= 2:
        return ": ".join(parts[1:])
    return check_name


def _parse_report_item(
    source_file: str,
    host_name: str,
    host_props: Dict[str, str],
    report_item: ET.Element,
) -> ComplianceRow:
    plugin_name = report_item.attrib.get("pluginName", "")
    plugin_id = report_item.attrib.get("pluginID", "")
    severity = report_item.attrib.get("severity", "")
    protocol = report_item.attrib.get("protocol", "")
    port = report_item.attrib.get("port", "")
    service_name = report_item.attrib.get("svc_name", "")

    check_name = (_child_text(report_item, "cm_compliance-check-name")
                  or _child_text(report_item, "compliance-check-name"))
    info = (_child_text(report_item, "cm_compliance-info")
            or _child_text(report_item, "compliance-info"))
    solution = (_child_text(report_item, "cm_compliance-solution")
                or _child_text(report_item, "compliance-solution"))
    result = (_child_text(report_item, "cm_compliance-result")
              or _child_text(report_item, "compliance-result") or "UNKNOWN")
    actual_value = (_child_text(report_item, "cm_compliance-actual-value")
                    or _child_text(report_item, "compliance-actual-value"))
    policy_value = (_child_text(report_item, "cm_compliance-policy-value")
                    or _child_text(report_item, "compliance-policy-value"))
    see_also = _child_text(report_item, "see_also")
    synopsis = _child_text(report_item, "synopsis")
    description = _child_text(report_item, "description")

    benchmark_level = _extract_benchmark_level(plugin_name, check_name, info, description)
    benchmark_profile = _extract_benchmark_profile(plugin_name, check_name, info, description)

    return ComplianceRow(
        source_file=source_file,
        host_name=host_name,
        ip_address=host_props.get("host-ip", ""),
        fqdn=host_props.get("host-fqdn", ""),
        operating_system=host_props.get("operating-system", ""),
        policy_name=plugin_name or "Policy Compliance",
        plugin_id=plugin_id,
        plugin_name=plugin_name,
        benchmark_level=benchmark_level,
        benchmark_profile=benchmark_profile,
        policy_setting=_split_policy_setting(check_name) if check_name else "",
        description_of_requirement=info,
        solution=solution,
        result=result.upper(),
        system_value_or_error=actual_value,
        compliance_requirement=policy_value,
        severity=severity,
        service_name=service_name,
        port=port,
        protocol=protocol,
        see_also=see_also,
        synopsis=synopsis,
        description=description,
    )


# ---------- public parsing entry points ----------

def _load_xml_root(raw: str) -> ET.Element:
    # Nessus compliance data may contain cm: tags without a declared
    # namespace. The legacy Perl parser tolerated this, but ElementTree
    # does not — strip the prefix.
    raw = raw.replace("<cm:", "<cm_").replace("</cm:", "</cm_")
    try:
        return ET.fromstring(raw)
    except ET.ParseError as exc:
        raise ValueError(f"XML parse error: {exc}") from exc


def parse_nessus_bytes(source_filename: str, raw_bytes: bytes) -> List[ComplianceRow]:
    """Parse one .nessus file (provided as raw bytes) into a list of
    ``ComplianceRow``. Returns an empty list if the file has no
    Policy Compliance items; raises ``ValueError`` if the file isn't a
    valid NessusClientData_v2 document.
    """
    raw = raw_bytes.decode("utf-8", errors="replace")
    root = _load_xml_root(raw)
    if _strip_tag(root.tag) != "NessusClientData_v2":
        raise ValueError(
            f"{source_filename} is not a NessusClientData_v2 file."
        )

    report = None
    for child in list(root):
        if _strip_tag(child.tag) == "Report":
            report = child
            break
    if report is None:
        return []

    rows: List[ComplianceRow] = []
    for report_host in _all_children(report, "ReportHost"):
        host_name = report_host.attrib.get("name", "")
        host_props = _parse_host_properties(report_host)
        for report_item in _all_children(report_host, "ReportItem"):
            if not _is_compliance_item(report_item):
                continue
            rows.append(_parse_report_item(source_filename, host_name, host_props, report_item))
    return rows


def parse_uploads(uploads: Iterable[Tuple[str, bytes]]) -> List[ComplianceRow]:
    """Parse one or many uploaded .nessus / .xml files into a flat
    list of ``ComplianceRow``. Each tuple is ``(filename, raw_bytes)``
    — the filename only feeds the ``source_file`` column on each row.

    Files with the wrong extension are rejected (the caller surfaces
    the error to the user). Files that turn out to be non-compliance
    Nessus scans are accepted but contribute zero rows.
    """
    all_rows: List[ComplianceRow] = []
    for filename, data in uploads:
        ext = "." + filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
        if ext not in SUPPORTED_EXTENSIONS:
            raise ValueError(
                f"{filename}: unsupported file type. Upload .nessus or .xml only."
            )
        rows = parse_nessus_bytes(filename, data)
        all_rows.extend(rows)
    return all_rows


# ---------- Excel writing ----------

def _safe_sheet_name(name: str, used: set[str]) -> str:
    cleaned = re.sub(r"[\\/*?:\[\]]", "", name).strip() or "Sheet"
    cleaned = cleaned[:31]
    original = cleaned
    counter = 2
    while cleaned in used:
        suffix = f"_{counter}"
        cleaned = f"{original[:31-len(suffix)]}{suffix}"
        counter += 1
    used.add(cleaned)
    return cleaned


def _write_dataframe_like(ws, rows: List[Dict[str, str]]) -> None:
    if not rows:
        ws.append(["No data"])
        return

    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        excel_row = []
        for h in headers:
            value = row.get(h, "")
            # Preserve leading '=' so Excel doesn't treat the cell as a formula.
            if isinstance(value, str) and value.startswith("="):
                value = "'" + value
            excel_row.append(value)
        ws.append(excel_row)

    _style_worksheet(ws, len(headers))
    _add_table(ws, len(rows) + 1, len(headers))


def _style_worksheet(ws, num_cols: int) -> None:
    header_fill = PatternFill("solid", fgColor="000000")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="000000")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{ws.max_row}"

    # Widths tuned for review usability — same map as the CLI tool.
    width_map = {
        "A": 18, "B": 18, "C": 15, "D": 28, "E": 28, "F": 14,
        "G": 40, "H": 14, "I": 25, "J": 60, "K": 60, "L": 18,
        "M": 24, "N": 22, "O": 40, "P": 22, "Q": 10, "R": 16,
        "S": 10, "T": 10, "U": 30, "V": 30, "W": 60,
    }
    for idx in range(1, num_cols + 1):
        col = get_column_letter(idx)
        ws.column_dimensions[col].width = width_map.get(col, 20)


def _add_table(ws, max_row: int, max_col: int) -> None:
    if max_row < 2:
        return
    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(
        displayName=f"Table_{re.sub(r'[^A-Za-z0-9]', '_', ws.title)}",
        ref=ref,
    )
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)


def _build_summary_rows(rows: List[ComplianceRow]) -> List[Dict[str, str]]:
    result_counter = Counter(row.result for row in rows)
    level_counter = Counter(row.benchmark_level for row in rows)
    policy_counter = Counter(row.policy_name for row in rows)
    host_counter = Counter(row.ip_address or row.host_name for row in rows)

    ordered_results = RESULT_ORDER + sorted(set(result_counter) - set(RESULT_ORDER))

    output: List[Dict[str, str]] = []
    output.append({"Metric": "Generated At",
                   "Value": datetime.now().strftime("%Y-%m-%d %H:%M:%S")})
    output.append({"Metric": "Total Compliance Rows", "Value": str(len(rows))})
    output.append({"Metric": "Unique Hosts", "Value": str(len(host_counter))})
    output.append({"Metric": "Unique Policies", "Value": str(len(policy_counter))})
    for result in ordered_results:
        if result in result_counter:
            output.append({"Metric": f"Result - {result}",
                           "Value": str(result_counter[result])})
    for level in sorted(level_counter):
        output.append({"Metric": f"Benchmark Level - {level}",
                       "Value": str(level_counter[level])})
    return output


def _build_host_summary_rows(rows: List[ComplianceRow]) -> List[Dict[str, str]]:
    grouped: Dict[str, Counter] = defaultdict(Counter)
    meta: Dict[str, Dict[str, str]] = {}
    for row in rows:
        host_key = row.ip_address or row.host_name
        grouped[host_key][row.result] += 1
        if host_key not in meta:
            meta[host_key] = {
                "Host": row.host_name,
                "IP Address": row.ip_address,
                "FQDN": row.fqdn,
                "Operating System": row.operating_system,
            }

    output: List[Dict[str, str]] = []
    for host_key in sorted(grouped):
        counts = grouped[host_key]
        record = dict(meta[host_key])
        for label in ("FAILED", "PASSED", "WARNING", "ERROR", "INFO", "UNKNOWN"):
            record[label] = str(counts.get(label, 0))
        record["Total"] = str(sum(counts.values()))
        output.append(record)
    return output


def _rows_to_dicts(rows: Iterable[ComplianceRow]) -> List[Dict[str, str]]:
    return [asdict(row) for row in rows]


def build_workbook_bytes(rows: List[ComplianceRow]) -> bytes:
    """Materialise the workbook in memory and return its bytes. The
    caller streams these to the user as a download — we never touch
    disk, so concurrent invocations from different consultants don't
    contend on a temp file."""
    wb = Workbook()
    wb.remove(wb.active)
    used_sheet_names: set[str] = set()

    summary_ws = wb.create_sheet(_safe_sheet_name("Summary", used_sheet_names))
    _write_dataframe_like(summary_ws, _build_summary_rows(rows))

    all_ws = wb.create_sheet(_safe_sheet_name("All Compliance", used_sheet_names))
    _write_dataframe_like(all_ws, _rows_to_dicts(rows))

    host_ws = wb.create_sheet(_safe_sheet_name("Host Summary", used_sheet_names))
    _write_dataframe_like(host_ws, _build_host_summary_rows(rows))

    by_policy: Dict[str, List[ComplianceRow]] = defaultdict(list)
    for row in rows:
        by_policy[row.policy_name or "Policy Compliance"].append(row)

    for policy_name in sorted(by_policy):
        ws = wb.create_sheet(_safe_sheet_name(policy_name, used_sheet_names))
        _write_dataframe_like(ws, _rows_to_dicts(by_policy[policy_name]))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def summary_stats(rows: List[ComplianceRow]) -> dict:
    """Lightweight summary for the UI's "preview" pane after upload —
    the consultant sees the counts before clicking Download."""
    result_counter = Counter(row.result for row in rows)
    level_counter = Counter(row.benchmark_level for row in rows)
    host_counter = Counter(row.ip_address or row.host_name for row in rows)
    policy_counter = Counter(row.policy_name for row in rows)
    return {
        "total_rows":     len(rows),
        "unique_hosts":   len(host_counter),
        "unique_policies": len(policy_counter),
        "results":  {k: result_counter[k] for k in result_counter},
        "levels":   {k: level_counter[k] for k in level_counter},
        "top_hosts": host_counter.most_common(5),
    }
