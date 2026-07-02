"""
Scope auto-population for Infrastructure VAPT / VA projects.

Two input shapes are accepted:

  * **Nessus CSV** — the export consultants get from Tenable. The
    `Host` column (sometimes named `IP Address` / `Asset Name`) carries
    one target per row.

  * **VibeDocs VAPT Excel tracker** — open the workbook, scan every
    sheet for IPv4 / IPv6 / CIDR / hostname-looking cells and dedupe.
    This intentionally over-matches: it's better to surface a
    near-target the consultant can prune than to silently miss one.

Both paths produce the same return shape:

    {
        "targets":     [ "10.1.2.3", "10.1.2.0/24", "host.example.com" ],
        "host_count":  N,
        "source":      "nessus_csv" | "excel_tracker",
        "warnings":    [ ... ],
    }

The list is order-preserving deduped so consultants see targets in the
order they appear in the source, which usually matches the engagement's
own ordering.
"""
from __future__ import annotations
import csv
import ipaddress
import io
import logging
import re
from pathlib import Path
from typing import Iterable

log = logging.getLogger(__name__)


# Hostname-ish strings: letters/digits/-/. with at least one dot and a
# multi-char TLD. Deliberately stricter than RFC-1123 — we'd rather skip a
# weird short token than pollute the scope with cell labels like "Pass".
_HOSTNAME_RE = re.compile(
    r"\b(?=[A-Za-z0-9.-]{4,253}\b)[A-Za-z0-9]"
    r"(?:[A-Za-z0-9-]*[A-Za-z0-9])?"
    r"(?:\.[A-Za-z0-9][A-Za-z0-9-]{0,62})+"
    r"\.[A-Za-z]{2,24}\b"
)

# CIDR e.g. 10.0.0.0/24, 2001:db8::/32
_CIDR_RE = re.compile(r"\b(?:\d{1,3}(?:\.\d{1,3}){3}/\d{1,2}|[0-9a-fA-F:]+/\d{1,3})\b")

# Range e.g. 10.0.0.5-10.0.0.20 or 10.0.0.5-20
_RANGE_RE = re.compile(
    r"\b(\d{1,3}(?:\.\d{1,3}){3})\s*[-–]\s*(\d{1,3}(?:\.\d{1,3}){0,3})\b"
)

# Plain IPv4 / IPv6 — we run ipaddress.ip_address() to validate before keeping.
_IPV4_RE = re.compile(r"\b\d{1,3}(?:\.\d{1,3}){3}\b")
_IPV6_RE = re.compile(r"\b(?:[0-9a-fA-F]{1,4}:){2,7}[0-9a-fA-F]{1,4}\b")


def _classify_token(tok: str) -> str | None:
    """Return a normalised representation if `tok` looks like a target, else None."""
    tok = tok.strip().rstrip(",;")
    if not tok:
        return None
    # CIDR
    m = _CIDR_RE.fullmatch(tok)
    if m:
        try:
            return str(ipaddress.ip_network(tok, strict=False))
        except ValueError:
            return None
    # Range a.b.c.d-e.f.g.h or a.b.c.d-h
    m = _RANGE_RE.fullmatch(tok)
    if m:
        start, end = m.group(1), m.group(2)
        if "." not in end:
            # last-octet shorthand
            base = start.rsplit(".", 1)[0]
            end = f"{base}.{end}"
        try:
            ipaddress.ip_address(start); ipaddress.ip_address(end)
            return f"{start}-{end}"
        except ValueError:
            return None
    # Plain IPv4
    if _IPV4_RE.fullmatch(tok):
        try:
            return str(ipaddress.ip_address(tok))
        except ValueError:
            return None
    # IPv6
    if _IPV6_RE.fullmatch(tok):
        try:
            return str(ipaddress.ip_address(tok))
        except ValueError:
            return None
    # Hostname
    if _HOSTNAME_RE.fullmatch(tok):
        return tok.lower()
    return None


def _dedupe_preserve_order(items: Iterable[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for x in items:
        if not x: continue
        if x in seen: continue
        seen.add(x); out.append(x)
    return out


# ============================================================
# Nessus CSV
# ============================================================

# Header candidates that hold the target identifier in a Nessus CSV export.
_NESSUS_HOST_HEADERS = ["host", "ip address", "ipaddress", "ip", "asset",
                         "asset name", "target", "name"]


def parse_nessus_csv(path: Path) -> dict:
    targets: list[str] = []
    warnings: list[str] = []

    # Nessus exports may be UTF-8-BOM. Use 'utf-8-sig' to strip the marker.
    with path.open("r", encoding="utf-8-sig", errors="replace", newline="") as f:
        # Some Nessus exports are tab-separated and named .csv. Use the Sniffer
        # as a fallback to detect.
        sample = f.read(4096); f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        header = next(reader, None)
        if not header:
            return {"targets": [], "host_count": 0, "source": "nessus_csv",
                    "warnings": ["empty CSV"]}
        header_l = [h.strip().lower() for h in header]
        host_col = None
        for cand in _NESSUS_HOST_HEADERS:
            if cand in header_l:
                host_col = header_l.index(cand); break
        if host_col is None:
            # Fallback: look at the first row and pick the first column
            # whose value classifies as a target.
            for row in reader:
                for idx, cell in enumerate(row):
                    if _classify_token(cell or ""):
                        host_col = idx; break
                if host_col is not None: break
            f.seek(0); next(reader, None)   # re-skip header
        if host_col is None:
            return {"targets": [], "host_count": 0, "source": "nessus_csv",
                    "warnings": ["could not locate a host column"]}
        for row in reader:
            if host_col >= len(row): continue
            val = (row[host_col] or "").strip()
            if not val: continue
            norm = _classify_token(val)
            if norm: targets.append(norm)
            else:    warnings.append(f"skipped value: {val}")

    deduped = _dedupe_preserve_order(targets)
    return {
        "targets": deduped,
        "host_count": len(deduped),
        "source": "nessus_csv",
        "warnings": warnings[:50],
    }


# ============================================================
# Excel tracker
# ============================================================

def parse_excel_tracker(path: Path) -> dict:
    """Walk every sheet in the workbook and harvest target-like strings."""
    from openpyxl import load_workbook
    wb = load_workbook(str(path), data_only=True, read_only=True)
    found: list[str] = []
    warnings: list[str] = []
    try:
        for ws in wb.worksheets:
            for row in ws.iter_rows(values_only=True):
                for cell in row:
                    if cell is None: continue
                    s = str(cell)
                    # Heuristic: cells often hold a single target,
                    # sometimes comma-separated lists. Tokenise on
                    # whitespace / commas / semicolons but also try the
                    # whole-cell match first so CIDRs and ranges survive.
                    whole = _classify_token(s)
                    if whole:
                        found.append(whole); continue
                    for tok in re.split(r"[,;\s]+", s):
                        norm = _classify_token(tok)
                        if norm: found.append(norm)
    finally:
        wb.close()
    deduped = _dedupe_preserve_order(found)
    return {
        "targets": deduped,
        "host_count": len(deduped),
        "source": "excel_tracker",
        "warnings": warnings[:50],
    }


def parse_any(path: Path) -> dict:
    """Dispatch on extension."""
    ext = path.suffix.lower()
    if ext == ".csv":
        return parse_nessus_csv(path)
    if ext in (".xlsx", ".xls", ".xlsm"):
        return parse_excel_tracker(path)
    raise ValueError(f"Unsupported file type: {ext}. Expected .csv / .xlsx / .xls / .xlsm")
